import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()

# ── Styles ──────────────────────────────────────────────────────────────────
header_font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
credit_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
level_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
criteria_font = Font(name="Calibri", bold=True, size=11, color="1F4E28")
question_font = Font(name="Calibri", size=10)
data_flag_font = Font(name="Calibri", size=10, italic=True, color="2E75B6")
condition_font = Font(name="Calibri", bold=True, size=10, color="7030A0")
wrap = Alignment(wrap_text=True, vertical="top")
center_wrap = Alignment(wrap_text=True, vertical="center", horizontal="center")

green_fill = PatternFill(start_color="1F4E28", end_color="1F4E28", fill_type="solid")
dark_green_fill = PatternFill(start_color="0D3318", end_color="0D3318", fill_type="solid")
level_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
criteria_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
question_fill = PatternFill(start_color="F1F8E9", end_color="F1F8E9", fill_type="solid")
condition_fill = PatternFill(start_color="EDE7F6", end_color="EDE7F6", fill_type="solid")
data_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

COL_WIDTHS = {"A": 8, "B": 20, "C": 22, "D": 28, "E": 16, "F": 55, "G": 50, "H": 40}

# Reusable Yes/No dropdown
yn_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
yn_dv.error = "Please select Yes or No"
yn_dv.errorTitle = "Invalid entry"
yn_dv.prompt = "Select Yes or No"
yn_dv.promptTitle = "Yes / No"

# Track which rows need the dropdown per sheet
yn_cells = {}  # sheet title -> list of cell refs


def setup_sheet(ws, title):
    ws.title = title
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    headers = [
        "Ref", "Credit", "Performance Level", "Criteria",
        "Question Type", "Question", "Response",
        "Data Collection / Research Notes",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = dark_green_fill
        cell.alignment = center_wrap
        cell.border = thin_border
    ws.row_dimensions[1].height = 35
    ws.freeze_panes = "A2"
    yn_cells[title] = []
    return 2


def add_credit_header(ws, row, credit_name, outcome):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = ws.cell(row=row, column=1, value=f"{credit_name} — {outcome}")
    cell.font = credit_font
    cell.fill = green_fill
    cell.alignment = wrap
    cell.border = thin_border
    ws.row_dimensions[row].height = 30
    return row + 1


def add_level_header(ws, row, level_name):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = ws.cell(row=row, column=1, value=level_name)
    cell.font = level_font
    cell.fill = level_fill
    cell.alignment = wrap
    cell.border = thin_border
    ws.row_dimensions[row].height = 22
    return row + 1


def add_criteria_header(ws, row, criteria_name):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = ws.cell(row=row, column=1, value=criteria_name)
    cell.font = criteria_font
    cell.fill = criteria_fill
    cell.alignment = wrap
    cell.border = thin_border
    ws.row_dimensions[row].height = 20
    return row + 1


def add_question(ws, row, ref, credit, level, criteria, q_type, question, data_note=""):
    is_yn = q_type == "Condition (Y/N)"
    values = [ref, credit, level, criteria, q_type, question, "", data_note]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.alignment = wrap
        cell.border = thin_border
        if is_yn:
            cell.fill = condition_fill
            cell.font = condition_font if col == 5 else question_font
        elif col == 8 and data_note:
            cell.fill = data_fill
            cell.font = data_flag_font
        else:
            cell.fill = question_fill if col != 7 else white_fill
            cell.font = question_font
    ws.row_dimensions[row].height = 45
    if is_yn:
        yn_cells[ws.title].append(f"G{row}")
    return row + 1


def apply_dropdowns(ws):
    """Add a single Yes/No data validation to the sheet covering all Y/N cells."""
    cells = yn_cells.get(ws.title, [])
    if not cells:
        return
    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    dv.error = "Please select Yes or No"
    dv.errorTitle = "Invalid entry"
    dv.prompt = "Select Yes or No"
    dv.promptTitle = "Yes / No"
    for c in cells:
        dv.add(c)
    ws.add_data_validation(dv)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 1: INDUSTRY DEVELOPMENT
# ════════════════════════════════════════════════════════════════════════════
ws = wb.active
row = setup_sheet(ws, "Industry Development")
row = add_credit_header(ws, row, "Industry Development",
    "The development facilitates industry transformation through partnership, collaboration and data sharing.")

row = add_level_header(ws, row, "Credit Achievement (1 point)")

# ── Green Star Accredited Professional ──
row = add_criteria_header(ws, row, "Green Star Accredited Professional")
ref_base = "ID"
q = 1

row = add_question(ws, row, f"{ref_base}.{q}", "Industry Development", "Credit Achievement",
    "Green Star Accredited Professional", "Descriptive",
    "Identify the GSAP(s) engaged, including name, organisation, accreditation number, and Green Star Buildings accreditation held.",
    "GSAP workforce capacity and distribution across projects.")
q += 1

row = add_question(ws, row, f"{ref_base}.{q}", "Industry Development", "Credit Achievement",
    "Green Star Accredited Professional", "Descriptive",
    "State the date and project phase when the GSAP was first engaged.",
    "Timing of sustainability expertise integration relative to design stage.")
q += 1

row = add_question(ws, row, f"{ref_base}.{q}", "Industry Development", "Credit Achievement",
    "Green Star Accredited Professional", "Condition (Y/N)",
    "Was the GSAP engaged within one month of project registration?",
    "")
q += 1

row = add_question(ws, row, f"{ref_base}.{q}", "Industry Development", "Credit Achievement",
    "Green Star Accredited Professional", "Descriptive",
    "Summarise the GSAP's scope of advisory and coordination activities on Green Star strategy, process and certification.",
    "Depth of sustainability advisory services on projects.")
q += 1

row = add_question(ws, row, f"{ref_base}.{q}", "Industry Development", "Credit Achievement",
    "Green Star Accredited Professional", "Condition (Y/N)",
    "Was the GSAP role fulfilled by more than one individual or organisation?",
    "Continuity of sustainability expertise across project lifecycle.")
q += 1

row = add_question(ws, row, f"{ref_base}.{q}", "Industry Development", "Credit Achievement",
    "Green Star Accredited Professional", "Descriptive",
    "If multiple GSAPs, explain transitions and confirm each held valid Green Star Buildings accreditation throughout their engagement.",
    "")
q += 1

row = add_question(ws, row, f"{ref_base}.{q}", "Industry Development", "Credit Achievement",
    "Green Star Accredited Professional", "Condition (Y/N)",
    "Is the GSAP nominated as the Project Contact for GBCA communications?",
    "")
q += 1

row = add_question(ws, row, f"{ref_base}.{q}", "Industry Development", "Credit Achievement",
    "Green Star Accredited Professional", "Descriptive",
    "Describe how ongoing GSAP involvement was maintained throughout the project (e.g. design meetings, workshops).",
    "")
q += 1

# ── Financial Transparency ──
row = add_criteria_header(ws, row, "Financial Transparency")

row = add_question(ws, row, f"{ref_base}.{q}", "Industry Development", "Credit Achievement",
    "Financial Transparency", "Condition (Y/N)",
    "Was the Financial Transparency template completed in its latest version and submitted in Excel format?",
    "Industry-wide benchmarking of sustainable building costs.")
q += 1

row = add_question(ws, row, f"{ref_base}.{q}", "Industry Development", "Credit Achievement",
    "Financial Transparency", "Descriptive",
    "Identify who prepared the cost data (e.g. quantity surveyor, head contractor, cost consultant).",
    "")
q += 1

row = add_question(ws, row, f"{ref_base}.{q}", "Industry Development", "Credit Achievement",
    "Financial Transparency", "Descriptive",
    "Explain how documentation and implementation costs for sustainable practices were isolated from the base (non-Green Star) requirement.",
    "Cost premiums/savings of green building practices.")
q += 1

row = add_question(ws, row, f"{ref_base}.{q}", "Industry Development", "Credit Achievement",
    "Financial Transparency", "Data",
    "Provide total project construction cost and total additional cost for sustainable practices (documentation + implementation).",
    "Cost-benefit analysis of green certification across the industry.")
q += 1

# ── Marketing Sustainability Achievements ──
row = add_criteria_header(ws, row, "Marketing Sustainability Achievements")

row = add_question(ws, row, f"{ref_base}.{q}", "Industry Development", "Credit Achievement",
    "Marketing Sustainability Achievements", "Descriptive",
    "List which three or more marketing activities were undertaken: (a) case study to GBCA, (b) digital screens, (c) construction hoarding, (d) marketing/communications strategy.",
    "Industry adoption of sustainability marketing practices.")
q += 1

row = add_question(ws, row, f"{ref_base}.{q}", "Industry Development", "Credit Achievement",
    "Marketing Sustainability Achievements", "Descriptive",
    "Describe how sustainability achievements are communicated to building users, the public, or prospective tenants/buyers.",
    "Effectiveness and reach of green building awareness campaigns.")
q += 1

row = add_question(ws, row, f"{ref_base}.{q}", "Industry Development", "Credit Achievement",
    "Marketing Sustainability Achievements", "Data",
    "Identify the target audience and estimated reach for each marketing activity.",
    "Public awareness exposure to green building benefits.")
q += 1

apply_dropdowns(ws)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 2: RESPONSIBLE CONSTRUCTION
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet()
row = setup_sheet(ws2, "Responsible Construction")
row = add_credit_header(ws2, row, "Responsible Construction",
    "The builder's construction practices reduce impacts and promote opportunities for improved environmental and social outcomes.")
ref_base = "RC"
q = 1

row = add_level_header(ws2, row, "Minimum Expectation (Nil points)")

# ── Environmental Management System ──
row = add_criteria_header(ws2, row, "Environmental Management System")

row = add_question(ws2, row, f"{ref_base}.{q}", "Responsible Construction", "Minimum Expectation",
    "Environmental Management System", "Condition (Y/N)",
    "Is any site works contract valued at $10 million or more?",
    "Contract sizes relative to EMS certification thresholds.")
q += 1

row = add_question(ws2, row, f"{ref_base}.{q}", "Responsible Construction", "Minimum Expectation",
    "Environmental Management System", "Descriptive",
    "For contracts under $10M, identify the EMS framework used and explain how it complies (e.g. NSW EMS Guidelines or equivalent).",
    "EMS framework adoption rates in construction.")
q += 1

row = add_question(ws2, row, f"{ref_base}.{q}", "Responsible Construction", "Minimum Expectation",
    "Environmental Management System", "Descriptive",
    "For contracts $10M+, state the certified standard (ISO 14001, BS 7750, or EMAS) and confirm certification validity for the full duration of site works.",
    "Uptake of certified environmental management in construction.")
q += 1

row = add_question(ws2, row, f"{ref_base}.{q}", "Responsible Construction", "Minimum Expectation",
    "Environmental Management System", "Condition (Y/N)",
    "Were different head contractors used for demolition, early works, and main works?",
    "")
q += 1

row = add_question(ws2, row, f"{ref_base}.{q}", "Responsible Construction", "Minimum Expectation",
    "Environmental Management System", "Descriptive",
    "If multiple head contractors, confirm each had an EMS in place and explain how contract values were apportioned.",
    "")
q += 1

row = add_question(ws2, row, f"{ref_base}.{q}", "Responsible Construction", "Minimum Expectation",
    "Environmental Management System", "Descriptive",
    "Explain how the EMS addresses implementation of the EMP and the key environmental impacts targeted.",
    "Relationship between management systems and on-site environmental outcomes.")
q += 1

# ── Environmental Management Plan ──
row = add_criteria_header(ws2, row, "Environmental Management Plan")

row = add_question(ws2, row, f"{ref_base}.{q}", "Responsible Construction", "Minimum Expectation",
    "Environmental Management Plan", "Descriptive",
    "Outline the project-specific EMP, including key impact areas addressed (e.g. noise, dust, stormwater, vegetation).",
    "Most common environmental risks managed during construction.")
q += 1

row = add_question(ws2, row, f"{ref_base}.{q}", "Responsible Construction", "Minimum Expectation",
    "Environmental Management Plan", "Condition (Y/N)",
    "Did the EMP cover the full duration of all site works?",
    "Construction duration and environmental management coverage.")
q += 1

row = add_question(ws2, row, f"{ref_base}.{q}", "Responsible Construction", "Minimum Expectation",
    "Environmental Management Plan", "Data",
    "State the EMP start and end dates.",
    "")
q += 1

row = add_question(ws2, row, f"{ref_base}.{q}", "Responsible Construction", "Minimum Expectation",
    "Environmental Management Plan", "Descriptive",
    "Describe the audit and reporting regime, including frequency and how non-conformances were closed out.",
    "Environmental compliance enforcement during construction.")
q += 1

row = add_question(ws2, row, f"{ref_base}.{q}", "Responsible Construction", "Minimum Expectation",
    "Environmental Management Plan", "Data",
    "Provide the total number of audits, non-conformances identified, and percentage closed out.",
    "Quantitative construction environmental management performance.")
q += 1

# ── Construction and Demolition Waste Diversion ──
row = add_criteria_header(ws2, row, "Construction and Demolition Waste Diversion")

row = add_question(ws2, row, f"{ref_base}.{q}", "Responsible Construction", "Minimum Expectation",
    "Construction and Demolition Waste Diversion", "Data",
    "State total site waste (tonnes), total diverted from landfill (tonnes), and diversion rate (%).",
    "Construction waste diversion rate benchmarking.")
q += 1

row = add_question(ws2, row, f"{ref_base}.{q}", "Responsible Construction", "Minimum Expectation",
    "Construction and Demolition Waste Diversion", "Condition (Y/N)",
    "Does the diversion rate meet the 80% threshold?",
    "")
q += 1

row = add_question(ws2, row, f"{ref_base}.{q}", "Responsible Construction", "Minimum Expectation",
    "Construction and Demolition Waste Diversion", "Descriptive",
    "List waste streams and diversion pathways (recycling, reuse, recovery). Note any excluded streams (special/excavation waste) with justification.",
    "Waste stream composition and recycling pathways in construction.")
q += 1

row = add_question(ws2, row, f"{ref_base}.{q}", "Responsible Construction", "Minimum Expectation",
    "Construction and Demolition Waste Diversion", "Descriptive",
    "Confirm waste contractors provided a Disclosure Statement aligned with the Green Star C&D Waste Reporting Criteria.",
    "Supply chain transparency in waste reporting.")
q += 1

row = add_question(ws2, row, f"{ref_base}.{q}", "Responsible Construction", "Minimum Expectation",
    "Construction and Demolition Waste Diversion", "Data",
    "Provide a breakdown by material type (e.g. concrete, timber, steel, plasterboard) showing tonnes generated and diverted.",
    "Material-specific waste benchmarking across projects.")
q += 1

# ── Sustainability Training ──
row = add_criteria_header(ws2, row, "Sustainability Training")

row = add_question(ws2, row, f"{ref_base}.{q}", "Responsible Construction", "Minimum Expectation",
    "Sustainability Training", "Data",
    "State total site workers on site 3+ days, number trained, and resulting percentage.",
    "Sustainability education reach in the construction workforce.")
q += 1

row = add_question(ws2, row, f"{ref_base}.{q}", "Responsible Construction", "Minimum Expectation",
    "Sustainability Training", "Condition (Y/N)",
    "Does the training rate meet the 95% threshold?",
    "")
q += 1

row = add_question(ws2, row, f"{ref_base}.{q}", "Responsible Construction", "Minimum Expectation",
    "Sustainability Training", "Descriptive",
    "Summarise training content covering: (a) project sustainability attributes, (b) value of certification, (c) site workers' role in delivery.",
    "Training content quality for workforce sustainability literacy.")
q += 1

row = add_question(ws2, row, f"{ref_base}.{q}", "Responsible Construction", "Minimum Expectation",
    "Sustainability Training", "Descriptive",
    "Describe the delivery method (e.g. induction, toolbox talks) and how attendance was tracked.",
    "Effective training delivery models for sustainability in construction.")
q += 1

# ── Credit Achievement ──
row = add_level_header(ws2, row, "Credit Achievement (1 point)")
row = add_criteria_header(ws2, row, "Increased Construction and Demolition Waste Diversion")

row = add_question(ws2, row, f"{ref_base}.{q}", "Responsible Construction", "Credit Achievement",
    "Increased C&D Waste Diversion", "Data",
    "State total site waste (tonnes), total diverted (tonnes), and diversion rate (%).",
    "Higher-tier waste diversion benchmarking.")
q += 1

row = add_question(ws2, row, f"{ref_base}.{q}", "Responsible Construction", "Credit Achievement",
    "Increased C&D Waste Diversion", "Condition (Y/N)",
    "Does the diversion rate meet the 90% threshold?",
    "")
q += 1

row = add_question(ws2, row, f"{ref_base}.{q}", "Responsible Construction", "Credit Achievement",
    "Increased C&D Waste Diversion", "Descriptive",
    "Confirm waste contractors/facilities provided a Compliance Verification Summary per the Green Star C&D Waste Reporting Criteria.",
    "Third-party waste reporting verification practices.")
q += 1

row = add_question(ws2, row, f"{ref_base}.{q}", "Responsible Construction", "Credit Achievement",
    "Increased C&D Waste Diversion", "Descriptive",
    "Identify the waste reporting auditor(s) and their credentials per the Green Star C&D Waste Reporting Criteria.",
    "Auditor capacity and verification standards in waste management.")
q += 1

row = add_question(ws2, row, f"{ref_base}.{q}", "Responsible Construction", "Credit Achievement",
    "Increased C&D Waste Diversion", "Data",
    "List waste processing facilities used, their location, waste types processed, and any GECA C&D Waste Services Standard certification held.",
    "Waste processing infrastructure availability and certification uptake.")
q += 1

apply_dropdowns(ws2)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 3: VERIFICATION AND HANDOVER
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet()
row = setup_sheet(ws3, "Verification and Handover")
row = add_credit_header(ws3, row, "Verification and Handover",
    "The building has been optimised and handed over to deliver a higher level of performance in operation.")
ref_base = "VH"
q = 1

row = add_level_header(ws3, row, "Minimum Expectation (Nil points)")

# ── Metering and Monitoring ──
row = add_criteria_header(ws3, row, "Metering and Monitoring")

row = add_question(ws3, row, f"{ref_base}.{q}", "Verification and Handover", "Minimum Expectation",
    "Metering and Monitoring", "Descriptive",
    "Outline the metering strategy for energy and water across all distinct uses, major uses, and tenancies/units, referencing the CIBSE TM39 schedule.",
    "Metering granularity across building types.")
q += 1

row = add_question(ws3, row, f"{ref_base}.{q}", "Verification and Handover", "Minimum Expectation",
    "Metering and Monitoring", "Data",
    "State the total number of energy and water meters (utility + sub-meters) and distinct end-uses metered.",
    "Metering density benchmarking across building typologies.")
q += 1

row = add_question(ws3, row, f"{ref_base}.{q}", "Verification and Handover", "Minimum Expectation",
    "Metering and Monitoring", "Descriptive",
    "Confirm all meters provide up to 1-hour interval readings, are validated per NABERS Metering Rules, and are NMI pattern-approved or meet an equivalent standard.",
    "Metering quality standards adoption.")
q += 1

row = add_question(ws3, row, f"{ref_base}.{q}", "Verification and Handover", "Minimum Expectation",
    "Metering and Monitoring", "Descriptive",
    "Describe the automatic monitoring system, including consumption trend reporting and alarm/alert functionality for the facilities manager.",
    "Monitoring system capabilities for operational performance.")
q += 1

row = add_question(ws3, row, f"{ref_base}.{q}", "Verification and Handover", "Minimum Expectation",
    "Metering and Monitoring", "Condition (Y/N)",
    "Is this a Class 2 build-to-sell apartment project?",
    "")
q += 1

row = add_question(ws3, row, f"{ref_base}.{q}", "Verification and Handover", "Minimum Expectation",
    "Metering and Monitoring", "Descriptive",
    "If Class 2 build-to-sell, confirm base building trends are provided to the FM and explain how unit meters are handled.",
    "")
q += 1

row = add_question(ws3, row, f"{ref_base}.{q}", "Verification and Handover", "Minimum Expectation",
    "Metering and Monitoring", "Condition (Y/N)",
    "Does the metering strategy rely on connection of tenant meters?",
    "")
q += 1

row = add_question(ws3, row, f"{ref_base}.{q}", "Verification and Handover", "Minimum Expectation",
    "Metering and Monitoring", "Descriptive",
    "If relying on tenant meters, describe the fitout guide or lease clauses ensuring meter connection and monitoring requirements.",
    "")
q += 1

# ── Commissioning and Tuning ──
row = add_criteria_header(ws3, row, "Commissioning and Tuning")

row = add_question(ws3, row, f"{ref_base}.{q}", "Verification and Handover", "Minimum Expectation",
    "Commissioning and Tuning", "Descriptive",
    "Summarise environmental performance targets (energy, water, IEQ, airtightness) set prior to schematic design.",
    "Target-setting practices driving building performance outcomes.")
q += 1

row = add_question(ws3, row, f"{ref_base}.{q}", "Verification and Handover", "Minimum Expectation",
    "Commissioning and Tuning", "Condition (Y/N)",
    "Was the design intent report or OPR signed off by the building owner?",
    "")
q += 1

row = add_question(ws3, row, f"{ref_base}.{q}", "Verification and Handover", "Minimum Expectation",
    "Commissioning and Tuning", "Data",
    "Provide numerical targets for: (a) energy use intensity, (b) water consumption, (c) IEQ parameters, (d) airtightness rate.",
    "Design target benchmarking against actual operational performance.")
q += 1

row = add_question(ws3, row, f"{ref_base}.{q}", "Verification and Handover", "Minimum Expectation",
    "Commissioning and Tuning", "Descriptive",
    "Summarise the services and maintainability review: participants, key outcomes, and close-out status in the Services and Maintainability Report.",
    "Stakeholder collaboration in pre-construction reviews.")
q += 1

row = add_question(ws3, row, f"{ref_base}.{q}", "Verification and Handover", "Minimum Expectation",
    "Commissioning and Tuning", "Descriptive",
    "Identify the commissioning standard followed (e.g. AIRAH DA27, ASHRAE 202, CIBSE Code M). Outline the commissioning plan scope and program.",
    "Commissioning standard adoption rates.")
q += 1

row = add_question(ws3, row, f"{ref_base}.{q}", "Verification and Handover", "Minimum Expectation",
    "Commissioning and Tuning", "Descriptive",
    "List all nominated building systems commissioned (e.g. HVAC, BMCS, lighting, electrical, hydraulic, fire, lifts).",
    "Extent of building systems commissioning across projects.")
q += 1

row = add_question(ws3, row, f"{ref_base}.{q}", "Verification and Handover", "Minimum Expectation",
    "Commissioning and Tuning", "Descriptive",
    "Explain how airtightness targets were set (per ATTMA Guide) and how the air barrier schematic was reviewed before end of design development.",
    "Airtightness design integration practices.")
q += 1

row = add_question(ws3, row, f"{ref_base}.{q}", "Verification and Handover", "Minimum Expectation",
    "Commissioning and Tuning", "Descriptive",
    "Describe airtightness testing: practitioner's ATTMA level, standard followed (AS/NZS ISO 9972), areas tested (whole building or sample), and selection of high-risk assemblies.",
    "Airtightness testing practices and practitioner qualifications.")
q += 1

row = add_question(ws3, row, f"{ref_base}.{q}", "Verification and Handover", "Minimum Expectation",
    "Commissioning and Tuning", "Data",
    "Provide airtightness results (air permeability rates) per tested area. Note whether targets were met and any improvement opportunities identified.",
    "Building envelope quality benchmarking.")
q += 1

row = add_question(ws3, row, f"{ref_base}.{q}", "Verification and Handover", "Minimum Expectation",
    "Commissioning and Tuning", "Descriptive",
    "Describe the tuning commitment: contractual arrangement, tuning plan, and team roles (FM, ICA, head contractor, subcontractors).",
    "Post-occupancy building optimisation approaches.")
q += 1

row = add_question(ws3, row, f"{ref_base}.{q}", "Verification and Handover", "Minimum Expectation",
    "Commissioning and Tuning", "Data",
    "State tuning duration (min. 12 months), frequency of adjustments (min. quarterly), and planned start date.",
    "Tuning duration and frequency indicators.")
q += 1

# ── Building Information ──
row = add_criteria_header(ws3, row, "Building Information")

row = add_question(ws3, row, f"{ref_base}.{q}", "Verification and Handover", "Minimum Expectation",
    "Building Information", "Descriptive",
    "Summarise the O&M information provided: maintenance procedures, schedules, contacts, warranties, and as-built drawings for nominated systems.",
    "Handover documentation completeness.")
q += 1

row = add_question(ws3, row, f"{ref_base}.{q}", "Verification and Handover", "Minimum Expectation",
    "Building Information", "Descriptive",
    "Explain how O&M information guides the FM team on keeping records current and responding to monitoring system alerts/faults.",
    "")
q += 1

row = add_question(ws3, row, f"{ref_base}.{q}", "Verification and Handover", "Minimum Expectation",
    "Building Information", "Descriptive",
    "Confirm a CIBSE TM31 building logbook was prepared covering all nominated systems and delivered to the owner prior to occupation.",
    "Building logbook adoption rates.")
q += 1

row = add_question(ws3, row, f"{ref_base}.{q}", "Verification and Handover", "Minimum Expectation",
    "Building Information", "Descriptive",
    "Describe building user information: availability to occupants, relevance to audience, and digital format used (e.g. website, app, signage).",
    "Approaches to engaging occupants in sustainable operations.")
q += 1

row = add_question(ws3, row, f"{ref_base}.{q}", "Verification and Handover", "Minimum Expectation",
    "Building Information", "Condition (Y/N)",
    "Is the building user information in an editable digital format accessible to the FM team?",
    "Digital information management maturity in building operations.")
q += 1

row = add_question(ws3, row, f"{ref_base}.{q}", "Verification and Handover", "Minimum Expectation",
    "Building Information", "Descriptive",
    "State the format and platform used for building user information.",
    "")
q += 1

# ── Credit Achievement ──
row = add_level_header(ws3, row, "Credit Achievement (1 point)")

row = add_question(ws3, row, f"{ref_base}.{q}", "Verification and Handover", "Credit Achievement",
    "General", "Condition (Y/N)",
    "Is the Total Building Services Value over $20 million? (If yes, both Soft Landings and ICA criteria must be met.)",
    "Building services expenditure relative to commissioning requirements.")
q += 1

# ── Soft Landings Approach ──
row = add_criteria_header(ws3, row, "Soft Landings Approach")

row = add_question(ws3, row, f"{ref_base}.{q}", "Verification and Handover", "Credit Achievement",
    "Soft Landings Approach", "Descriptive",
    "Describe implementation of CIBSE ANZ Soft Landings Stages 1-4.",
    "Soft landings adoption and implementation quality.")
q += 1

row = add_question(ws3, row, f"{ref_base}.{q}", "Verification and Handover", "Credit Achievement",
    "Soft Landings Approach", "Condition (Y/N)",
    "Are the sample worksheets for Stages 1-3 completed and Stage 4 actions identified?",
    "")
q += 1

row = add_question(ws3, row, f"{ref_base}.{q}", "Verification and Handover", "Credit Achievement",
    "Soft Landings Approach", "Descriptive",
    "Describe the FM team's involvement: commissioning participation, O&M manual development and sign-off, and pre-handover training received.",
    "FM involvement in building transition and performance gap.")
q += 1

row = add_question(ws3, row, f"{ref_base}.{q}", "Verification and Handover", "Credit Achievement",
    "Soft Landings Approach", "Descriptive",
    "Explain arrangements for FM access to design and construction team members for two years post practical completion.",
    "Post-handover support duration and structure.")
q += 1

row = add_question(ws3, row, f"{ref_base}.{q}", "Verification and Handover", "Credit Achievement",
    "Soft Landings Approach", "Condition (Y/N)",
    "Has Stage 5 (post-occupancy evaluation) been planned or implemented?",
    "Voluntary post-occupancy evaluation uptake.")
q += 1

# ── Independent Commissioning Agent ──
row = add_criteria_header(ws3, row, "Independent Commissioning Agent")

row = add_question(ws3, row, f"{ref_base}.{q}", "Verification and Handover", "Credit Achievement",
    "Independent Commissioning Agent", "Descriptive",
    "Identify the ICA: qualifications, commissioning knowledge, and experience with 2+ similar projects.",
    "ICA workforce capacity and qualification levels.")
q += 1

row = add_question(ws3, row, f"{ref_base}.{q}", "Verification and Handover", "Credit Achievement",
    "Independent Commissioning Agent", "Descriptive",
    "Confirm the ICA was appointed before design development and is independent of all design/installation consultants and contractors.",
    "Independence arrangements in commissioning oversight.")
q += 1

row = add_question(ws3, row, f"{ref_base}.{q}", "Verification and Handover", "Credit Achievement",
    "Independent Commissioning Agent", "Descriptive",
    "Summarise the ICA's involvement across phases: design development, tender, construction, commissioning, and tuning.",
    "Breadth of ICA involvement for commissioning effectiveness.")
q += 1

row = add_question(ws3, row, f"{ref_base}.{q}", "Verification and Handover", "Credit Achievement",
    "Independent Commissioning Agent", "Condition (Y/N)",
    "Is the ICA role fulfilled by more than one person?",
    "")
q += 1

apply_dropdowns(ws3)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 4: RESPONSIBLE RESOURCE MANAGEMENT
# ════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet()
row = setup_sheet(ws4, "Responsible Resource Mgmt")
row = add_credit_header(ws4, row, "Responsible Resource Management",
    "Operational waste and resources can be separated and recovered in a safe and efficient manner.")
ref_base = "RRM"
q = 1

row = add_level_header(ws4, row, "Minimum Expectation (Nil points)")

# ── Collection of Waste Streams ──
row = add_criteria_header(ws4, row, "Collection of Waste Streams")

row = add_question(ws4, row, f"{ref_base}.{q}", "Responsible Resource Management", "Minimum Expectation",
    "Collection of Waste Streams", "Descriptive",
    "List all separately collected waste streams (min: general waste, paper/cardboard, glass, plastic, plus one additional). Justify the additional stream selected.",
    "Diversity of operational waste separation across building types.")
q += 1

row = add_question(ws4, row, f"{ref_base}.{q}", "Responsible Resource Management", "Minimum Expectation",
    "Collection of Waste Streams", "Condition (Y/N)",
    "Is any single non-food waste stream expected to exceed 5% of total annual operational waste by volume?",
    "Dominant waste streams for targeted reduction strategies.")
q += 1

row = add_question(ws4, row, f"{ref_base}.{q}", "Responsible Resource Management", "Minimum Expectation",
    "Collection of Waste Streams", "Descriptive",
    "If yes, identify the stream(s) exceeding 5% and describe their separate collection provisions.",
    "")
q += 1

row = add_question(ws4, row, f"{ref_base}.{q}", "Responsible Resource Management", "Minimum Expectation",
    "Collection of Waste Streams", "Descriptive",
    "Describe bin/chute intake locations, proximity to waste generation points, and labelling approach.",
    "Waste infrastructure design for occupant convenience.")
q += 1

row = add_question(ws4, row, f"{ref_base}.{q}", "Responsible Resource Management", "Minimum Expectation",
    "Collection of Waste Streams", "Condition (Y/N)",
    "Does the project include cold shell or excluded tenancy spaces outside the rating scope?",
    "")
q += 1

row = add_question(ws4, row, f"{ref_base}.{q}", "Responsible Resource Management", "Minimum Expectation",
    "Collection of Waste Streams", "Descriptive",
    "If yes, describe fitout guide, lease clauses, or contracts ensuring waste separation in those spaces.",
    "")
q += 1

row = add_question(ws4, row, f"{ref_base}.{q}", "Responsible Resource Management", "Minimum Expectation",
    "Collection of Waste Streams", "Condition (Y/N)",
    "Is co-mingled recycling used for any waste streams?",
    "Prevalence of co-mingled vs source-separated recycling.")
q += 1

row = add_question(ws4, row, f"{ref_base}.{q}", "Responsible Resource Management", "Minimum Expectation",
    "Collection of Waste Streams", "Descriptive",
    "If co-mingled, identify which streams and confirm acceptance by the waste collection service.",
    "")
q += 1

# ── Dedicated Waste Storage Area ──
row = add_criteria_header(ws4, row, "Dedicated Waste Storage Area")

row = add_question(ws4, row, f"{ref_base}.{q}", "Responsible Resource Management", "Minimum Expectation",
    "Dedicated Waste Storage Area", "Descriptive",
    "Describe the storage area(s): location, total area, and layout for keeping waste streams separate.",
    "Waste storage design patterns across building types.")
q += 1

row = add_question(ws4, row, f"{ref_base}.{q}", "Responsible Resource Management", "Minimum Expectation",
    "Dedicated Waste Storage Area", "Data",
    "Provide forecasted waste generation rates, collection frequency per stream, and storage capacity calculations. Identify the best practice guideline used.",
    "Waste generation rate data for cross-project benchmarking.")
q += 1

row = add_question(ws4, row, f"{ref_base}.{q}", "Responsible Resource Management", "Minimum Expectation",
    "Dedicated Waste Storage Area", "Descriptive",
    "Describe collection vehicle access: parking, driveways, height clearances, and manoeuvring per AS 2890.2:2018.",
    "Waste collection logistics design.")
q += 1

row = add_question(ws4, row, f"{ref_base}.{q}", "Responsible Resource Management", "Minimum Expectation",
    "Dedicated Waste Storage Area", "Condition (Y/N)",
    "Is this a tenanted building where excluded tenancies contribute to the waste storage strategy?",
    "")
q += 1

row = add_question(ws4, row, f"{ref_base}.{q}", "Responsible Resource Management", "Minimum Expectation",
    "Dedicated Waste Storage Area", "Descriptive",
    "If yes, explain how waste from excluded tenancies was estimated and factored into storage sizing.",
    "")
q += 1

# ── Safe and Efficient Access to Waste Storage ──
row = add_criteria_header(ws4, row, "Safe and Efficient Access to Waste Storage")

row = add_question(ws4, row, f"{ref_base}.{q}", "Responsible Resource Management", "Minimum Expectation",
    "Safe and Efficient Access to Waste Storage", "Descriptive",
    "Identify the waste specialist/contractor who signed off on designs, including their organisation and relevant experience (min. 3 years).",
    "Waste specialist involvement in building design.")
q += 1

row = add_question(ws4, row, f"{ref_base}.{q}", "Responsible Resource Management", "Minimum Expectation",
    "Safe and Efficient Access to Waste Storage", "Descriptive",
    "Summarise the sign-off findings confirming storage areas are adequately sized and located for safe collection.",
    "Waste management design validation practices.")
q += 1

row = add_question(ws4, row, f"{ref_base}.{q}", "Responsible Resource Management", "Minimum Expectation",
    "Safe and Efficient Access to Waste Storage", "Data",
    "Provide: building GFA, number of occupants/units, waste storage area (m²), and estimated annual operational waste (tonnes or m³/year).",
    "Waste generation benchmarks normalised by building size and occupancy.")
q += 1

apply_dropdowns(ws4)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 5: RESPONSIBLE PROCUREMENT
# ════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet()
row = setup_sheet(ws5, "Responsible Procurement")
row = add_credit_header(ws5, row, "Responsible Procurement",
    "The procurement process for key products, materials, and services follows best practice environmental and social principles.")
ref_base = "RP"
q = 1

row = add_level_header(ws5, row, "Credit Achievement (1 point)")

# ── Risk and Opportunity Assessment ──
row = add_criteria_header(ws5, row, "Risk and Opportunity Assessment")

row = add_question(ws5, row, f"{ref_base}.{q}", "Responsible Procurement", "Credit Achievement",
    "Risk and Opportunity Assessment", "Condition (Y/N)",
    "Was the risk and opportunity assessment completed before appointment of the head contractor?",
    "Timing of supply chain risk assessment relative to procurement.")
q += 1

row = add_question(ws5, row, f"{ref_base}.{q}", "Responsible Procurement", "Credit Achievement",
    "Risk and Opportunity Assessment", "Condition (Y/N)",
    "Did the building owner provide input into the assessment?",
    "Stakeholder involvement in supply chain risk assessment.")
q += 1

row = add_question(ws5, row, f"{ref_base}.{q}", "Responsible Procurement", "Credit Achievement",
    "Risk and Opportunity Assessment", "Descriptive",
    "Identify who conducted the assessment.",
    "")
q += 1

row = add_question(ws5, row, f"{ref_base}.{q}", "Responsible Procurement", "Credit Achievement",
    "Risk and Opportunity Assessment", "Descriptive",
    "List the 10+ key supply chain items (min. 2 building services, 1 building material). Briefly justify each selection.",
    "Supply chain risk hotspots across building projects.")
q += 1

row = add_question(ws5, row, f"{ref_base}.{q}", "Responsible Procurement", "Credit Achievement",
    "Risk and Opportunity Assessment", "Descriptive",
    "Describe how risks and opportunities were evaluated per ISO 20400 Clause 4.3: human rights, labour, environment, fair practices, consumer issues, community.",
    "Depth of sustainability risk analysis in procurement.")
q += 1

row = add_question(ws5, row, f"{ref_base}.{q}", "Responsible Procurement", "Credit Achievement",
    "Risk and Opportunity Assessment", "Data",
    "For each key item, summarise priority risks/opportunities and risk ratings (high/medium/low) per ISO 20400 issue area.",
    "Aggregated supply chain risk profiles for industry guidance.")
q += 1

row = add_question(ws5, row, f"{ref_base}.{q}", "Responsible Procurement", "Credit Achievement",
    "Risk and Opportunity Assessment", "Descriptive",
    "Explain the methodology used to analyse and prioritise risks. Note any tools or references beyond ISO 20400 Annex A.",
    "Risk assessment methodologies for industry knowledge sharing.")
q += 1

# ── Responsible Procurement Plan ──
row = add_criteria_header(ws5, row, "Responsible Procurement Plan")

row = add_question(ws5, row, f"{ref_base}.{q}", "Responsible Procurement", "Credit Achievement",
    "Responsible Procurement Plan", "Descriptive",
    "Outline the plan's environmental, social, and economic objectives addressing the identified risks and opportunities.",
    "Procurement sustainability objective ambition across projects.")
q += 1

row = add_question(ws5, row, f"{ref_base}.{q}", "Responsible Procurement", "Credit Achievement",
    "Responsible Procurement Plan", "Descriptive",
    "Describe data collection, monitoring, and reporting requirements per ISO 20400 Clause 6.5. State metrics tracked and frequency.",
    "Procurement monitoring and reporting approaches.")
q += 1

row = add_question(ws5, row, f"{ref_base}.{q}", "Responsible Procurement", "Credit Achievement",
    "Responsible Procurement Plan", "Descriptive",
    "Describe the framework for incentivising contractors and trades. Provide examples of incentive mechanisms.",
    "Supply chain incentive models for sustainability outcomes.")
q += 1

row = add_question(ws5, row, f"{ref_base}.{q}", "Responsible Procurement", "Credit Achievement",
    "Responsible Procurement Plan", "Descriptive",
    "Explain how the plan was embedded in tender documentation for the head contractor and relevant trades.",
    "Integration of sustainability requirements into procurement workflows.")
q += 1

row = add_question(ws5, row, f"{ref_base}.{q}", "Responsible Procurement", "Credit Achievement",
    "Responsible Procurement Plan", "Condition (Y/N)",
    "Was the head contractor engaged under a design and construct (D&C) contract?",
    "Procurement models and their impact on sustainability integration.")
q += 1

row = add_question(ws5, row, f"{ref_base}.{q}", "Responsible Procurement", "Credit Achievement",
    "Responsible Procurement Plan", "Descriptive",
    "If D&C, explain the head contractor's role in developing the plan and how it was embedded in subcontractor tenders.",
    "")
q += 1

row = add_question(ws5, row, f"{ref_base}.{q}", "Responsible Procurement", "Credit Achievement",
    "Responsible Procurement Plan", "Descriptive",
    "Describe plan implementation during construction: data collection, monitoring, and reporting activities per ISO 20400 Clause 7.",
    "Real-world sustainable procurement implementation effectiveness.")
q += 1

row = add_question(ws5, row, f"{ref_base}.{q}", "Responsible Procurement", "Credit Achievement",
    "Responsible Procurement Plan", "Data",
    "List key items with procurement actions taken. Summarise the sustainability outcome per item (e.g. modern slavery risk mitigated, local supply used).",
    "Outcome-level data on responsible procurement for advocacy/policy.")
q += 1

row = add_question(ws5, row, f"{ref_base}.{q}", "Responsible Procurement", "Credit Achievement",
    "Responsible Procurement Plan", "Data",
    "Note any supply chain risks that materialised and corrective actions taken. State items fully vs partially implemented.",
    "Procurement plan implementation rates and real-world risk events.")
q += 1

apply_dropdowns(ws5)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 6: RESPONSIBLE STRUCTURE
# ════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet()
row = setup_sheet(ws6, "Responsible Structure")
row = add_credit_header(ws6, row, "Responsible Structure",
    "The building's structure is comprised of responsibly manufactured products.")
ref_base = "RS"
q = 1

row = add_level_header(ws6, row, "Credit Achievement (1 point)")
row = add_criteria_header(ws6, row, "Good Practice Products")

row = add_question(ws6, row, f"{ref_base}.{q}", "Responsible Structure", "Credit Achievement",
    "Good Practice Products", "Data",
    "State the percentage of structural components (by cost) that meet a Responsible Products Value (RPV) of at least 10.",
    "Tracks responsible product uptake in structural materials.")
q += 1

row = add_question(ws6, row, f"{ref_base}.{q}", "Responsible Structure", "Credit Achievement",
    "Good Practice Products", "Condition (Y/N)",
    "Does the percentage meet the required threshold for Credit Achievement?",
    "")
q += 1

row = add_question(ws6, row, f"{ref_base}.{q}", "Responsible Structure", "Credit Achievement",
    "Good Practice Products", "Descriptive",
    "List the structural products included in the calculation and their individual RPV scores from the Responsible Products Calculator.",
    "Material-level data on responsible certification uptake.")
q += 1

row = add_question(ws6, row, f"{ref_base}.{q}", "Responsible Structure", "Credit Achievement",
    "Good Practice Products", "Descriptive",
    "Describe how product data was collected and verified (e.g. EPDs, supplier declarations, certification evidence).",
    "Documents verification practices for responsible product claims.")
q += 1

row = add_level_header(ws6, row, "Exceptional Performance (2 points)")
row = add_criteria_header(ws6, row, "Best Practice Products")

row = add_question(ws6, row, f"{ref_base}.{q}", "Responsible Structure", "Exceptional Performance",
    "Best Practice Products", "Data",
    "State the percentage of structural components (by cost) that meet an RPV of at least 15.",
    "Higher-tier responsible product benchmarking.")
q += 1

row = add_question(ws6, row, f"{ref_base}.{q}", "Responsible Structure", "Exceptional Performance",
    "Best Practice Products", "Condition (Y/N)",
    "Does the percentage meet the required threshold for Exceptional Performance?",
    "")
q += 1

row = add_criteria_header(ws6, row, "Good Practice Products (Alternative)")

row = add_question(ws6, row, f"{ref_base}.{q}", "Responsible Structure", "Exceptional Performance",
    "Good Practice Products (Alt)", "Data",
    "Alternatively, state if 95% or more of structural components (by cost) meet an RPV of at least 10.",
    "Near-universal responsible product adoption tracking.")
q += 1

apply_dropdowns(ws6)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 7: RESPONSIBLE ENVELOPE
# ════════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet()
row = setup_sheet(ws7, "Responsible Envelope")
row = add_credit_header(ws7, row, "Responsible Envelope",
    "The building's envelope is comprised of responsibly manufactured products.")
ref_base = "RE"
q = 1

row = add_level_header(ws7, row, "Credit Achievement (1 point)")
row = add_criteria_header(ws7, row, "Good Practice Products")

row = add_question(ws7, row, f"{ref_base}.{q}", "Responsible Envelope", "Credit Achievement",
    "Good Practice Products", "Data",
    "State the percentage of envelope components (by cost) that meet an RPV of at least 10.",
    "Tracks responsible product uptake in building envelope materials.")
q += 1

row = add_question(ws7, row, f"{ref_base}.{q}", "Responsible Envelope", "Credit Achievement",
    "Good Practice Products", "Condition (Y/N)",
    "Does the percentage meet the required threshold for Credit Achievement?",
    "")
q += 1

row = add_question(ws7, row, f"{ref_base}.{q}", "Responsible Envelope", "Credit Achievement",
    "Good Practice Products", "Descriptive",
    "List the envelope products included (façade, glazing, roofing, external walls) and their individual RPV scores.",
    "Material-level data on envelope product responsibility.")
q += 1

row = add_level_header(ws7, row, "Exceptional Performance (2 points)")
row = add_criteria_header(ws7, row, "Best Practice Products")

row = add_question(ws7, row, f"{ref_base}.{q}", "Responsible Envelope", "Exceptional Performance",
    "Best Practice Products", "Data",
    "State the percentage of envelope components (by cost) that meet an RPV of at least 15.",
    "Higher-tier envelope material responsibility benchmarking.")
q += 1

row = add_question(ws7, row, f"{ref_base}.{q}", "Responsible Envelope", "Exceptional Performance",
    "Best Practice Products", "Condition (Y/N)",
    "Does the percentage meet the required threshold for Exceptional Performance?",
    "")
q += 1

apply_dropdowns(ws7)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 8: RESPONSIBLE SYSTEMS
# ════════════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet()
row = setup_sheet(ws8, "Responsible Systems")
row = add_credit_header(ws8, row, "Responsible Systems",
    "The building's systems are comprised of responsibly manufactured products.")
ref_base = "RSy"
q = 1

row = add_level_header(ws8, row, "Credit Achievement (1 point)")
row = add_criteria_header(ws8, row, "Good Practice Products")

row = add_question(ws8, row, f"{ref_base}.{q}", "Responsible Systems", "Credit Achievement",
    "Good Practice Products", "Data",
    "State the percentage of building systems (by cost) that meet an RPV of at least 10.",
    "Tracks responsible product uptake in HVAC, electrical, hydraulic systems.")
q += 1

row = add_question(ws8, row, f"{ref_base}.{q}", "Responsible Systems", "Credit Achievement",
    "Good Practice Products", "Condition (Y/N)",
    "Does the percentage meet the required threshold for Credit Achievement?",
    "")
q += 1

row = add_question(ws8, row, f"{ref_base}.{q}", "Responsible Systems", "Credit Achievement",
    "Good Practice Products", "Descriptive",
    "List the building systems products included (HVAC, electrical, hydraulic, fire, vertical transport) and their individual RPV scores.",
    "Systems-level data on responsible product procurement.")
q += 1

row = add_level_header(ws8, row, "Exceptional Performance (2 points)")
row = add_criteria_header(ws8, row, "Best Practice Products")

row = add_question(ws8, row, f"{ref_base}.{q}", "Responsible Systems", "Exceptional Performance",
    "Best Practice Products", "Data",
    "State the percentage of building systems (by cost) that meet an RPV of at least 15.",
    "Higher-tier building systems responsibility benchmarking.")
q += 1

row = add_question(ws8, row, f"{ref_base}.{q}", "Responsible Systems", "Exceptional Performance",
    "Best Practice Products", "Condition (Y/N)",
    "Does the percentage meet the required threshold for Exceptional Performance?",
    "")
q += 1

apply_dropdowns(ws8)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 9: RESPONSIBLE FINISHES
# ════════════════════════════════════════════════════════════════════════════
ws9 = wb.create_sheet()
row = setup_sheet(ws9, "Responsible Finishes")
row = add_credit_header(ws9, row, "Responsible Finishes",
    "The building's finishes are comprised of responsibly manufactured products.")
ref_base = "RF"
q = 1

row = add_level_header(ws9, row, "Credit Achievement (1 point)")
row = add_criteria_header(ws9, row, "Good Practice Products")

row = add_question(ws9, row, f"{ref_base}.{q}", "Responsible Finishes", "Credit Achievement",
    "Good Practice Products", "Data",
    "State the percentage of finishes (by cost) that meet an RPV of at least 10.",
    "Tracks responsible product uptake in interior finishes.")
q += 1

row = add_question(ws9, row, f"{ref_base}.{q}", "Responsible Finishes", "Credit Achievement",
    "Good Practice Products", "Condition (Y/N)",
    "Does the percentage meet the required threshold for Credit Achievement?",
    "")
q += 1

row = add_question(ws9, row, f"{ref_base}.{q}", "Responsible Finishes", "Credit Achievement",
    "Good Practice Products", "Descriptive",
    "List the finish products included (flooring, ceilings, internal walls, joinery, paints) and their individual RPV scores.",
    "Material-level data on interior finish responsibility.")
q += 1

row = add_level_header(ws9, row, "Exceptional Performance (2 points)")
row = add_criteria_header(ws9, row, "Best Practice Products")

row = add_question(ws9, row, f"{ref_base}.{q}", "Responsible Finishes", "Exceptional Performance",
    "Best Practice Products", "Data",
    "State the percentage of finishes (by cost) that meet an RPV of at least 15.",
    "Higher-tier interior finish responsibility benchmarking.")
q += 1

row = add_question(ws9, row, f"{ref_base}.{q}", "Responsible Finishes", "Exceptional Performance",
    "Best Practice Products", "Condition (Y/N)",
    "Does the percentage meet the required threshold for Exceptional Performance?",
    "")
q += 1

apply_dropdowns(ws9)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 10: IMPACTS DISCLOSURE
# ════════════════════════════════════════════════════════════════════════════
ws10 = wb.create_sheet()
row = setup_sheet(ws10, "Impacts Disclosure")
row = add_credit_header(ws10, row, "Impacts Disclosure",
    "Environmental impacts of building products are disclosed and data is made available for research.")
ref_base = "ID2"
q = 1

row = add_level_header(ws10, row, "Credit Achievement (1 point)")
row = add_criteria_header(ws10, row, "Environmental Performance Disclosure")

row = add_question(ws10, row, f"{ref_base}.{q}", "Impacts Disclosure", "Credit Achievement",
    "Environmental Performance Disclosure", "Data",
    "State the percentage of products (by cost) for which Environmental Product Declarations (EPDs) have been obtained.",
    "Tracks EPD availability and uptake in the construction supply chain.")
q += 1

row = add_question(ws10, row, f"{ref_base}.{q}", "Impacts Disclosure", "Credit Achievement",
    "Environmental Performance Disclosure", "Condition (Y/N)",
    "Does the EPD coverage meet the required threshold?",
    "")
q += 1

row = add_question(ws10, row, f"{ref_base}.{q}", "Impacts Disclosure", "Credit Achievement",
    "Environmental Performance Disclosure", "Descriptive",
    "List the products with EPDs, identifying which are product-specific vs industry-average EPDs.",
    "Documents EPD granularity and specificity across product categories.")
q += 1

row = add_question(ws10, row, f"{ref_base}.{q}", "Impacts Disclosure", "Credit Achievement",
    "Environmental Performance Disclosure", "Descriptive",
    "Describe how EPD data was collected and how compliance with EN 15804 or ISO 21930 was verified.",
    "EPD standard compliance verification practices.")
q += 1

row = add_criteria_header(ws10, row, "Data Sharing")

row = add_question(ws10, row, f"{ref_base}.{q}", "Impacts Disclosure", "Credit Achievement",
    "Data Sharing", "Condition (Y/N)",
    "Has the project agreed to share product environmental data with GBCA for research purposes?",
    "Tracks willingness to contribute to industry-wide environmental benchmarking.")
q += 1

row = add_question(ws10, row, f"{ref_base}.{q}", "Impacts Disclosure", "Credit Achievement",
    "Data Sharing", "Descriptive",
    "Describe what data will be shared and any confidentiality arrangements in place.",
    "")
q += 1

row = add_level_header(ws10, row, "Exceptional Performance (2 points)")
row = add_criteria_header(ws10, row, "Enhanced Disclosure")

row = add_question(ws10, row, f"{ref_base}.{q}", "Impacts Disclosure", "Exceptional Performance",
    "Enhanced Disclosure", "Data",
    "State the percentage of products (by cost) with product-specific EPDs (not industry-average).",
    "Measures product-specific environmental transparency.")
q += 1

row = add_question(ws10, row, f"{ref_base}.{q}", "Impacts Disclosure", "Exceptional Performance",
    "Enhanced Disclosure", "Condition (Y/N)",
    "Does the product-specific EPD coverage meet the enhanced threshold?",
    "")
q += 1

apply_dropdowns(ws10)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 11: CLEAN AIR
# ════════════════════════════════════════════════════════════════════════════
ws11 = wb.create_sheet()
row = setup_sheet(ws11, "Clean Air")
row = add_credit_header(ws11, row, "Clean Air",
    "Pollutants entering the building are minimised, and a high level of air quality is provided.")
ref_base = "CA"
q = 1

row = add_level_header(ws11, row, "Minimum Expectation (Nil points)")

# Ventilation System Attributes
row = add_criteria_header(ws11, row, "Ventilation System Attributes")

row = add_question(ws11, row, f"{ref_base}.{q}", "Clean Air", "Minimum Expectation",
    "Ventilation System Attributes", "Descriptive",
    "Describe how ventilation intakes meet minimum separation distances from pollution sources in 95% of regularly occupied areas.",
    "Documents air intake design relative to pollution sources.")
q += 1

row = add_question(ws11, row, f"{ref_base}.{q}", "Clean Air", "Minimum Expectation",
    "Ventilation System Attributes", "Descriptive",
    "Describe ductwork cleaning procedures undertaken prior to occupation, identifying the standard followed (ACR 2021 or SMACNA).",
    "Tracks pre-occupation ductwork hygiene practices.")
q += 1

row = add_question(ws11, row, f"{ref_base}.{q}", "Clean Air", "Minimum Expectation",
    "Ventilation System Attributes", "Condition (Y/N)",
    "Does the building have ductwork requiring cleaning?",
    "")
q += 1

# Provision of Outdoor Air
row = add_criteria_header(ws11, row, "Provision of Outdoor Air")

row = add_question(ws11, row, f"{ref_base}.{q}", "Clean Air", "Minimum Expectation",
    "Provision of Outdoor Air", "Descriptive",
    "Describe the pathway used to demonstrate high levels of effective outdoor air to 95% of regularly occupied areas.",
    "Documents ventilation design approach by building type.")
q += 1

row = add_question(ws11, row, f"{ref_base}.{q}", "Clean Air", "Minimum Expectation",
    "Provision of Outdoor Air", "Data",
    "State the outdoor air rate provided (L/s per person or air changes per hour) and the standard/code used as reference.",
    "Outdoor air rate benchmarking across building types.")
q += 1

# Exhaust or Elimination of Pollutants
row = add_criteria_header(ws11, row, "Exhaust or Elimination of Pollutants")

row = add_question(ws11, row, f"{ref_base}.{q}", "Clean Air", "Minimum Expectation",
    "Exhaust or Elimination of Pollutants", "Descriptive",
    "Describe how pollutants from printing equipment, cooking, and vehicles are exhausted or eliminated in 95% of regularly occupied areas.",
    "Maps pollutant control strategies across building uses.")
q += 1

row = add_question(ws11, row, f"{ref_base}.{q}", "Clean Air", "Minimum Expectation",
    "Exhaust or Elimination of Pollutants", "Condition (Y/N)",
    "Are there printing/photocopying rooms, commercial kitchens, or enclosed car parks in the building?",
    "")
q += 1

row = add_level_header(ws11, row, "Credit Achievement (1 point)")

row = add_criteria_header(ws11, row, "Ventilation System Attributes (Enhanced)")

row = add_question(ws11, row, f"{ref_base}.{q}", "Clean Air", "Credit Achievement",
    "Ventilation System Attributes", "Descriptive",
    "Describe access provisions for maintenance of moisture and debris-catching components in 95% of regularly occupied areas.",
    "Documents maintainability of ventilation systems for air quality.")
q += 1

row = add_question(ws11, row, f"{ref_base}.{q}", "Clean Air", "Credit Achievement",
    "Ventilation System Attributes", "Condition (Y/N)",
    "Are there any fan coil units where access to both sides for cleaning is not possible?",
    "")
q += 1

row = add_question(ws11, row, f"{ref_base}.{q}", "Clean Air", "Credit Achievement",
    "Ventilation System Attributes", "Descriptive",
    "If yes, describe the alternative compliance pathway used (MERV 8+ filters, UV-C treatment, or antimicrobial coating).",
    "")
q += 1

row = add_criteria_header(ws11, row, "Provision of Outdoor Air (Enhanced)")

row = add_question(ws11, row, f"{ref_base}.{q}", "Clean Air", "Credit Achievement",
    "Provision of Outdoor Air", "Descriptive",
    "Describe how enhanced outdoor air levels are provided in 95% of regularly occupied areas per the relevant pathway.",
    "Enhanced ventilation design benchmarking.")
q += 1

row = add_question(ws11, row, f"{ref_base}.{q}", "Clean Air", "Credit Achievement",
    "Provision of Outdoor Air", "Data",
    "State the enhanced outdoor air rate achieved and the percentage improvement over Minimum Expectation.",
    "Quantifies ventilation performance uplift.")
q += 1

apply_dropdowns(ws11)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 12: LIGHT QUALITY
# ════════════════════════════════════════════════════════════════════════════
ws12 = wb.create_sheet()
row = setup_sheet(ws12, "Light Quality")
row = add_credit_header(ws12, row, "Light Quality",
    "The building provides good daylight and its lighting is of high quality.")
ref_base = "LQ"
q = 1

row = add_level_header(ws12, row, "Minimum Expectation (Nil points)")

row = add_criteria_header(ws12, row, "Lighting Comfort")

row = add_question(ws12, row, f"{ref_base}.{q}", "Light Quality", "Minimum Expectation",
    "Lighting Comfort", "Descriptive",
    "Describe how lighting in 95% of regularly occupied areas meets the requirements for the relevant building class pathway.",
    "Lighting comfort compliance approach by building type.")
q += 1

row = add_question(ws12, row, f"{ref_base}.{q}", "Light Quality", "Minimum Expectation",
    "Lighting Comfort", "Data",
    "State the illuminance levels (lux) provided in key space types and the standard referenced.",
    "Illuminance benchmarking data across space types.")
q += 1

row = add_criteria_header(ws12, row, "Glare from Light Sources")

row = add_question(ws12, row, f"{ref_base}.{q}", "Light Quality", "Minimum Expectation",
    "Glare from Light Sources", "Descriptive",
    "Describe the pathway used to limit glare from light sources in 95% of regularly occupied areas.",
    "Glare control strategies by building type.")
q += 1

row = add_question(ws12, row, f"{ref_base}.{q}", "Light Quality", "Minimum Expectation",
    "Glare from Light Sources", "Data",
    "State the UGR (Unified Glare Rating) achieved or the glare control measures implemented.",
    "Glare performance benchmarking.")
q += 1

row = add_criteria_header(ws12, row, "Daylight Strategy")

row = add_question(ws12, row, f"{ref_base}.{q}", "Light Quality", "Minimum Expectation",
    "Daylight Strategy", "Descriptive",
    "Summarise the daylight strategy prepared by the project team, including how design maximises daylight access and controls external glare.",
    "Documents daylight design intent and strategies.")
q += 1

row = add_question(ws12, row, f"{ref_base}.{q}", "Light Quality", "Minimum Expectation",
    "Daylight Strategy", "Data",
    "State the proportion of regularly occupied areas with access to daylight (per GBCA calculation guide).",
    "Daylight access benchmarking across building types.")
q += 1

row = add_level_header(ws12, row, "Credit Achievement (1 point)")

row = add_criteria_header(ws12, row, "Artificial Lighting OR Daylight")

row = add_question(ws12, row, f"{ref_base}.{q}", "Light Quality", "Credit Achievement",
    "Artificial Lighting / Daylight", "Condition (Y/N)",
    "Is Credit Achievement being claimed via the Artificial Lighting criterion?",
    "")
q += 1

row = add_question(ws12, row, f"{ref_base}.{q}", "Light Quality", "Credit Achievement",
    "Artificial Lighting", "Descriptive",
    "If Artificial Lighting: Describe how lighting provides high quality light exposure supporting task visibility, visual comfort and well-being per the relevant pathway.",
    "Artificial lighting quality approach by building type.")
q += 1

row = add_question(ws12, row, f"{ref_base}.{q}", "Light Quality", "Credit Achievement",
    "Daylight", "Descriptive",
    "If Daylight: Describe how the building provides high daylight levels per the relevant pathway, including the Daylight Autonomy calculation method.",
    "Daylight performance modelling approach.")
q += 1

row = add_question(ws12, row, f"{ref_base}.{q}", "Light Quality", "Credit Achievement",
    "Daylight", "Data",
    "State the Daylight Autonomy achieved (target: 160 lux for 80% of nominated hours) and the percentage of regularly occupied areas meeting this.",
    "Daylight Autonomy benchmarking for research.")
q += 1

row = add_question(ws12, row, f"{ref_base}.{q}", "Light Quality", "Credit Achievement",
    "Daylight", "Descriptive",
    "Describe the external glare control measures for viewing façades and skylights in regularly occupied areas.",
    "External glare control strategies.")
q += 1

row = add_level_header(ws12, row, "Exceptional Performance (2 points)")

row = add_question(ws12, row, f"{ref_base}.{q}", "Light Quality", "Exceptional Performance",
    "Both Criteria", "Condition (Y/N)",
    "Are both Artificial Lighting AND Daylight criteria being met for Exceptional Performance?",
    "")
q += 1

apply_dropdowns(ws12)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 13: ACOUSTIC COMFORT
# ════════════════════════════════════════════════════════════════════════════
ws13 = wb.create_sheet()
row = setup_sheet(ws13, "Acoustic Comfort")
row = add_credit_header(ws13, row, "Acoustic Comfort",
    "The building provides acoustic comfort for building occupants.")
ref_base = "AC"
q = 1

row = add_level_header(ws13, row, "Minimum Expectation (Nil points)")

row = add_criteria_header(ws13, row, "Acoustic Comfort Strategy")

row = add_question(ws13, row, f"{ref_base}.{q}", "Acoustic Comfort", "Minimum Expectation",
    "Acoustic Comfort Strategy", "Condition (Y/N)",
    "Was an Acoustic Comfort Strategy prepared by a qualified acoustic consultant during design?",
    "Tracks acoustic consultant engagement on projects.")
q += 1

row = add_question(ws13, row, f"{ref_base}.{q}", "Acoustic Comfort", "Minimum Expectation",
    "Acoustic Comfort Strategy", "Descriptive",
    "List the standards, legislation, and guidelines identified as applicable in the strategy.",
    "Documents acoustic regulatory and best practice context.")
q += 1

row = add_question(ws13, row, f"{ref_base}.{q}", "Acoustic Comfort", "Minimum Expectation",
    "Acoustic Comfort Strategy", "Descriptive",
    "Summarise which acoustic considerations are relevant for each space type: quiet enjoyment, functional use, intrusive noise control, privacy, noise transfer, speech intelligibility.",
    "Maps acoustic priorities across building space types.")
q += 1

row = add_question(ws13, row, f"{ref_base}.{q}", "Acoustic Comfort", "Minimum Expectation",
    "Acoustic Comfort Strategy", "Descriptive",
    "Describe how the design solution achieves the proposed performance metrics.",
    "Documents acoustic design intent translation to solutions.")
q += 1

row = add_level_header(ws13, row, "Credit Achievement (1 point)")

row = add_criteria_header(ws13, row, "Acoustic Performance")

row = add_question(ws13, row, f"{ref_base}.{q}", "Acoustic Comfort", "Credit Achievement",
    "Acoustic Performance", "Descriptive",
    "Identify which acoustic criteria are being met for the building class (refer to rating tool table for required number of criteria).",
    "Tracks acoustic criteria selection by building type.")
q += 1

row = add_criteria_header(ws13, row, "Maximum Internal Noise Levels")

row = add_question(ws13, row, f"{ref_base}.{q}", "Acoustic Comfort", "Credit Achievement",
    "Maximum Internal Noise Levels", "Data",
    "State the measured internal ambient noise levels in regularly occupied areas and compare to AS/NZS 2107:2016 Table 1 upper limits.",
    "Internal noise level benchmarking data.")
q += 1

row = add_question(ws13, row, f"{ref_base}.{q}", "Acoustic Comfort", "Credit Achievement",
    "Maximum Internal Noise Levels", "Condition (Y/N)",
    "For Class 2, 3 and 9 buildings: Do bedroom/sleeping spaces meet the NSW Road Noise Policy Sleep Disturbance criteria?",
    "Sleep disturbance compliance tracking for residential buildings.")
q += 1

row = add_criteria_header(ws13, row, "Minimum Internal Noise Levels")

row = add_question(ws13, row, f"{ref_base}.{q}", "Acoustic Comfort", "Credit Achievement",
    "Minimum Internal Noise Levels", "Data",
    "State whether internal ambient noise levels are no less than 5 dB below AS/NZS 2107:2016 Table 1 lower limits.",
    "Minimum noise floor benchmarking for acoustic masking.")
q += 1

row = add_criteria_header(ws13, row, "Acoustic Separation")

row = add_question(ws13, row, f"{ref_base}.{q}", "Acoustic Comfort", "Credit Achievement",
    "Acoustic Separation", "Descriptive",
    "Describe the pathway used to address noise transmission through walls and floors (sound transmission or sound insulation).",
    "Acoustic separation approach by building type.")
q += 1

row = add_question(ws13, row, f"{ref_base}.{q}", "Acoustic Comfort", "Credit Achievement",
    "Acoustic Separation", "Data",
    "State the sound transmission class (STC) or weighted sound reduction index (Rw) achieved for key partitions.",
    "Acoustic partition performance benchmarking.")
q += 1

row = add_criteria_header(ws13, row, "Impact Noise Transfer")

row = add_question(ws13, row, f"{ref_base}.{q}", "Acoustic Comfort", "Credit Achievement",
    "Impact Noise Transfer", "Data",
    "State the impact noise transfer performance achieved for floors above regularly occupied areas (per ISO 16283-2:2020).",
    "Impact noise performance benchmarking.")
q += 1

row = add_criteria_header(ws13, row, "Reverberation Control")

row = add_question(ws13, row, f"{ref_base}.{q}", "Acoustic Comfort", "Credit Achievement",
    "Reverberation Control", "Data",
    "State the measured reverberation times against AS/NZS 2107:2016 Table 1 recommendations.",
    "Reverberation time benchmarking across space types.")
q += 1

row = add_question(ws13, row, f"{ref_base}.{q}", "Acoustic Comfort", "Credit Achievement",
    "Reverberation Control", "Descriptive",
    "For open plan spaces: Describe the acoustic absorption treatment (percentage of floor/ceiling area with NRC ≥0.5).",
    "Open plan acoustic treatment quantification.")
q += 1

apply_dropdowns(ws13)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 14: EXPOSURE TO TOXINS
# ════════════════════════════════════════════════════════════════════════════
ws14 = wb.create_sheet()
row = setup_sheet(ws14, "Exposure to Toxins")
row = add_credit_header(ws14, row, "Exposure to Toxins",
    "The building reduces occupant exposure to toxins.")
ref_base = "ET"
q = 1

row = add_level_header(ws14, row, "Minimum Expectation (Nil points)")

row = add_criteria_header(ws14, row, "Hazardous Materials Survey")

row = add_question(ws14, row, f"{ref_base}.{q}", "Exposure to Toxins", "Minimum Expectation",
    "Hazardous Materials Survey", "Condition (Y/N)",
    "Does the project involve refurbishment or alteration of an existing building constructed before 2004?",
    "Tracks projects requiring hazardous materials assessment.")
q += 1

row = add_question(ws14, row, f"{ref_base}.{q}", "Exposure to Toxins", "Minimum Expectation",
    "Hazardous Materials Survey", "Descriptive",
    "If yes, describe the hazardous materials survey undertaken (asbestos, lead paint, PCBs) and the findings.",
    "Documents hazardous material prevalence in building stock.")
q += 1

row = add_question(ws14, row, f"{ref_base}.{q}", "Exposure to Toxins", "Minimum Expectation",
    "Hazardous Materials Survey", "Descriptive",
    "Describe the management or removal plan for any hazardous materials identified.",
    "Hazardous material remediation approaches.")
q += 1

row = add_criteria_header(ws14, row, "Reduced Exposure to Toxins")

row = add_question(ws14, row, f"{ref_base}.{q}", "Exposure to Toxins", "Minimum Expectation",
    "Reduced Exposure to Toxins", "Descriptive",
    "Describe how the building design reduces occupant exposure to toxins through material selection and ventilation.",
    "Toxin reduction design strategies.")
q += 1

row = add_level_header(ws14, row, "Credit Achievement (1 point)")

row = add_criteria_header(ws14, row, "Low-Emission Products")

row = add_question(ws14, row, f"{ref_base}.{q}", "Exposure to Toxins", "Credit Achievement",
    "Low-Emission Products", "Data",
    "State the percentage of paints, adhesives, sealants, and carpets (by area or cost) that meet low-VOC emission standards.",
    "Low-VOC product uptake benchmarking.")
q += 1

row = add_question(ws14, row, f"{ref_base}.{q}", "Exposure to Toxins", "Credit Achievement",
    "Low-Emission Products", "Descriptive",
    "List the emission standards or certifications met by key products (e.g. GECA, Green Tag, Declare, Cradle to Cradle).",
    "Third-party certification uptake in interior products.")
q += 1

row = add_question(ws14, row, f"{ref_base}.{q}", "Exposure to Toxins", "Credit Achievement",
    "Low-Emission Products", "Descriptive",
    "Describe verification process for product emission compliance (test reports, certificates, supplier declarations).",
    "Product emission verification practices.")
q += 1

row = add_criteria_header(ws14, row, "Formaldehyde Limits")

row = add_question(ws14, row, f"{ref_base}.{q}", "Exposure to Toxins", "Credit Achievement",
    "Formaldehyde Limits", "Data",
    "State the percentage of engineered wood products meeting E0 or E1 formaldehyde emission classification.",
    "Formaldehyde emission compliance in timber products.")
q += 1

row = add_question(ws14, row, f"{ref_base}.{q}", "Exposure to Toxins", "Credit Achievement",
    "Formaldehyde Limits", "Descriptive",
    "List the engineered wood products used and their formaldehyde emission classifications.",
    "Timber product formaldehyde benchmarking.")
q += 1

apply_dropdowns(ws14)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 15: AMENITY AND COMFORT
# ════════════════════════════════════════════════════════════════════════════
ws15 = wb.create_sheet()
row = setup_sheet(ws15, "Amenity and Comfort")
row = add_credit_header(ws15, row, "Amenity and Comfort",
    "The building provides comfortable thermal conditions and occupant amenity.")
ref_base = "AmC"
q = 1

row = add_level_header(ws15, row, "Minimum Expectation (Nil points)")

row = add_criteria_header(ws15, row, "Thermal Comfort")

row = add_question(ws15, row, f"{ref_base}.{q}", "Amenity and Comfort", "Minimum Expectation",
    "Thermal Comfort", "Descriptive",
    "Describe the thermal comfort strategy for regularly occupied areas, identifying the comfort standard applied (ASHRAE 55, ISO 7730, or equivalent).",
    "Thermal comfort standard adoption by building type.")
q += 1

row = add_question(ws15, row, f"{ref_base}.{q}", "Amenity and Comfort", "Minimum Expectation",
    "Thermal Comfort", "Data",
    "State the predicted percentage of people dissatisfied (PPD) or the thermal comfort category achieved.",
    "Thermal comfort performance benchmarking.")
q += 1

row = add_criteria_header(ws15, row, "End of Trip Facilities")

row = add_question(ws15, row, f"{ref_base}.{q}", "Amenity and Comfort", "Minimum Expectation",
    "End of Trip Facilities", "Condition (Y/N)",
    "Does the building provide end-of-trip facilities for active transport users?",
    "Tracks end-of-trip facility provision rates.")
q += 1

row = add_question(ws15, row, f"{ref_base}.{q}", "Amenity and Comfort", "Minimum Expectation",
    "End of Trip Facilities", "Data",
    "State the number of bicycle parking spaces, showers, and lockers provided.",
    "End-of-trip facility provision benchmarking.")
q += 1

row = add_level_header(ws15, row, "Credit Achievement (1 point)")

row = add_criteria_header(ws15, row, "Enhanced Thermal Comfort")

row = add_question(ws15, row, f"{ref_base}.{q}", "Amenity and Comfort", "Credit Achievement",
    "Enhanced Thermal Comfort", "Descriptive",
    "Describe how enhanced thermal comfort is provided (individual control, mixed-mode ventilation, or other measures).",
    "Enhanced thermal comfort strategies.")
q += 1

row = add_question(ws15, row, f"{ref_base}.{q}", "Amenity and Comfort", "Credit Achievement",
    "Enhanced Thermal Comfort", "Condition (Y/N)",
    "Are occupants provided with individual control over their thermal environment?",
    "Individual thermal control provision tracking.")
q += 1

row = add_criteria_header(ws15, row, "Enhanced End of Trip Facilities")

row = add_question(ws15, row, f"{ref_base}.{q}", "Amenity and Comfort", "Credit Achievement",
    "Enhanced End of Trip Facilities", "Data",
    "State the enhanced provision of bicycle parking, showers, and lockers compared to minimum requirements.",
    "Enhanced end-of-trip facility benchmarking.")
q += 1

apply_dropdowns(ws15)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 16: CONNECTION TO NATURE
# ════════════════════════════════════════════════════════════════════════════
ws16 = wb.create_sheet()
row = setup_sheet(ws16, "Connection to Nature")
row = add_credit_header(ws16, row, "Connection to Nature",
    "Building occupants can experience nature through views and access.")
ref_base = "CN"
q = 1

row = add_level_header(ws16, row, "Credit Achievement (1 point)")

row = add_criteria_header(ws16, row, "Views to Nature")

row = add_question(ws16, row, f"{ref_base}.{q}", "Connection to Nature", "Credit Achievement",
    "Views to Nature", "Data",
    "State the percentage of regularly occupied areas with quality views to nature (vegetation, water, sky).",
    "Nature view access benchmarking across building types.")
q += 1

row = add_question(ws16, row, f"{ref_base}.{q}", "Connection to Nature", "Credit Achievement",
    "Views to Nature", "Descriptive",
    "Describe the nature elements visible from occupied spaces (external vegetation, green walls, water features, distant natural landscapes).",
    "Categorises nature view types provided.")
q += 1

row = add_question(ws16, row, f"{ref_base}.{q}", "Connection to Nature", "Credit Achievement",
    "Views to Nature", "Descriptive",
    "Describe any internal biophilic elements (indoor plants, living walls, nature imagery) where external views are limited.",
    "Documents internal biophilic design elements.")
q += 1

row = add_criteria_header(ws16, row, "Access to Nature")

row = add_question(ws16, row, f"{ref_base}.{q}", "Connection to Nature", "Credit Achievement",
    "Access to Nature", "Condition (Y/N)",
    "Does the building provide direct access to outdoor green space or nature for occupants?",
    "Tracks outdoor green space access provision.")
q += 1

row = add_question(ws16, row, f"{ref_base}.{q}", "Connection to Nature", "Credit Achievement",
    "Access to Nature", "Data",
    "State the total area of accessible outdoor green space provided and the distance from regularly occupied areas.",
    "Outdoor green space area and proximity benchmarking.")
q += 1

row = add_question(ws16, row, f"{ref_base}.{q}", "Connection to Nature", "Credit Achievement",
    "Access to Nature", "Descriptive",
    "Describe the nature elements in accessible outdoor spaces (gardens, courtyards, rooftop terraces, balconies).",
    "Outdoor nature space typology documentation.")
q += 1

row = add_level_header(ws16, row, "Exceptional Performance (2 points)")

row = add_question(ws16, row, f"{ref_base}.{q}", "Connection to Nature", "Exceptional Performance",
    "Enhanced Connection", "Condition (Y/N)",
    "Are both Views to Nature and Access to Nature criteria met at an enhanced level?",
    "")
q += 1

row = add_question(ws16, row, f"{ref_base}.{q}", "Connection to Nature", "Exceptional Performance",
    "Enhanced Connection", "Data",
    "State the enhanced percentage of occupied areas with quality nature views and/or enhanced outdoor space provision.",
    "Enhanced nature connection benchmarking.")
q += 1

apply_dropdowns(ws16)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 17: CLIMATE RESILIENCE
# ════════════════════════════════════════════════════════════════════════════
ws17 = wb.create_sheet()
row = setup_sheet(ws17, "Climate Resilience")
row = add_credit_header(ws17, row, "Climate Resilience",
    "The building is designed to adapt to and mitigate climate change impacts.")
ref_base = "CR"
q = 1

row = add_level_header(ws17, row, "Minimum Expectation (Nil points)")

row = add_criteria_header(ws17, row, "Climate Risk Assessment")

row = add_question(ws17, row, f"{ref_base}.{q}", "Climate Resilience", "Minimum Expectation",
    "Climate Risk Assessment", "Condition (Y/N)",
    "Was a climate risk assessment undertaken for the project?",
    "Tracks climate risk assessment adoption in building projects.")
q += 1

row = add_question(ws17, row, f"{ref_base}.{q}", "Climate Resilience", "Minimum Expectation",
    "Climate Risk Assessment", "Descriptive",
    "Identify the climate scenarios and time horizons assessed (e.g. RCP 4.5, RCP 8.5; 2050, 2090).",
    "Climate scenario usage in building risk assessment.")
q += 1

row = add_question(ws17, row, f"{ref_base}.{q}", "Climate Resilience", "Minimum Expectation",
    "Climate Risk Assessment", "Descriptive",
    "List the climate hazards assessed (extreme heat, flooding, bushfire, sea level rise, storms, drought) and their projected impacts on the building.",
    "Documents climate hazard exposure by location and building type.")
q += 1

row = add_question(ws17, row, f"{ref_base}.{q}", "Climate Resilience", "Minimum Expectation",
    "Climate Risk Assessment", "Descriptive",
    "Describe the risk assessment methodology used and the data sources referenced (e.g. CSIRO, BOM, local council data).",
    "Climate risk assessment methodology documentation.")
q += 1

row = add_criteria_header(ws17, row, "Adaptation Plan")

row = add_question(ws17, row, f"{ref_base}.{q}", "Climate Resilience", "Minimum Expectation",
    "Adaptation Plan", "Descriptive",
    "Summarise the adaptation measures identified to address the key climate risks.",
    "Catalogues climate adaptation strategies in the built environment.")
q += 1

row = add_question(ws17, row, f"{ref_base}.{q}", "Climate Resilience", "Minimum Expectation",
    "Adaptation Plan", "Descriptive",
    "Describe how the adaptation plan has been integrated into the building design.",
    "Documents design integration of climate adaptation.")
q += 1

row = add_level_header(ws17, row, "Credit Achievement (1 point)")

row = add_criteria_header(ws17, row, "Enhanced Climate Resilience")

row = add_question(ws17, row, f"{ref_base}.{q}", "Climate Resilience", "Credit Achievement",
    "Enhanced Climate Resilience", "Descriptive",
    "Describe the enhanced climate resilience measures implemented beyond Minimum Expectation.",
    "Enhanced climate resilience strategies.")
q += 1

row = add_question(ws17, row, f"{ref_base}.{q}", "Climate Resilience", "Credit Achievement",
    "Enhanced Climate Resilience", "Data",
    "State the design life of the building and the climate scenario it has been designed to withstand.",
    "Building design life and climate scenario alignment.")
q += 1

apply_dropdowns(ws17)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 18: OPERATIONS RESILIENCE
# ════════════════════════════════════════════════════════════════════════════
ws18 = wb.create_sheet()
row = setup_sheet(ws18, "Operations Resilience")
row = add_credit_header(ws18, row, "Operations Resilience",
    "The building can maintain essential services during disruptions.")
ref_base = "OR"
q = 1

row = add_level_header(ws18, row, "Credit Achievement (1 point)")

row = add_criteria_header(ws18, row, "Essential Services Resilience")

row = add_question(ws18, row, f"{ref_base}.{q}", "Operations Resilience", "Credit Achievement",
    "Essential Services Resilience", "Descriptive",
    "Identify the essential services and functions the building needs to maintain during disruptions.",
    "Essential services identification by building type.")
q += 1

row = add_question(ws18, row, f"{ref_base}.{q}", "Operations Resilience", "Credit Achievement",
    "Essential Services Resilience", "Descriptive",
    "Describe the backup systems and redundancy measures in place for power, water, and communications.",
    "Backup systems and redundancy strategies.")
q += 1

row = add_question(ws18, row, f"{ref_base}.{q}", "Operations Resilience", "Credit Achievement",
    "Essential Services Resilience", "Data",
    "State the duration of backup power capacity (hours/days) for essential services.",
    "Backup power duration benchmarking.")
q += 1

row = add_question(ws18, row, f"{ref_base}.{q}", "Operations Resilience", "Credit Achievement",
    "Essential Services Resilience", "Condition (Y/N)",
    "Is on-site water storage or alternative water supply provided for emergencies?",
    "Emergency water supply provision tracking.")
q += 1

row = add_question(ws18, row, f"{ref_base}.{q}", "Operations Resilience", "Credit Achievement",
    "Essential Services Resilience", "Data",
    "If yes, state the emergency water storage capacity (litres) and the days of supply this represents.",
    "Emergency water storage benchmarking.")
q += 1

row = add_criteria_header(ws18, row, "Business Continuity Plan")

row = add_question(ws18, row, f"{ref_base}.{q}", "Operations Resilience", "Credit Achievement",
    "Business Continuity Plan", "Condition (Y/N)",
    "Has a business continuity plan been developed for building operations?",
    "Business continuity planning adoption tracking.")
q += 1

row = add_question(ws18, row, f"{ref_base}.{q}", "Operations Resilience", "Credit Achievement",
    "Business Continuity Plan", "Descriptive",
    "Summarise the key elements of the business continuity plan and the disruption scenarios addressed.",
    "Business continuity plan scope documentation.")
q += 1

apply_dropdowns(ws18)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 19: COMMUNITY RESILIENCE
# ════════════════════════════════════════════════════════════════════════════
ws19 = wb.create_sheet()
row = setup_sheet(ws19, "Community Resilience")
row = add_credit_header(ws19, row, "Community Resilience",
    "The building supports the resilience of the broader community.")
ref_base = "CoR"
q = 1

row = add_level_header(ws19, row, "Credit Achievement (1 point)")

row = add_criteria_header(ws19, row, "Community Refuge or Support")

row = add_question(ws19, row, f"{ref_base}.{q}", "Community Resilience", "Credit Achievement",
    "Community Refuge or Support", "Condition (Y/N)",
    "Can the building serve as a community refuge during extreme weather events or emergencies?",
    "Community refuge capacity in building stock.")
q += 1

row = add_question(ws19, row, f"{ref_base}.{q}", "Community Resilience", "Credit Achievement",
    "Community Refuge or Support", "Descriptive",
    "Describe the building features that enable it to support the community during emergencies (shelter capacity, cooling/heating refuge, power/water access).",
    "Community resilience features in buildings.")
q += 1

row = add_question(ws19, row, f"{ref_base}.{q}", "Community Resilience", "Credit Achievement",
    "Community Refuge or Support", "Data",
    "State the estimated number of people the building could shelter or support during an emergency.",
    "Community shelter capacity benchmarking.")
q += 1

row = add_question(ws19, row, f"{ref_base}.{q}", "Community Resilience", "Credit Achievement",
    "Community Refuge or Support", "Descriptive",
    "Describe any agreements or arrangements with local authorities for the building to serve as community refuge.",
    "Documents formal community resilience arrangements.")
q += 1

apply_dropdowns(ws19)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 20: HEAT RESILIENCE
# ════════════════════════════════════════════════════════════════════════════
ws20 = wb.create_sheet()
row = setup_sheet(ws20, "Heat Resilience")
row = add_credit_header(ws20, row, "Heat Resilience",
    "The building maintains safe conditions during extreme heat events.")
ref_base = "HR"
q = 1

row = add_level_header(ws20, row, "Credit Achievement (1 point)")

row = add_criteria_header(ws20, row, "Passive Survivability")

row = add_question(ws20, row, f"{ref_base}.{q}", "Heat Resilience", "Credit Achievement",
    "Passive Survivability", "Descriptive",
    "Describe how the building maintains safe thermal conditions during power outages in extreme heat events.",
    "Passive survivability design strategies.")
q += 1

row = add_question(ws20, row, f"{ref_base}.{q}", "Heat Resilience", "Credit Achievement",
    "Passive Survivability", "Data",
    "State the maximum internal temperature modelled during a design extreme heat event with no active cooling.",
    "Passive survivability thermal performance.")
q += 1

row = add_question(ws20, row, f"{ref_base}.{q}", "Heat Resilience", "Credit Achievement",
    "Passive Survivability", "Descriptive",
    "Describe the passive design features that support thermal resilience (thermal mass, insulation, shading, natural ventilation).",
    "Passive thermal resilience design features.")
q += 1

row = add_criteria_header(ws20, row, "Urban Heat Mitigation")

row = add_question(ws20, row, f"{ref_base}.{q}", "Heat Resilience", "Credit Achievement",
    "Urban Heat Mitigation", "Descriptive",
    "Describe measures to reduce the building's contribution to urban heat island effect.",
    "Urban heat mitigation strategies.")
q += 1

row = add_question(ws20, row, f"{ref_base}.{q}", "Heat Resilience", "Credit Achievement",
    "Urban Heat Mitigation", "Data",
    "State the Solar Reflectance Index (SRI) of roof and hardscape surfaces.",
    "Surface reflectance benchmarking for heat mitigation.")
q += 1

row = add_question(ws20, row, f"{ref_base}.{q}", "Heat Resilience", "Credit Achievement",
    "Urban Heat Mitigation", "Data",
    "State the percentage of site area with green cover or permeable surfaces.",
    "Green cover and permeability benchmarking.")
q += 1

apply_dropdowns(ws20)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 21: GRID RESILIENCE
# ════════════════════════════════════════════════════════════════════════════
ws21 = wb.create_sheet()
row = setup_sheet(ws21, "Grid Resilience")
row = add_credit_header(ws21, row, "Grid Resilience",
    "The building supports electricity grid stability and resilience.")
ref_base = "GR"
q = 1

row = add_level_header(ws21, row, "Credit Achievement (1 point)")

row = add_criteria_header(ws21, row, "Demand Response Capability")

row = add_question(ws21, row, f"{ref_base}.{q}", "Grid Resilience", "Credit Achievement",
    "Demand Response Capability", "Condition (Y/N)",
    "Is the building capable of participating in demand response programs?",
    "Demand response capability in building stock.")
q += 1

row = add_question(ws21, row, f"{ref_base}.{q}", "Grid Resilience", "Credit Achievement",
    "Demand Response Capability", "Descriptive",
    "Describe the systems and controls that enable demand response participation.",
    "Demand response enabling technologies.")
q += 1

row = add_question(ws21, row, f"{ref_base}.{q}", "Grid Resilience", "Credit Achievement",
    "Demand Response Capability", "Data",
    "State the peak demand reduction capacity (kW) available through demand response.",
    "Demand response capacity benchmarking.")
q += 1

row = add_criteria_header(ws21, row, "On-site Energy Storage")

row = add_question(ws21, row, f"{ref_base}.{q}", "Grid Resilience", "Credit Achievement",
    "On-site Energy Storage", "Condition (Y/N)",
    "Is on-site energy storage (battery) installed?",
    "Battery storage adoption tracking.")
q += 1

row = add_question(ws21, row, f"{ref_base}.{q}", "Grid Resilience", "Credit Achievement",
    "On-site Energy Storage", "Data",
    "State the battery storage capacity (kWh) and the usable capacity for grid services.",
    "Battery storage capacity benchmarking.")
q += 1

row = add_question(ws21, row, f"{ref_base}.{q}", "Grid Resilience", "Credit Achievement",
    "On-site Energy Storage", "Descriptive",
    "Describe how the battery storage is configured to support grid resilience (load shifting, peak shaving, backup power).",
    "Battery storage use case documentation.")
q += 1

row = add_criteria_header(ws21, row, "Vehicle-to-Building")

row = add_question(ws21, row, f"{ref_base}.{q}", "Grid Resilience", "Credit Achievement",
    "Vehicle-to-Building", "Condition (Y/N)",
    "Is vehicle-to-building (V2B) or vehicle-to-grid (V2G) capability provided?",
    "V2B/V2G capability adoption tracking.")
q += 1

row = add_question(ws21, row, f"{ref_base}.{q}", "Grid Resilience", "Credit Achievement",
    "Vehicle-to-Building", "Data",
    "State the number of EV charging points with V2B/V2G capability.",
    "V2B/V2G infrastructure benchmarking.")
q += 1

apply_dropdowns(ws21)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 22: ENERGY SOURCE
# ════════════════════════════════════════════════════════════════════════════
ws22 = wb.create_sheet()
row = setup_sheet(ws22, "Energy Source")
row = add_credit_header(ws22, row, "Energy Source",
    "The building is powered by renewable energy sources.")
ref_base = "ES"
q = 1

row = add_level_header(ws22, row, "Pathway Selection")

row = add_question(ws22, row, f"{ref_base}.{q}", "Energy Source", "Pathway Selection",
    "Pathway", "Condition (Y/N)",
    "Is this an owner-operated building (Pathway A)?",
    "Building ownership model relative to energy source pathway.")
q += 1

row = add_question(ws22, row, f"{ref_base}.{q}", "Energy Source", "Pathway Selection",
    "Pathway", "Condition (Y/N)",
    "Is this a tenant-operated building (Pathway B)?",
    "")
q += 1

row = add_level_header(ws22, row, "Minimum Expectation (Nil points)")

row = add_criteria_header(ws22, row, "Renewable Energy Procurement")

row = add_question(ws22, row, f"{ref_base}.{q}", "Energy Source", "Minimum Expectation",
    "Renewable Energy Procurement", "Data",
    "State the percentage of building energy to be sourced from renewables.",
    "Renewable energy procurement targets benchmarking.")
q += 1

row = add_question(ws22, row, f"{ref_base}.{q}", "Energy Source", "Minimum Expectation",
    "Renewable Energy Procurement", "Descriptive",
    "Describe the renewable energy sources (on-site solar PV, PPAs, green power, RECs) and the mix of each.",
    "Renewable energy source mix documentation.")
q += 1

row = add_question(ws22, row, f"{ref_base}.{q}", "Energy Source", "Minimum Expectation",
    "Renewable Energy Procurement", "Condition (Y/N)",
    "Is on-site renewable energy generation installed?",
    "On-site renewable generation adoption tracking.")
q += 1

row = add_question(ws22, row, f"{ref_base}.{q}", "Energy Source", "Minimum Expectation",
    "Renewable Energy Procurement", "Data",
    "If yes, state the on-site renewable capacity (kWp) and estimated annual generation (kWh).",
    "On-site renewable capacity benchmarking.")
q += 1

row = add_level_header(ws22, row, "Credit Achievement (1-3 points)")

row = add_criteria_header(ws22, row, "Enhanced Renewable Energy")

row = add_question(ws22, row, f"{ref_base}.{q}", "Energy Source", "Credit Achievement",
    "Enhanced Renewable Energy", "Data",
    "State the percentage of building energy from renewables for Credit Achievement level claimed.",
    "Renewable energy percentage tiers.")
q += 1

row = add_question(ws22, row, f"{ref_base}.{q}", "Energy Source", "Credit Achievement",
    "Enhanced Renewable Energy", "Descriptive",
    "Describe contractual arrangements for renewable energy (PPA terms, GreenPower percentage, LGC procurement).",
    "Renewable energy contract structures.")
q += 1

row = add_question(ws22, row, f"{ref_base}.{q}", "Energy Source", "Credit Achievement",
    "Enhanced Renewable Energy", "Data",
    "State the contract duration for renewable energy procurement (years).",
    "Renewable energy contract duration benchmarking.")
q += 1

row = add_level_header(ws22, row, "Exceptional Performance")

row = add_question(ws22, row, f"{ref_base}.{q}", "Energy Source", "Exceptional Performance",
    "100% Renewable", "Condition (Y/N)",
    "Is 100% of the building's energy sourced from renewables?",
    "100% renewable building tracking.")
q += 1

apply_dropdowns(ws22)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 23: ENERGY USE
# ════════════════════════════════════════════════════════════════════════════
ws23 = wb.create_sheet()
row = setup_sheet(ws23, "Energy Use")
row = add_credit_header(ws23, row, "Energy Use",
    "The building minimises operational energy use.")
ref_base = "EU"
q = 1

row = add_level_header(ws23, row, "Pathway Selection")

row = add_question(ws23, row, f"{ref_base}.{q}", "Energy Use", "Pathway Selection",
    "Pathway", "Descriptive",
    "Identify the pathway used: A (Reference building), B (NABERS Commitment Agreement), C (Residential), or D (Small non-residential).",
    "Energy use pathway selection by building type.")
q += 1

row = add_level_header(ws23, row, "Minimum Expectation (Nil points)")

row = add_criteria_header(ws23, row, "Energy Performance")

row = add_question(ws23, row, f"{ref_base}.{q}", "Energy Use", "Minimum Expectation",
    "Energy Performance", "Data",
    "State the predicted energy use intensity (kWh/m²/year or MJ/m²/year).",
    "Energy use intensity benchmarking across building types.")
q += 1

row = add_question(ws23, row, f"{ref_base}.{q}", "Energy Use", "Minimum Expectation",
    "Energy Performance", "Descriptive",
    "Describe the energy modelling methodology used and the reference standard (NCC Section J, NABERS, NatHERS).",
    "Energy modelling methodology documentation.")
q += 1

row = add_question(ws23, row, f"{ref_base}.{q}", "Energy Use", "Minimum Expectation",
    "Energy Performance", "Data",
    "State the percentage improvement over the reference building or code baseline.",
    "Energy performance improvement quantification.")
q += 1

row = add_level_header(ws23, row, "Credit Achievement (1-3 points)")

row = add_criteria_header(ws23, row, "Enhanced Energy Performance")

row = add_question(ws23, row, f"{ref_base}.{q}", "Energy Use", "Credit Achievement",
    "Enhanced Energy Performance", "Data",
    "State the energy performance level achieved for the Credit Achievement tier claimed.",
    "Enhanced energy performance tiers.")
q += 1

row = add_question(ws23, row, f"{ref_base}.{q}", "Energy Use", "Credit Achievement",
    "Enhanced Energy Performance", "Descriptive",
    "For Pathway B: State the NABERS Energy Commitment Agreement rating achieved.",
    "NABERS commitment agreement ratings.")
q += 1

row = add_question(ws23, row, f"{ref_base}.{q}", "Energy Use", "Credit Achievement",
    "Enhanced Energy Performance", "Data",
    "For Pathway C (Residential): State the NatHERS rating achieved.",
    "NatHERS rating benchmarking for residential.")
q += 1

row = add_question(ws23, row, f"{ref_base}.{q}", "Energy Use", "Credit Achievement",
    "Enhanced Energy Performance", "Descriptive",
    "Describe the key energy efficiency measures contributing to performance (envelope, HVAC, lighting, equipment).",
    "Energy efficiency measure cataloguing.")
q += 1

row = add_level_header(ws23, row, "Exceptional Performance")

row = add_question(ws23, row, f"{ref_base}.{q}", "Energy Use", "Exceptional Performance",
    "Leading Performance", "Data",
    "State the energy performance level achieved for Exceptional Performance.",
    "Leading energy performance benchmarking.")
q += 1

apply_dropdowns(ws23)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 24: UPFRONT CARBON REDUCTION
# ════════════════════════════════════════════════════════════════════════════
ws24 = wb.create_sheet()
row = setup_sheet(ws24, "Upfront Carbon Reduction")
row = add_credit_header(ws24, row, "Upfront Carbon Reduction",
    "The building minimises embodied carbon in construction.")
ref_base = "UCR"
q = 1

row = add_level_header(ws24, row, "Pathway Selection")

row = add_question(ws24, row, f"{ref_base}.{q}", "Upfront Carbon Reduction", "Pathway Selection",
    "Pathway", "Descriptive",
    "Identify the pathway used: A (Benchmark pathway) or B (Reference building pathway).",
    "Embodied carbon assessment pathway selection.")
q += 1

row = add_level_header(ws24, row, "Minimum Expectation (Nil points)")

row = add_criteria_header(ws24, row, "Life Cycle Assessment")

row = add_question(ws24, row, f"{ref_base}.{q}", "Upfront Carbon Reduction", "Minimum Expectation",
    "Life Cycle Assessment", "Condition (Y/N)",
    "Was a life cycle assessment (LCA) undertaken for the building?",
    "LCA adoption in building projects.")
q += 1

row = add_question(ws24, row, f"{ref_base}.{q}", "Upfront Carbon Reduction", "Minimum Expectation",
    "Life Cycle Assessment", "Descriptive",
    "Identify the LCA practitioner and their qualifications. State the LCA tool or software used.",
    "LCA practitioner capacity and tool usage.")
q += 1

row = add_question(ws24, row, f"{ref_base}.{q}", "Upfront Carbon Reduction", "Minimum Expectation",
    "Life Cycle Assessment", "Descriptive",
    "Describe the LCA scope: life cycle stages included (A1-A5, B, C, D) and building elements covered.",
    "LCA scope and boundary documentation.")
q += 1

row = add_question(ws24, row, f"{ref_base}.{q}", "Upfront Carbon Reduction", "Minimum Expectation",
    "Life Cycle Assessment", "Data",
    "State the total upfront carbon (kgCO2e) and the upfront carbon intensity (kgCO2e/m² GFA).",
    "Upfront carbon benchmarking across building types.")
q += 1

row = add_level_header(ws24, row, "Credit Achievement (1-3 points)")

row = add_criteria_header(ws24, row, "Upfront Carbon Reduction")

row = add_question(ws24, row, f"{ref_base}.{q}", "Upfront Carbon Reduction", "Credit Achievement",
    "Upfront Carbon Reduction", "Data",
    "State the percentage reduction in upfront carbon compared to the benchmark or reference building.",
    "Embodied carbon reduction quantification.")
q += 1

row = add_question(ws24, row, f"{ref_base}.{q}", "Upfront Carbon Reduction", "Credit Achievement",
    "Upfront Carbon Reduction", "Descriptive",
    "Describe the key strategies used to reduce upfront carbon (material substitution, optimised design, recycled content, local sourcing).",
    "Embodied carbon reduction strategy cataloguing.")
q += 1

row = add_question(ws24, row, f"{ref_base}.{q}", "Upfront Carbon Reduction", "Credit Achievement",
    "Upfront Carbon Reduction", "Data",
    "Identify the top 5 materials by embodied carbon contribution and the reduction strategies applied to each.",
    "Material-level embodied carbon data.")
q += 1

row = add_level_header(ws24, row, "Exceptional Performance")

row = add_question(ws24, row, f"{ref_base}.{q}", "Upfront Carbon Reduction", "Exceptional Performance",
    "Leading Reduction", "Data",
    "State the percentage reduction achieved for Exceptional Performance.",
    "Leading embodied carbon reduction benchmarking.")
q += 1

apply_dropdowns(ws24)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 25: UPFRONT CARBON COMPENSATION
# ════════════════════════════════════════════════════════════════════════════
ws25 = wb.create_sheet()
row = setup_sheet(ws25, "Upfront Carbon Compensation")
row = add_credit_header(ws25, row, "Upfront Carbon Compensation",
    "Residual upfront carbon is compensated through offsets.")
ref_base = "UCC"
q = 1

row = add_level_header(ws25, row, "Credit Achievement (1 point)")

row = add_criteria_header(ws25, row, "Carbon Offset Procurement")

row = add_question(ws25, row, f"{ref_base}.{q}", "Upfront Carbon Compensation", "Credit Achievement",
    "Carbon Offset Procurement", "Data",
    "State the total upfront carbon to be offset (tCO2e).",
    "Carbon offset volume benchmarking.")
q += 1

row = add_question(ws25, row, f"{ref_base}.{q}", "Upfront Carbon Compensation", "Credit Achievement",
    "Carbon Offset Procurement", "Data",
    "State the percentage of upfront carbon being offset (minimum threshold for credit).",
    "Offset coverage percentage tracking.")
q += 1

row = add_question(ws25, row, f"{ref_base}.{q}", "Upfront Carbon Compensation", "Credit Achievement",
    "Carbon Offset Procurement", "Descriptive",
    "Identify the offset type(s) purchased and the certification standard (Gold Standard, VCS, ACCUs, etc.).",
    "Carbon offset type and standard tracking.")
q += 1

row = add_question(ws25, row, f"{ref_base}.{q}", "Upfront Carbon Compensation", "Credit Achievement",
    "Carbon Offset Procurement", "Descriptive",
    "Describe the offset projects supported (location, project type, co-benefits).",
    "Offset project characteristics documentation.")
q += 1

row = add_question(ws25, row, f"{ref_base}.{q}", "Upfront Carbon Compensation", "Credit Achievement",
    "Carbon Offset Procurement", "Condition (Y/N)",
    "Have the offsets been retired or will they be retired within 12 months of practical completion?",
    "Offset retirement timing tracking.")
q += 1

apply_dropdowns(ws25)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 26: REFRIGERANT SYSTEMS IMPACTS
# ════════════════════════════════════════════════════════════════════════════
ws26 = wb.create_sheet()
row = setup_sheet(ws26, "Refrigerant Systems Impacts")
row = add_credit_header(ws26, row, "Refrigerant Systems Impacts",
    "The building minimises refrigerant-related environmental impacts.")
ref_base = "RSI"
q = 1

row = add_level_header(ws26, row, "Minimum Expectation (Nil points)")

row = add_criteria_header(ws26, row, "Refrigerant Management")

row = add_question(ws26, row, f"{ref_base}.{q}", "Refrigerant Systems Impacts", "Minimum Expectation",
    "Refrigerant Management", "Descriptive",
    "List all refrigerant-containing systems in the building (HVAC, refrigeration, fire suppression) and the refrigerant type in each.",
    "Refrigerant inventory across building systems.")
q += 1

row = add_question(ws26, row, f"{ref_base}.{q}", "Refrigerant Systems Impacts", "Minimum Expectation",
    "Refrigerant Management", "Data",
    "State the total refrigerant charge (kg) and the Global Warming Potential (GWP) of each refrigerant used.",
    "Refrigerant charge and GWP benchmarking.")
q += 1

row = add_question(ws26, row, f"{ref_base}.{q}", "Refrigerant Systems Impacts", "Minimum Expectation",
    "Refrigerant Management", "Descriptive",
    "Describe leak detection and containment measures for refrigerant systems.",
    "Refrigerant leak prevention strategies.")
q += 1

row = add_level_header(ws26, row, "Credit Achievement (1 point)")

row = add_criteria_header(ws26, row, "Low-GWP Refrigerants")

row = add_question(ws26, row, f"{ref_base}.{q}", "Refrigerant Systems Impacts", "Credit Achievement",
    "Low-GWP Refrigerants", "Condition (Y/N)",
    "Do all refrigerant systems use refrigerants with GWP below the threshold (e.g. GWP < 750)?",
    "Low-GWP refrigerant adoption tracking.")
q += 1

row = add_question(ws26, row, f"{ref_base}.{q}", "Refrigerant Systems Impacts", "Credit Achievement",
    "Low-GWP Refrigerants", "Descriptive",
    "List the low-GWP refrigerants used and their applications (e.g. R-32, R-290, R-744).",
    "Low-GWP refrigerant type cataloguing.")
q += 1

row = add_question(ws26, row, f"{ref_base}.{q}", "Refrigerant Systems Impacts", "Credit Achievement",
    "Low-GWP Refrigerants", "Data",
    "Calculate the Total Equivalent Warming Impact (TEWI) or Life Cycle Climate Performance (LCCP) for refrigerant systems.",
    "Refrigerant system climate impact benchmarking.")
q += 1

apply_dropdowns(ws26)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 27: LOW-EMISSIONS TRANSPORT
# ════════════════════════════════════════════════════════════════════════════
ws27 = wb.create_sheet()
row = setup_sheet(ws27, "Low-Emissions Transport")
row = add_credit_header(ws27, row, "Low-Emissions Transport",
    "The building supports transition to low-emissions transport.")
ref_base = "LET"
q = 1

row = add_level_header(ws27, row, "Credit Achievement (1 point)")

row = add_criteria_header(ws27, row, "Electric Vehicle Infrastructure")

row = add_question(ws27, row, f"{ref_base}.{q}", "Low-Emissions Transport", "Credit Achievement",
    "Electric Vehicle Infrastructure", "Data",
    "State the total number of car parking spaces and the number with EV charging capability.",
    "EV charging provision rate benchmarking.")
q += 1

row = add_question(ws27, row, f"{ref_base}.{q}", "Low-Emissions Transport", "Credit Achievement",
    "Electric Vehicle Infrastructure", "Data",
    "State the percentage of car parking spaces that are EV-ready (conduit and capacity for future chargers).",
    "EV-ready infrastructure benchmarking.")
q += 1

row = add_question(ws27, row, f"{ref_base}.{q}", "Low-Emissions Transport", "Credit Achievement",
    "Electric Vehicle Infrastructure", "Descriptive",
    "Describe the EV charger types provided (Level 2, DC fast charging) and their power ratings.",
    "EV charger type and capacity documentation.")
q += 1

row = add_criteria_header(ws27, row, "Active Transport Support")

row = add_question(ws27, row, f"{ref_base}.{q}", "Low-Emissions Transport", "Credit Achievement",
    "Active Transport Support", "Data",
    "State the number of secure bicycle parking spaces provided relative to building occupancy.",
    "Bicycle parking provision benchmarking.")
q += 1

row = add_question(ws27, row, f"{ref_base}.{q}", "Low-Emissions Transport", "Credit Achievement",
    "Active Transport Support", "Condition (Y/N)",
    "Are e-bike charging facilities provided?",
    "E-bike charging provision tracking.")
q += 1

apply_dropdowns(ws27)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 28: DESIGN FOR CIRCULARITY
# ════════════════════════════════════════════════════════════════════════════
ws28 = wb.create_sheet()
row = setup_sheet(ws28, "Design for Circularity")
row = add_credit_header(ws28, row, "Design for Circularity",
    "The building is designed for adaptability, disassembly, and material recovery.")
ref_base = "DC"
q = 1

row = add_level_header(ws28, row, "Credit Achievement (1 point)")

row = add_criteria_header(ws28, row, "Design for Adaptability")

row = add_question(ws28, row, f"{ref_base}.{q}", "Design for Circularity", "Credit Achievement",
    "Design for Adaptability", "Descriptive",
    "Describe how the building design allows for future adaptation (flexible floor plates, modular systems, accessible services).",
    "Adaptable design feature cataloguing.")
q += 1

row = add_question(ws28, row, f"{ref_base}.{q}", "Design for Circularity", "Credit Achievement",
    "Design for Adaptability", "Condition (Y/N)",
    "Can the building accommodate a change of use without major structural modifications?",
    "Building adaptability for use change tracking.")
q += 1

row = add_criteria_header(ws28, row, "Design for Disassembly")

row = add_question(ws28, row, f"{ref_base}.{q}", "Design for Circularity", "Credit Achievement",
    "Design for Disassembly", "Descriptive",
    "Describe how building elements are designed for future disassembly and reuse (mechanical connections, material passports, reversible finishes).",
    "Design for disassembly strategies.")
q += 1

row = add_question(ws28, row, f"{ref_base}.{q}", "Design for Circularity", "Credit Achievement",
    "Design for Disassembly", "Condition (Y/N)",
    "Has a material passport or asset register been created for key building components?",
    "Material passport adoption tracking.")
q += 1

row = add_question(ws28, row, f"{ref_base}.{q}", "Design for Circularity", "Credit Achievement",
    "Design for Disassembly", "Data",
    "Estimate the percentage of building materials (by mass) that could be recovered for reuse at end of life.",
    "Material recovery potential benchmarking.")
q += 1

row = add_criteria_header(ws28, row, "Reused or Recycled Content")

row = add_question(ws28, row, f"{ref_base}.{q}", "Design for Circularity", "Credit Achievement",
    "Reused or Recycled Content", "Data",
    "State the percentage of materials (by cost or mass) that are reused or contain recycled content.",
    "Reused/recycled content benchmarking.")
q += 1

row = add_question(ws28, row, f"{ref_base}.{q}", "Design for Circularity", "Credit Achievement",
    "Reused or Recycled Content", "Descriptive",
    "List key materials with reused or recycled content and their sources.",
    "Reused/recycled material sourcing documentation.")
q += 1

apply_dropdowns(ws28)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 29: WATER USE
# ════════════════════════════════════════════════════════════════════════════
ws29 = wb.create_sheet()
row = setup_sheet(ws29, "Water Use")
row = add_credit_header(ws29, row, "Water Use",
    "The building minimises potable water consumption.")
ref_base = "WU"
q = 1

row = add_level_header(ws29, row, "Minimum Expectation (Nil points)")

row = add_criteria_header(ws29, row, "Water Efficient Fixtures")

row = add_question(ws29, row, f"{ref_base}.{q}", "Water Use", "Minimum Expectation",
    "Water Efficient Fixtures", "Descriptive",
    "List the WELS ratings achieved for sanitary fixtures (toilets, taps, showers, urinals).",
    "WELS rating compliance documentation.")
q += 1

row = add_question(ws29, row, f"{ref_base}.{q}", "Water Use", "Minimum Expectation",
    "Water Efficient Fixtures", "Data",
    "State the predicted annual potable water consumption (kL/year) and water use intensity (L/m²/year or L/person/year).",
    "Water use intensity benchmarking.")
q += 1

row = add_level_header(ws29, row, "Credit Achievement (1-2 points)")

row = add_criteria_header(ws29, row, "Enhanced Water Efficiency")

row = add_question(ws29, row, f"{ref_base}.{q}", "Water Use", "Credit Achievement",
    "Enhanced Water Efficiency", "Data",
    "State the percentage reduction in potable water use compared to the reference case.",
    "Potable water reduction quantification.")
q += 1

row = add_question(ws29, row, f"{ref_base}.{q}", "Water Use", "Credit Achievement",
    "Enhanced Water Efficiency", "Descriptive",
    "Describe water efficiency measures beyond minimum fixtures (rainwater harvesting, greywater recycling, efficient irrigation, cooling tower optimisation).",
    "Water efficiency strategy cataloguing.")
q += 1

row = add_criteria_header(ws29, row, "Alternative Water Sources")

row = add_question(ws29, row, f"{ref_base}.{q}", "Water Use", "Credit Achievement",
    "Alternative Water Sources", "Condition (Y/N)",
    "Is rainwater harvesting installed?",
    "Rainwater harvesting adoption tracking.")
q += 1

row = add_question(ws29, row, f"{ref_base}.{q}", "Water Use", "Credit Achievement",
    "Alternative Water Sources", "Data",
    "State the rainwater tank capacity (kL) and estimated annual rainwater capture (kL/year).",
    "Rainwater harvesting capacity benchmarking.")
q += 1

row = add_question(ws29, row, f"{ref_base}.{q}", "Water Use", "Credit Achievement",
    "Alternative Water Sources", "Condition (Y/N)",
    "Is greywater or blackwater recycling installed?",
    "Water recycling adoption tracking.")
q += 1

row = add_question(ws29, row, f"{ref_base}.{q}", "Water Use", "Credit Achievement",
    "Alternative Water Sources", "Data",
    "If yes, state the recycled water treatment capacity and estimated annual volume recycled (kL/year).",
    "Water recycling capacity benchmarking.")
q += 1

apply_dropdowns(ws29)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 30: MOVEMENT AND PLACE
# ════════════════════════════════════════════════════════════════════════════
ws30 = wb.create_sheet()
row = setup_sheet(ws30, "Movement and Place")
row = add_credit_header(ws30, row, "Movement and Place",
    "The building supports sustainable transport and enhances place connectivity.")
ref_base = "MP"
q = 1

row = add_level_header(ws30, row, "Credit Achievement (1 point)")

row = add_criteria_header(ws30, row, "Public Transport Access")

row = add_question(ws30, row, f"{ref_base}.{q}", "Movement and Place", "Credit Achievement",
    "Public Transport Access", "Data",
    "State the distance to nearest public transport stops (bus, train, tram, ferry) in metres.",
    "Public transport proximity benchmarking.")
q += 1

row = add_question(ws30, row, f"{ref_base}.{q}", "Movement and Place", "Credit Achievement",
    "Public Transport Access", "Data",
    "State the frequency of public transport services during peak periods.",
    "Public transport service frequency documentation.")
q += 1

row = add_criteria_header(ws30, row, "Pedestrian and Cyclist Connectivity")

row = add_question(ws30, row, f"{ref_base}.{q}", "Movement and Place", "Credit Achievement",
    "Pedestrian and Cyclist Connectivity", "Descriptive",
    "Describe the pedestrian and cyclist connections from the building to surrounding amenities and transport.",
    "Active transport connectivity documentation.")
q += 1

row = add_question(ws30, row, f"{ref_base}.{q}", "Movement and Place", "Credit Achievement",
    "Pedestrian and Cyclist Connectivity", "Data",
    "State the Walk Score or equivalent walkability index for the location.",
    "Walkability benchmarking.")
q += 1

row = add_criteria_header(ws30, row, "Reduced Car Dependency")

row = add_question(ws30, row, f"{ref_base}.{q}", "Movement and Place", "Credit Achievement",
    "Reduced Car Dependency", "Data",
    "State the car parking provision rate (spaces per dwelling or per 100m² GFA) compared to maximum allowable.",
    "Car parking provision benchmarking.")
q += 1

row = add_question(ws30, row, f"{ref_base}.{q}", "Movement and Place", "Credit Achievement",
    "Reduced Car Dependency", "Condition (Y/N)",
    "Is car parking provision below the maximum allowable rate?",
    "Reduced car parking tracking.")
q += 1

apply_dropdowns(ws30)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 31: ENJOYABLE PLACES
# ════════════════════════════════════════════════════════════════════════════
ws31 = wb.create_sheet()
row = setup_sheet(ws31, "Enjoyable Places")
row = add_credit_header(ws31, row, "Enjoyable Places",
    "The building creates high-quality public realm and occupant spaces.")
ref_base = "EP"
q = 1

row = add_level_header(ws31, row, "Credit Achievement (1 point)")

row = add_criteria_header(ws31, row, "Public Realm Quality")

row = add_question(ws31, row, f"{ref_base}.{q}", "Enjoyable Places", "Credit Achievement",
    "Public Realm Quality", "Descriptive",
    "Describe how the building design contributes to the quality of the public realm (street activation, weather protection, seating, landscaping).",
    "Public realm contribution strategies.")
q += 1

row = add_question(ws31, row, f"{ref_base}.{q}", "Enjoyable Places", "Credit Achievement",
    "Public Realm Quality", "Condition (Y/N)",
    "Does the building provide publicly accessible ground floor uses or spaces?",
    "Public accessibility at ground level tracking.")
q += 1

row = add_criteria_header(ws31, row, "Communal Spaces")

row = add_question(ws31, row, f"{ref_base}.{q}", "Enjoyable Places", "Credit Achievement",
    "Communal Spaces", "Data",
    "State the area of communal spaces provided for building occupants (m²) and the types of spaces.",
    "Communal space provision benchmarking.")
q += 1

row = add_question(ws31, row, f"{ref_base}.{q}", "Enjoyable Places", "Credit Achievement",
    "Communal Spaces", "Descriptive",
    "Describe the communal facilities provided (rooftop terraces, gardens, lounges, BBQ areas, gyms).",
    "Communal facility type documentation.")
q += 1

row = add_criteria_header(ws31, row, "Safety and Security")

row = add_question(ws31, row, f"{ref_base}.{q}", "Enjoyable Places", "Credit Achievement",
    "Safety and Security", "Descriptive",
    "Describe how the building design promotes safety through natural surveillance, lighting, and clear sightlines.",
    "Crime Prevention Through Environmental Design (CPTED) strategies.")
q += 1

apply_dropdowns(ws31)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 32: CONTRIBUTION TO PLACE
# ════════════════════════════════════════════════════════════════════════════
ws32 = wb.create_sheet()
row = setup_sheet(ws32, "Contribution to Place")
row = add_credit_header(ws32, row, "Contribution to Place",
    "The building makes a positive contribution to its local context and community.")
ref_base = "CP"
q = 1

row = add_level_header(ws32, row, "Credit Achievement (1 point)")

row = add_criteria_header(ws32, row, "Community Engagement")

row = add_question(ws32, row, f"{ref_base}.{q}", "Contribution to Place", "Credit Achievement",
    "Community Engagement", "Condition (Y/N)",
    "Was community engagement undertaken during the design process?",
    "Community engagement in building design tracking.")
q += 1

row = add_question(ws32, row, f"{ref_base}.{q}", "Contribution to Place", "Credit Achievement",
    "Community Engagement", "Descriptive",
    "Describe the community engagement process and how feedback influenced the design.",
    "Community engagement methodology documentation.")
q += 1

row = add_criteria_header(ws32, row, "Local Character Response")

row = add_question(ws32, row, f"{ref_base}.{q}", "Contribution to Place", "Credit Achievement",
    "Local Character Response", "Descriptive",
    "Describe how the building design responds to local character, context, and urban design guidelines.",
    "Contextual design response documentation.")
q += 1

row = add_criteria_header(ws32, row, "Local Economic Contribution")

row = add_question(ws32, row, f"{ref_base}.{q}", "Contribution to Place", "Credit Achievement",
    "Local Economic Contribution", "Descriptive",
    "Describe how the project contributes to the local economy (local procurement, employment, businesses).",
    "Local economic impact documentation.")
q += 1

row = add_question(ws32, row, f"{ref_base}.{q}", "Contribution to Place", "Credit Achievement",
    "Local Economic Contribution", "Data",
    "State the percentage of construction contract value spent with local suppliers/contractors.",
    "Local procurement benchmarking.")
q += 1

apply_dropdowns(ws32)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 33: CULTURE, HERITAGE AND IDENTITY
# ════════════════════════════════════════════════════════════════════════════
ws33 = wb.create_sheet()
row = setup_sheet(ws33, "Culture Heritage Identity")
row = add_credit_header(ws33, row, "Culture, Heritage and Identity",
    "The building respects and celebrates cultural heritage and identity.")
ref_base = "CHI"
q = 1

row = add_level_header(ws33, row, "Credit Achievement (1 point)")

row = add_criteria_header(ws33, row, "Heritage Response")

row = add_question(ws33, row, f"{ref_base}.{q}", "Culture, Heritage and Identity", "Credit Achievement",
    "Heritage Response", "Condition (Y/N)",
    "Is the site or surrounding area subject to heritage controls or significance?",
    "Heritage context identification.")
q += 1

row = add_question(ws33, row, f"{ref_base}.{q}", "Culture, Heritage and Identity", "Credit Achievement",
    "Heritage Response", "Descriptive",
    "Describe how the building design responds to and respects heritage significance.",
    "Heritage response design strategies.")
q += 1

row = add_criteria_header(ws33, row, "Cultural Identity")

row = add_question(ws33, row, f"{ref_base}.{q}", "Culture, Heritage and Identity", "Credit Achievement",
    "Cultural Identity", "Descriptive",
    "Describe how the building design reflects or celebrates local cultural identity and stories.",
    "Cultural identity expression in building design.")
q += 1

row = add_question(ws33, row, f"{ref_base}.{q}", "Culture, Heritage and Identity", "Credit Achievement",
    "Cultural Identity", "Descriptive",
    "Describe any art, interpretation, or placemaking elements that communicate cultural narratives.",
    "Cultural interpretation documentation.")
q += 1

apply_dropdowns(ws33)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 34: INCLUSIVE CONSTRUCTION PRACTICES
# ════════════════════════════════════════════════════════════════════════════
ws34 = wb.create_sheet()
row = setup_sheet(ws34, "Inclusive Construction")
row = add_credit_header(ws34, row, "Inclusive Construction Practices",
    "Construction practices promote diversity, inclusion, and worker wellbeing.")
ref_base = "ICP"
q = 1

row = add_level_header(ws34, row, "Minimum Expectation (Nil points)")

row = add_criteria_header(ws34, row, "Fair Work Practices")

row = add_question(ws34, row, f"{ref_base}.{q}", "Inclusive Construction Practices", "Minimum Expectation",
    "Fair Work Practices", "Condition (Y/N)",
    "Do all contractors on site comply with relevant industrial awards and workplace laws?",
    "Industrial compliance tracking in construction.")
q += 1

row = add_question(ws34, row, f"{ref_base}.{q}", "Inclusive Construction Practices", "Minimum Expectation",
    "Fair Work Practices", "Descriptive",
    "Describe how compliance with fair work practices is verified for the head contractor and subcontractors.",
    "Fair work verification practices.")
q += 1

row = add_level_header(ws34, row, "Credit Achievement (1 point)")

row = add_criteria_header(ws34, row, "Workforce Diversity")

row = add_question(ws34, row, f"{ref_base}.{q}", "Inclusive Construction Practices", "Credit Achievement",
    "Workforce Diversity", "Data",
    "State the percentage of women employed on the construction project.",
    "Gender diversity in construction benchmarking.")
q += 1

row = add_question(ws34, row, f"{ref_base}.{q}", "Inclusive Construction Practices", "Credit Achievement",
    "Workforce Diversity", "Data",
    "State the percentage of apprentices and trainees employed on the project.",
    "Apprentice/trainee employment benchmarking.")
q += 1

row = add_question(ws34, row, f"{ref_base}.{q}", "Inclusive Construction Practices", "Credit Achievement",
    "Workforce Diversity", "Descriptive",
    "Describe initiatives to promote workforce diversity and inclusion on the project.",
    "Diversity and inclusion initiative documentation.")
q += 1

row = add_criteria_header(ws34, row, "Worker Wellbeing")

row = add_question(ws34, row, f"{ref_base}.{q}", "Inclusive Construction Practices", "Credit Achievement",
    "Worker Wellbeing", "Descriptive",
    "Describe mental health and wellbeing initiatives implemented on the construction site.",
    "Construction worker wellbeing program documentation.")
q += 1

row = add_question(ws34, row, f"{ref_base}.{q}", "Inclusive Construction Practices", "Credit Achievement",
    "Worker Wellbeing", "Condition (Y/N)",
    "Are site amenities provided beyond minimum WHS requirements?",
    "Enhanced site amenities tracking.")
q += 1

apply_dropdowns(ws34)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 35: FIRST NATIONS INCLUSION
# ════════════════════════════════════════════════════════════════════════════
ws35 = wb.create_sheet()
row = setup_sheet(ws35, "First Nations Inclusion")
row = add_credit_header(ws35, row, "First Nations Inclusion",
    "The project supports First Nations peoples and acknowledges their connection to Country.")
ref_base = "FNI"
q = 1

row = add_level_header(ws35, row, "Credit Achievement (1 point)")

row = add_criteria_header(ws35, row, "Acknowledgement of Country")

row = add_question(ws35, row, f"{ref_base}.{q}", "First Nations Inclusion", "Credit Achievement",
    "Acknowledgement of Country", "Condition (Y/N)",
    "Has engagement with Traditional Custodians occurred for the project?",
    "Traditional Custodian engagement tracking.")
q += 1

row = add_question(ws35, row, f"{ref_base}.{q}", "First Nations Inclusion", "Credit Achievement",
    "Acknowledgement of Country", "Descriptive",
    "Describe how the project acknowledges and celebrates the Traditional Custodians' connection to Country.",
    "Acknowledgement of Country implementation documentation.")
q += 1

row = add_question(ws35, row, f"{ref_base}.{q}", "First Nations Inclusion", "Credit Achievement",
    "Acknowledgement of Country", "Descriptive",
    "Describe any design elements that reflect First Nations culture, knowledge, or stories (with appropriate permissions).",
    "First Nations cultural expression in design.")
q += 1

row = add_criteria_header(ws35, row, "First Nations Employment and Procurement")

row = add_question(ws35, row, f"{ref_base}.{q}", "First Nations Inclusion", "Credit Achievement",
    "First Nations Employment and Procurement", "Data",
    "State the percentage of contract value with First Nations-owned businesses (certified Supply Nation or equivalent).",
    "First Nations procurement benchmarking.")
q += 1

row = add_question(ws35, row, f"{ref_base}.{q}", "First Nations Inclusion", "Credit Achievement",
    "First Nations Employment and Procurement", "Data",
    "State the number or percentage of First Nations people employed on the project.",
    "First Nations employment benchmarking.")
q += 1

row = add_question(ws35, row, f"{ref_base}.{q}", "First Nations Inclusion", "Credit Achievement",
    "First Nations Employment and Procurement", "Descriptive",
    "Describe initiatives to support First Nations employment, training, or business development on the project.",
    "First Nations economic participation initiatives.")
q += 1

apply_dropdowns(ws35)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 36: PROCUREMENT AND WORKFORCE INCLUSION
# ════════════════════════════════════════════════════════════════════════════
ws36 = wb.create_sheet()
row = setup_sheet(ws36, "Procurement Workforce Inclusion")
row = add_credit_header(ws36, row, "Procurement and Workforce Inclusion",
    "Procurement and employment practices support social inclusion.")
ref_base = "PWI"
q = 1

row = add_level_header(ws36, row, "Credit Achievement (1 point)")

row = add_criteria_header(ws36, row, "Social Procurement")

row = add_question(ws36, row, f"{ref_base}.{q}", "Procurement and Workforce Inclusion", "Credit Achievement",
    "Social Procurement", "Descriptive",
    "Describe the social procurement policy or targets for the project.",
    "Social procurement policy documentation.")
q += 1

row = add_question(ws36, row, f"{ref_base}.{q}", "Procurement and Workforce Inclusion", "Credit Achievement",
    "Social Procurement", "Data",
    "State the percentage of contract value with social enterprises or certified B Corps.",
    "Social enterprise procurement benchmarking.")
q += 1

row = add_question(ws36, row, f"{ref_base}.{q}", "Procurement and Workforce Inclusion", "Credit Achievement",
    "Social Procurement", "Descriptive",
    "List the social enterprises or organisations engaged and the services/products they provided.",
    "Social enterprise engagement documentation.")
q += 1

row = add_criteria_header(ws36, row, "Priority Group Employment")

row = add_question(ws36, row, f"{ref_base}.{q}", "Procurement and Workforce Inclusion", "Credit Achievement",
    "Priority Group Employment", "Data",
    "State the number of people from priority groups employed on the project (long-term unemployed, people with disability, refugees, ex-offenders).",
    "Priority group employment benchmarking.")
q += 1

row = add_question(ws36, row, f"{ref_base}.{q}", "Procurement and Workforce Inclusion", "Credit Achievement",
    "Priority Group Employment", "Descriptive",
    "Describe partnerships with employment services or social enterprises to facilitate priority group employment.",
    "Priority group employment pathway documentation.")
q += 1

apply_dropdowns(ws36)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 37: DESIGN FOR EQUITY
# ════════════════════════════════════════════════════════════════════════════
ws37 = wb.create_sheet()
row = setup_sheet(ws37, "Design for Equity")
row = add_credit_header(ws37, row, "Design for Equity",
    "The building is designed for equitable access and use by all people.")
ref_base = "DE"
q = 1

row = add_level_header(ws37, row, "Minimum Expectation (Nil points)")

row = add_criteria_header(ws37, row, "Universal Design Principles")

row = add_question(ws37, row, f"{ref_base}.{q}", "Design for Equity", "Minimum Expectation",
    "Universal Design Principles", "Condition (Y/N)",
    "Has an Access Consultant been engaged for the project?",
    "Access consultant engagement tracking.")
q += 1

row = add_question(ws37, row, f"{ref_base}.{q}", "Design for Equity", "Minimum Expectation",
    "Universal Design Principles", "Descriptive",
    "Describe how the building design applies universal design principles beyond minimum DDA/BCA requirements.",
    "Universal design implementation documentation.")
q += 1

row = add_level_header(ws37, row, "Credit Achievement (1 point)")

row = add_criteria_header(ws37, row, "Enhanced Accessibility")

row = add_question(ws37, row, f"{ref_base}.{q}", "Design for Equity", "Credit Achievement",
    "Enhanced Accessibility", "Descriptive",
    "Describe enhanced accessibility features beyond code requirements (tactile indicators, hearing loops, accessible way-finding, quiet spaces).",
    "Enhanced accessibility feature cataloguing.")
q += 1

row = add_question(ws37, row, f"{ref_base}.{q}", "Design for Equity", "Credit Achievement",
    "Enhanced Accessibility", "Condition (Y/N)",
    "Does the building achieve certification under an accessibility standard (e.g. Liveable Housing Design Guidelines Silver/Gold/Platinum)?",
    "Accessibility certification tracking.")
q += 1

row = add_question(ws37, row, f"{ref_base}.{q}", "Design for Equity", "Credit Achievement",
    "Enhanced Accessibility", "Descriptive",
    "Identify the accessibility certification or standard achieved and the level.",
    "")
q += 1

row = add_criteria_header(ws37, row, "Inclusive Facilities")

row = add_question(ws37, row, f"{ref_base}.{q}", "Design for Equity", "Credit Achievement",
    "Inclusive Facilities", "Condition (Y/N)",
    "Are all-gender bathrooms provided?",
    "All-gender bathroom provision tracking.")
q += 1

row = add_question(ws37, row, f"{ref_base}.{q}", "Design for Equity", "Credit Achievement",
    "Inclusive Facilities", "Condition (Y/N)",
    "Are Changing Places or accessible adult change facilities provided?",
    "Changing Places facility provision tracking.")
q += 1

row = add_question(ws37, row, f"{ref_base}.{q}", "Design for Equity", "Credit Achievement",
    "Inclusive Facilities", "Descriptive",
    "Describe other inclusive facilities provided (parents rooms, prayer rooms, sensory rooms).",
    "Inclusive facility provision documentation.")
q += 1

apply_dropdowns(ws37)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 38: IMPACTS TO NATURE
# ════════════════════════════════════════════════════════════════════════════
ws38 = wb.create_sheet()
row = setup_sheet(ws38, "Impacts to Nature")
row = add_credit_header(ws38, row, "Impacts to Nature",
    "The project minimises negative impacts on nature and ecological systems.")
ref_base = "IN"
q = 1

row = add_level_header(ws38, row, "Minimum Expectation (Nil points)")

row = add_criteria_header(ws38, row, "Ecological Assessment")

row = add_question(ws38, row, f"{ref_base}.{q}", "Impacts to Nature", "Minimum Expectation",
    "Ecological Assessment", "Condition (Y/N)",
    "Was an ecological assessment undertaken for the site?",
    "Ecological assessment adoption tracking.")
q += 1

row = add_question(ws38, row, f"{ref_base}.{q}", "Impacts to Nature", "Minimum Expectation",
    "Ecological Assessment", "Descriptive",
    "Identify the ecologist and summarise the ecological values found on or adjacent to the site.",
    "Site ecological value documentation.")
q += 1

row = add_question(ws38, row, f"{ref_base}.{q}", "Impacts to Nature", "Minimum Expectation",
    "Ecological Assessment", "Descriptive",
    "Describe any threatened species, endangered ecological communities, or significant habitat identified.",
    "Threatened species/habitat identification.")
q += 1

row = add_criteria_header(ws38, row, "Impact Mitigation")

row = add_question(ws38, row, f"{ref_base}.{q}", "Impacts to Nature", "Minimum Expectation",
    "Impact Mitigation", "Descriptive",
    "Describe the mitigation hierarchy applied: avoid, minimise, restore, offset ecological impacts.",
    "Ecological mitigation hierarchy application.")
q += 1

row = add_question(ws38, row, f"{ref_base}.{q}", "Impacts to Nature", "Minimum Expectation",
    "Impact Mitigation", "Data",
    "State the area of native vegetation cleared (m²) and the area retained or protected.",
    "Vegetation clearing and retention benchmarking.")
q += 1

row = add_level_header(ws38, row, "Credit Achievement (1 point)")

row = add_criteria_header(ws38, row, "Net Positive Impact")

row = add_question(ws38, row, f"{ref_base}.{q}", "Impacts to Nature", "Credit Achievement",
    "Net Positive Impact", "Condition (Y/N)",
    "Does the project achieve a net positive outcome for biodiversity?",
    "Net positive biodiversity outcome tracking.")
q += 1

row = add_question(ws38, row, f"{ref_base}.{q}", "Impacts to Nature", "Credit Achievement",
    "Net Positive Impact", "Descriptive",
    "Describe how net positive biodiversity is demonstrated (habitat created exceeds habitat lost, offsets, etc.).",
    "Net positive biodiversity methodology.")
q += 1

apply_dropdowns(ws38)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 39: BIODIVERSITY ENHANCEMENT
# ════════════════════════════════════════════════════════════════════════════
ws39 = wb.create_sheet()
row = setup_sheet(ws39, "Biodiversity Enhancement")
row = add_credit_header(ws39, row, "Biodiversity Enhancement",
    "The project enhances biodiversity through design and landscaping.")
ref_base = "BE"
q = 1

row = add_level_header(ws39, row, "Credit Achievement (1 point)")

row = add_criteria_header(ws39, row, "Native Planting")

row = add_question(ws39, row, f"{ref_base}.{q}", "Biodiversity Enhancement", "Credit Achievement",
    "Native Planting", "Data",
    "State the percentage of landscaping using locally native species.",
    "Native planting percentage benchmarking.")
q += 1

row = add_question(ws39, row, f"{ref_base}.{q}", "Biodiversity Enhancement", "Credit Achievement",
    "Native Planting", "Descriptive",
    "List the native species planted and their ecological function (food source, habitat, pollinator support).",
    "Native species selection documentation.")
q += 1

row = add_criteria_header(ws39, row, "Habitat Creation")

row = add_question(ws39, row, f"{ref_base}.{q}", "Biodiversity Enhancement", "Credit Achievement",
    "Habitat Creation", "Descriptive",
    "Describe habitat features created (nesting boxes, insect hotels, frog ponds, green roofs, habitat logs).",
    "Created habitat feature cataloguing.")
q += 1

row = add_question(ws39, row, f"{ref_base}.{q}", "Biodiversity Enhancement", "Credit Achievement",
    "Habitat Creation", "Data",
    "State the area of habitat created (m²) including green roofs, walls, and ground-level habitat.",
    "Habitat area benchmarking.")
q += 1

row = add_criteria_header(ws39, row, "Invasive Species Management")

row = add_question(ws39, row, f"{ref_base}.{q}", "Biodiversity Enhancement", "Credit Achievement",
    "Invasive Species Management", "Descriptive",
    "Describe invasive species removal and ongoing management commitments.",
    "Invasive species management documentation.")
q += 1

apply_dropdowns(ws39)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 40: NATURE CONNECTIVITY
# ════════════════════════════════════════════════════════════════════════════
ws40 = wb.create_sheet()
row = setup_sheet(ws40, "Nature Connectivity")
row = add_credit_header(ws40, row, "Nature Connectivity",
    "The project supports ecological connectivity and wildlife movement.")
ref_base = "NC"
q = 1

row = add_level_header(ws40, row, "Credit Achievement (1 point)")

row = add_criteria_header(ws40, row, "Wildlife Corridors")

row = add_question(ws40, row, f"{ref_base}.{q}", "Nature Connectivity", "Credit Achievement",
    "Wildlife Corridors", "Condition (Y/N)",
    "Does the site connect to or enhance existing ecological corridors?",
    "Ecological corridor connection tracking.")
q += 1

row = add_question(ws40, row, f"{ref_base}.{q}", "Nature Connectivity", "Credit Achievement",
    "Wildlife Corridors", "Descriptive",
    "Describe how the landscape design supports wildlife movement and connectivity.",
    "Wildlife connectivity design strategies.")
q += 1

row = add_criteria_header(ws40, row, "Fencing and Barriers")

row = add_question(ws40, row, f"{ref_base}.{q}", "Nature Connectivity", "Credit Achievement",
    "Fencing and Barriers", "Condition (Y/N)",
    "Is wildlife-permeable fencing used where appropriate?",
    "Wildlife-permeable fencing adoption.")
q += 1

row = add_question(ws40, row, f"{ref_base}.{q}", "Nature Connectivity", "Credit Achievement",
    "Fencing and Barriers", "Descriptive",
    "Describe measures to minimise barriers to wildlife movement (fauna underpasses, glider poles, etc.).",
    "Wildlife movement barrier mitigation.")
q += 1

row = add_criteria_header(ws40, row, "Bird-Safe Design")

row = add_question(ws40, row, f"{ref_base}.{q}", "Nature Connectivity", "Credit Achievement",
    "Bird-Safe Design", "Condition (Y/N)",
    "Has bird-safe glazing or design been implemented to reduce bird strike risk?",
    "Bird-safe design adoption tracking.")
q += 1

row = add_question(ws40, row, f"{ref_base}.{q}", "Nature Connectivity", "Credit Achievement",
    "Bird-Safe Design", "Descriptive",
    "Describe bird-safe design measures (fritting, external screens, reduced reflectivity, lighting design).",
    "Bird-safe design measure cataloguing.")
q += 1

apply_dropdowns(ws40)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 41: NATURE STEWARDSHIP
# ════════════════════════════════════════════════════════════════════════════
ws41 = wb.create_sheet()
row = setup_sheet(ws41, "Nature Stewardship")
row = add_credit_header(ws41, row, "Nature Stewardship",
    "Long-term stewardship of ecological values is committed.")
ref_base = "NS"
q = 1

row = add_level_header(ws41, row, "Credit Achievement (1 point)")

row = add_criteria_header(ws41, row, "Ecological Management Plan")

row = add_question(ws41, row, f"{ref_base}.{q}", "Nature Stewardship", "Credit Achievement",
    "Ecological Management Plan", "Condition (Y/N)",
    "Has an ongoing ecological management plan been developed?",
    "Ecological management plan adoption.")
q += 1

row = add_question(ws41, row, f"{ref_base}.{q}", "Nature Stewardship", "Credit Achievement",
    "Ecological Management Plan", "Descriptive",
    "Summarise the ecological management plan including maintenance schedules, monitoring, and responsible parties.",
    "Ecological management plan scope documentation.")
q += 1

row = add_question(ws41, row, f"{ref_base}.{q}", "Nature Stewardship", "Credit Achievement",
    "Ecological Management Plan", "Data",
    "State the duration of the ecological management commitment (years).",
    "Ecological management duration benchmarking.")
q += 1

row = add_criteria_header(ws41, row, "Funding and Governance")

row = add_question(ws41, row, f"{ref_base}.{q}", "Nature Stewardship", "Credit Achievement",
    "Funding and Governance", "Descriptive",
    "Describe the funding mechanism for ongoing ecological management (sinking fund, body corporate levy, etc.).",
    "Ecological management funding mechanisms.")
q += 1

row = add_question(ws41, row, f"{ref_base}.{q}", "Nature Stewardship", "Credit Achievement",
    "Funding and Governance", "Descriptive",
    "Identify who is responsible for ecological management and their qualifications.",
    "Ecological management governance documentation.")
q += 1

apply_dropdowns(ws41)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 42: WATERWAY PROTECTION
# ════════════════════════════════════════════════════════════════════════════
ws42 = wb.create_sheet()
row = setup_sheet(ws42, "Waterway Protection")
row = add_credit_header(ws42, row, "Waterway Protection",
    "The project protects waterway health and aquatic ecosystems.")
ref_base = "WP"
q = 1

row = add_level_header(ws42, row, "Minimum Expectation (Nil points)")

row = add_criteria_header(ws42, row, "Stormwater Management")

row = add_question(ws42, row, f"{ref_base}.{q}", "Waterway Protection", "Minimum Expectation",
    "Stormwater Management", "Descriptive",
    "Describe the stormwater management strategy and how it meets regulatory requirements.",
    "Stormwater management compliance documentation.")
q += 1

row = add_question(ws42, row, f"{ref_base}.{q}", "Waterway Protection", "Minimum Expectation",
    "Stormwater Management", "Data",
    "State the post-development stormwater runoff rate compared to pre-development or regulatory target.",
    "Stormwater runoff rate benchmarking.")
q += 1

row = add_level_header(ws42, row, "Credit Achievement (1 point)")

row = add_criteria_header(ws42, row, "Enhanced Stormwater Treatment")

row = add_question(ws42, row, f"{ref_base}.{q}", "Waterway Protection", "Credit Achievement",
    "Enhanced Stormwater Treatment", "Descriptive",
    "Describe the Water Sensitive Urban Design (WSUD) features implemented (bioretention, permeable paving, rain gardens, wetlands).",
    "WSUD feature cataloguing.")
q += 1

row = add_question(ws42, row, f"{ref_base}.{q}", "Waterway Protection", "Credit Achievement",
    "Enhanced Stormwater Treatment", "Data",
    "State the pollutant reduction targets achieved (Total Suspended Solids, Total Nitrogen, Total Phosphorus, Gross Pollutants).",
    "Stormwater pollutant reduction benchmarking.")
q += 1

row = add_question(ws42, row, f"{ref_base}.{q}", "Waterway Protection", "Credit Achievement",
    "Enhanced Stormwater Treatment", "Data",
    "State the percentage of site area treated by WSUD measures.",
    "WSUD treatment coverage benchmarking.")
q += 1

row = add_criteria_header(ws42, row, "Stream and Riparian Protection")

row = add_question(ws42, row, f"{ref_base}.{q}", "Waterway Protection", "Credit Achievement",
    "Stream and Riparian Protection", "Condition (Y/N)",
    "Is there a waterway or riparian zone on or adjacent to the site?",
    "Waterway/riparian presence identification.")
q += 1

row = add_question(ws42, row, f"{ref_base}.{q}", "Waterway Protection", "Credit Achievement",
    "Stream and Riparian Protection", "Descriptive",
    "If yes, describe protection and enhancement measures for the waterway and riparian vegetation.",
    "Riparian protection measures documentation.")
q += 1

apply_dropdowns(ws42)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 43: MARKET TRANSFORMATION
# ════════════════════════════════════════════════════════════════════════════
ws43 = wb.create_sheet()
row = setup_sheet(ws43, "Market Transformation")
row = add_credit_header(ws43, row, "Market Transformation",
    "The project demonstrates industry leadership and innovation.")
ref_base = "MT"
q = 1

row = add_level_header(ws43, row, "Credit Achievement (1 point)")

row = add_criteria_header(ws43, row, "Industry First or Innovation")

row = add_question(ws43, row, f"{ref_base}.{q}", "Market Transformation", "Credit Achievement",
    "Industry First or Innovation", "Descriptive",
    "Describe the innovation or industry-first initiative being claimed for this credit.",
    "Green building innovation documentation.")
q += 1

row = add_question(ws43, row, f"{ref_base}.{q}", "Market Transformation", "Credit Achievement",
    "Industry First or Innovation", "Descriptive",
    "Explain why this is considered innovative or an industry first (not already common practice).",
    "Innovation justification for market transformation.")
q += 1

row = add_question(ws43, row, f"{ref_base}.{q}", "Market Transformation", "Credit Achievement",
    "Industry First or Innovation", "Descriptive",
    "Describe the potential for the innovation to be replicated across the industry.",
    "Innovation replicability assessment.")
q += 1

row = add_criteria_header(ws43, row, "Knowledge Sharing")

row = add_question(ws43, row, f"{ref_base}.{q}", "Market Transformation", "Credit Achievement",
    "Knowledge Sharing", "Condition (Y/N)",
    "Has the project team committed to sharing learnings from the innovation with the industry?",
    "Knowledge sharing commitment tracking.")
q += 1

row = add_question(ws43, row, f"{ref_base}.{q}", "Market Transformation", "Credit Achievement",
    "Knowledge Sharing", "Descriptive",
    "Describe the knowledge sharing activities planned (case studies, conference presentations, site tours, publications).",
    "Knowledge sharing activity documentation.")
q += 1

apply_dropdowns(ws43)


# ════════════════════════════════════════════════════════════════════════════
# CREDIT 44: LEADERSHIP CHALLENGES
# ════════════════════════════════════════════════════════════════════════════
ws44 = wb.create_sheet()
row = setup_sheet(ws44, "Leadership Challenges")
row = add_credit_header(ws44, row, "Leadership Challenges",
    "The project addresses GBCA Leadership Challenges for additional recognition.")
ref_base = "LC"
q = 1

row = add_level_header(ws44, row, "Credit Achievement (Variable points)")

row = add_criteria_header(ws44, row, "Leadership Challenge Selection")

row = add_question(ws44, row, f"{ref_base}.{q}", "Leadership Challenges", "Credit Achievement",
    "Leadership Challenge Selection", "Descriptive",
    "Identify which GBCA Leadership Challenge(s) the project is addressing.",
    "Leadership Challenge uptake tracking.")
q += 1

row = add_question(ws44, row, f"{ref_base}.{q}", "Leadership Challenges", "Credit Achievement",
    "Leadership Challenge Selection", "Descriptive",
    "Describe how the project meets the requirements of the selected Leadership Challenge(s).",
    "Leadership Challenge compliance documentation.")
q += 1

row = add_question(ws44, row, f"{ref_base}.{q}", "Leadership Challenges", "Credit Achievement",
    "Leadership Challenge Selection", "Data",
    "State the number of Innovation Points being claimed through Leadership Challenges.",
    "Leadership Challenge points tracking.")
q += 1

row = add_question(ws44, row, f"{ref_base}.{q}", "Leadership Challenges", "Credit Achievement",
    "Leadership Challenge Selection", "Descriptive",
    "Describe any additional verification or documentation required for the Leadership Challenge.",
    "Leadership Challenge verification documentation.")
q += 1

apply_dropdowns(ws44)


# ════════════════════════════════════════════════════════════════════════════
# Save
# ════════════════════════════════════════════════════════════════════════════
output_path = "/home/user/submissionforms/Green_Star_Buildings_v1.1_Submission_Questions.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
