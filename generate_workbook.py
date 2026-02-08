#!/usr/bin/env python3
"""
Generate a macro-enabled Excel workbook (.xlsm) from the Green Star Buildings
submission data, with full interactivity via VBA:

  - Dashboard with progress tracking
  - Per-credit sheets with formatted questions, guidance, and input cells
  - Sidebar-style navigation (Dashboard index with hyperlinks)
  - Conditional row visibility (Y/N gateway questions)
  - N/A toggle per credit
  - Review mode (highlight unanswered)
  - Search across all questions
  - Version history on a hidden sheet
  - Dark mode toggle
  - Sheet protection (users can only edit response cells)
"""

import json
import os
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Protection, numbers
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from docx import Document as DocxDocument

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 1: Parse Excel questions
# ═══════════════════════════════════════════════════════════════════════════════
print("Parsing Excel questions...")
wb_src = load_workbook("Green_Star_Buildings_v1.1_Submission_Questions.xlsx")

CATEGORIES = {
    "Responsible": [
        "Industry Development", "Responsible Construction",
        "Verification and Handover", "Responsible Resource Mgmt",
        "Responsible Procurement", "Responsible Structure",
        "Responsible Envelope", "Responsible Systems",
        "Responsible Finishes", "Impacts Disclosure",
    ],
    "Healthy": [
        "Clean Air", "Light Quality", "Acoustic Comfort",
        "Exposure to Toxins", "Amenity and Comfort", "Connection to Nature",
    ],
    "Resilient": [
        "Climate Resilience", "Operations Resilience",
        "Community Resilience", "Heat Resilience", "Grid Resilience",
    ],
    "Positive": [
        "Energy Source", "Energy Use", "Upfront Carbon Reduction",
        "Upfront Carbon Compensation", "Refrigerant Systems Impacts",
        "Low-Emissions Transport", "Design for Circularity", "Water Use",
    ],
    "Places": [
        "Movement and Place", "Enjoyable Places",
        "Contribution to Place", "Culture Heritage Identity",
    ],
    "People": [
        "Inclusive Construction", "First Nations Inclusion",
        "Procurement Workforce Inclusion", "Design for Equity",
    ],
    "Nature": [
        "Impacts to Nature", "Biodiversity Enhancement",
        "Nature Connectivity", "Nature Stewardship", "Waterway Protection",
    ],
    "Leadership": [
        "Market Transformation", "Leadership Challenges",
    ],
}

CATEGORY_COLORS = {
    "Responsible": "1F4E28",
    "Healthy": "1565C0",
    "Resilient": "E65100",
    "Positive": "2E7D32",
    "Places": "6A1B9A",
    "People": "C62828",
    "Nature": "00695C",
    "Leadership": "F57F17",
}

all_credits = []
for sheet_name in wb_src.sheetnames:
    ws = wb_src[sheet_name]
    credit = {
        "sheet_name": sheet_name,
        "title": sheet_name,
        "sections": [],
        "questions": [],
    }
    current_section = None
    for row_idx in range(2, ws.max_row + 1):
        a = ws.cell(row=row_idx, column=1).value
        b = ws.cell(row=row_idx, column=2).value
        e = ws.cell(row=row_idx, column=5).value
        f_val = ws.cell(row=row_idx, column=6).value
        h = ws.cell(row=row_idx, column=8).value
        font = ws.cell(row=row_idx, column=1).font

        if a and not b and not e:
            text = str(a)
            if font.color and hasattr(font.color, 'rgb') and font.color.rgb:
                rgb = str(font.color.rgb)
                if rgb == "00FFFFFF" and font.size == 12:
                    credit["title"] = text
                    continue
                elif rgb == "00FFFFFF" and font.size == 11:
                    current_section = {"type": "level", "title": text, "criteria": []}
                    credit["sections"].append(current_section)
                    continue
                elif "1F4E28" in rgb:
                    if current_section is not None:
                        current_section["criteria"].append({"name": text, "questions": []})
                    continue
            if font.bold and font.size and font.size >= 12:
                credit["title"] = text
                continue
            elif font.bold and font.size and font.size >= 11:
                if font.color and hasattr(font.color, 'rgb') and "1F4E28" in str(font.color.rgb):
                    if current_section is not None:
                        current_section["criteria"].append({"name": text, "questions": []})
                    continue
                current_section = {"type": "level", "title": text, "criteria": []}
                credit["sections"].append(current_section)
                continue

        if e and f_val:
            c_val = ws.cell(row=row_idx, column=3).value
            d_val = ws.cell(row=row_idx, column=4).value
            q = {
                "ref": str(a) if a else "",
                "credit": str(b) if b else "",
                "level": str(c_val) if c_val else "",
                "criteria": str(d_val) if d_val else "",
                "type": str(e) if e else "",
                "question": str(f_val) if f_val else "",
                "data_note": str(h) if h else "",
            }
            credit["questions"].append(q)
            if current_section:
                if not current_section["criteria"]:
                    current_section["criteria"].append({"name": "General", "questions": []})
                current_section["criteria"][-1]["questions"].append(q)

    all_credits.append(credit)

def find_category(sheet_name):
    for cat, sheets in CATEGORIES.items():
        for s in sheets:
            if s.lower().replace(" ", "") in sheet_name.lower().replace(" ", ""):
                return cat
    return "Other"

for c in all_credits:
    c["category"] = find_category(c["sheet_name"])

total_credits = len(all_credits)
total_questions = sum(len(c["questions"]) for c in all_credits)
print(f"  {total_credits} credits, {total_questions} questions")

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 2: Parse DOCX guidance
# ═══════════════════════════════════════════════════════════════════════════════
print("Parsing DOCX guidelines...")
docx_doc = DocxDocument("Green Star Buildings v1.1_Submission Guidelines_RevA.docx")

_skip_h1 = {"Version control", "Table of contents", "Introduction",
            "Responsible", "Healthy", "Resilient", "Positive",
            "Places", "People", "Nature", "Leadership"}

docx_guidance = {}
_cc = _h2 = _h3 = _h4 = _h6 = None

for para in docx_doc.paragraphs:
    sn = para.style.name if para.style else ""
    txt = para.text.strip()
    if not txt:
        continue
    if "Heading 1" in sn:
        if txt in _skip_h1 or txt.startswith("Appendix"):
            _cc = None
            continue
        _cc = txt
        _h2 = _h3 = _h4 = _h6 = None
        docx_guidance[_cc] = {"outcome": "", "requirements": {}, "guidance": {}, "evidence": {}, "definitions": []}
    elif _cc and _cc in docx_guidance:
        g = docx_guidance[_cc]
        if "Heading 2" in sn:
            _h2 = txt; _h3 = _h4 = _h6 = None
        elif "Heading 3" in sn:
            _h3 = txt; _h4 = _h6 = None
            if _h2 == "Requirements" and _h3 not in g["requirements"]:
                g["requirements"][_h3] = {}
        elif "Heading 4" in sn:
            _h4 = txt; _h6 = None
            if _h2 == "Requirements" and _h3 and _h3 in g["requirements"]:
                g["requirements"][_h3][_h4] = ""
        elif any(f"Heading {n}" in sn for n in [5, 6, 7]):
            _h6 = txt
            if _h2 == "Guidance":
                g["guidance"][_h6] = ""
            elif _h2 == "Submission content":
                g["evidence"][_h6] = []
        else:
            if _h2 == "Outcome":
                g["outcome"] += txt + " "
            elif _h2 == "Requirements" and _h3 and _h4:
                if _h3 in g["requirements"] and _h4 in g["requirements"][_h3]:
                    g["requirements"][_h3][_h4] += txt + " "
            elif _h2 == "Guidance":
                if _h6 and _h6 in g["guidance"]:
                    g["guidance"][_h6] += txt + " "
                else:
                    g["guidance"]["_general"] = g["guidance"].get("_general", "") + txt + " "
            elif _h2 == "Submission content":
                if _h6 and _h6 in g["evidence"]:
                    g["evidence"][_h6].append(txt)
            elif _h2 == "Definitions":
                g["definitions"].append(txt)

print(f"  {len(docx_guidance)} credits from DOCX")


def _find_docx(sheet_name):
    sn = sheet_name.lower().replace(" ", "")
    for dn, data in docx_guidance.items():
        if sn in dn.lower().replace(" ", "") or dn.lower().replace(" ", "") in sn:
            return data
    return None


def _match_criteria(crit_name, req_dict, guide_dict, ev_dict):
    cn = crit_name.lower().replace(" ", "").replace("-", "").replace("–", "")
    if not cn or cn == "general":
        return None, None, None
    req_match = guide_match = ev_match = None
    for level, crits in req_dict.items():
        for cname, ctext in crits.items():
            if cn in cname.lower().replace(" ", "").replace("-", "") or \
               cname.lower().replace(" ", "").replace("-", "") in cn:
                req_match = (level, cname, ctext.strip())
                break
    for topic, gtext in guide_dict.items():
        if topic == "_general":
            continue
        tn = topic.lower().replace(" ", "").replace("-", "")
        if cn in tn or tn in cn:
            guide_match = gtext.strip()
            break
    for topic, items in ev_dict.items():
        tn = topic.lower().replace(" ", "").replace("-", "")
        if cn in tn or tn in cn:
            ev_match = items
            break
    return req_match, guide_match, ev_match


def get_guidance_text(sheet_name, crit_name, q_type, data_note):
    """Build a plain-text guidance string for a question."""
    g = _find_docx(sheet_name)
    parts = []
    if g:
        if g["outcome"]:
            parts.append(f"OUTCOME: {g['outcome'].strip()[:250]}")
        req_m, guide_m, ev_m = _match_criteria(crit_name or "", g["requirements"], g["guidance"], g["evidence"])
        if req_m:
            parts.append(f"REQUIREMENT ({req_m[0]} — {req_m[1]}): {req_m[2][:350]}")
        if guide_m:
            parts.append(f"WATCH OUT: {guide_m[:350]}")
        if ev_m:
            parts.append("EVIDENCE NEEDED: " + "; ".join(e[:100] for e in ev_m[:4]))
        if g["definitions"]:
            parts.append("DEFINITIONS: " + "; ".join(d[:100] for d in g["definitions"][:2]))
    if data_note:
        parts.append(f"NOTE: {data_note}")
    return "\n\n".join(parts) if parts else data_note or ""


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 3: Build conditional rules (same as website)
# ═══════════════════════════════════════════════════════════════════════════════
print("Building conditional rules...")

conditional_rules = {}  # (credit_idx, follower_ref) -> (credit_idx, gateway_ref, show_when)

for ci, c in enumerate(all_credits):
    sheet = c["sheet_name"]
    qs = c["questions"]
    refs = {q["ref"]: q for q in qs}
    ref_list = [q["ref"] for q in qs]
    ref_idx = {r: i for i, r in enumerate(ref_list)}
    rules = []

    def add_rules(gw, followers, val="Yes"):
        for f in followers:
            if gw in ref_idx and f in ref_idx:
                rules.append((f, gw, val))

    sn = sheet.lower().replace(" ", "")
    if "industrydevelopment" in sn:
        add_rules("ID.5", ["ID.6"], "Yes")
    elif "responsibleconstruction" in sn:
        add_rules("RC.1", ["RC.3"], "Yes")
        add_rules("RC.1", ["RC.2"], "No")
        add_rules("RC.4", ["RC.5"], "Yes")
    elif "verificationandhandover" in sn:
        add_rules("VH.5", ["VH.6"], "Yes")
        add_rules("VH.7", ["VH.8"], "Yes")
        add_rules("VH.26", ["VH.27"], "Yes")
    elif "responsibleresource" in sn:
        add_rules("RRM.2", ["RRM.3"], "Yes")
        add_rules("RRM.5", ["RRM.6"], "Yes")
        add_rules("RRM.7", ["RRM.8"], "Yes")
        add_rules("RRM.12", ["RRM.13"], "Yes")
    elif "responsibleprocurement" in sn:
        add_rules("RP.12", ["RP.13"], "Yes")
    elif "cleanair" in sn:
        add_rules("CA.9", ["CA.10"], "Yes")
    elif "lightquality" in sn:
        add_rules("LQ.7", ["LQ.8"], "Yes")
        add_rules("LQ.7", ["LQ.9", "LQ.10", "LQ.11"], "No")
    elif "exposuretotoxins" in sn:
        add_rules("ET.1", ["ET.2", "ET.3"], "Yes")
    elif "amenityandcomfort" in sn or "amenity" in sn:
        add_rules("AmC.3", ["AmC.4"], "Yes")
    elif "connectiontonature" in sn:
        add_rules("CN.4", ["CN.5", "CN.6"], "Yes")
    elif "climateresilience" in sn:
        add_rules("CR.1", ["CR.2", "CR.3", "CR.4"], "Yes")
    elif "operationsresilience" in sn:
        add_rules("OR.4", ["OR.5"], "Yes")
        add_rules("OR.6", ["OR.7"], "Yes")
    elif "communityresilience" in sn:
        add_rules("CoR.1", ["CoR.2", "CoR.3", "CoR.4"], "Yes")
    elif "gridresilience" in sn:
        add_rules("GR.1", ["GR.2", "GR.3"], "Yes")
        add_rules("GR.4", ["GR.5", "GR.6"], "Yes")
        add_rules("GR.7", ["GR.8"], "Yes")
    elif "energysource" in sn:
        add_rules("ES.5", ["ES.6"], "Yes")
    elif "upfrontcarbonreduction" in sn:
        add_rules("UCR.2", ["UCR.3", "UCR.4", "UCR.5"], "Yes")
    elif "wateruse" in sn:
        add_rules("WU.3", ["WU.4"], "Yes")
        add_rules("WU.5", ["WU.6"], "Yes")
    elif "contributiontoplace" in sn:
        add_rules("CP.1", ["CP.2"], "Yes")
    elif "cultureheritage" in sn:
        add_rules("CHI.1", ["CHI.2"], "Yes")
    elif "firstnations" in sn:
        add_rules("FNI.1", ["FNI.2", "FNI.3"], "Yes")
    elif "designforequity" in sn:
        add_rules("DE.4", ["DE.5"], "Yes")
    elif "impactstonature" in sn:
        add_rules("IN.1", ["IN.2", "IN.3"], "Yes")
        add_rules("IN.7", ["IN.8"], "Yes")
    elif "natureconnectivity" in sn:
        add_rules("NC.1", ["NC.2"], "Yes")
        add_rules("NC.5", ["NC.6"], "Yes")
    elif "naturestewardship" in sn:
        add_rules("NS.1", ["NS.2", "NS.3"], "Yes")
    elif "markettransformation" in sn:
        add_rules("MT.4", ["MT.5"], "Yes")
    elif "waterwayprotection" in sn:
        add_rules("WP.5", ["WP.6"], "Yes")
    elif "impactsdisclosure" in sn:
        add_rules("ID2.5", ["ID2.6"], "Yes")

    for follower_ref, gateway_ref, show_val in rules:
        conditional_rules[(ci, follower_ref)] = (ci, gateway_ref, show_val)

print(f"  {len(conditional_rules)} conditional rules")

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 4: Build VBA source code (exported as .bas for manual import)
# ═══════════════════════════════════════════════════════════════════════════════
print("Preparing VBA code...")


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 4 (continued): Build VBA source code
# ═══════════════════════════════════════════════════════════════════════════════

# Build a lookup table for conditional rules as VBA readable format
# Format: "SheetName|FollowerRef|GatewayRef|ShowWhen"
cond_rules_vba = []
for (ci, fref), (_, gref, sval) in conditional_rules.items():
    sname = all_credits[ci]["sheet_name"][:31]
    cond_rules_vba.append(f"{sname}|{fref}|{gref}|{sval}")

# Build the credit metadata for VBA
credit_meta_vba = []
for ci, c in enumerate(all_credits):
    cat = c["category"]
    color = CATEGORY_COLORS.get(cat, "333333")
    qcount = len(c["questions"])
    credit_meta_vba.append(f"{c['sheet_name'][:31]}|{cat}|{color}|{qcount}")

# Build question row map: SheetName -> list of (ref, row_number, q_type)
# We'll populate row numbers during workbook creation (Step 6)
# For now, store refs and types per credit
question_refs_per_credit = {}
for ci, c in enumerate(all_credits):
    sname = c["sheet_name"][:31]
    question_refs_per_credit[sname] = [(q["ref"], q["type"]) for q in c["questions"]]


# ── VBA ThisWorkbook module ──
THISWORKBOOK_CODE = '''Attribute VB_Name = "ThisWorkbook"
Attribute VB_Base = "0{00020819-0000-0000-C000-000000000046}"
Attribute VB_GlobalNameSpace = False
Attribute VB_Creatable = False
Attribute VB_PredeclaredId = True
Attribute VB_Exposed = True
Attribute VB_TemplateDerived = False
Attribute VB_Customizable = True

Private Sub Workbook_SheetChange(ByVal Sh As Object, ByVal Target As Range)
    On Error Resume Next
    If Target.Column = 7 And Target.Count = 1 Then
        Application.EnableEvents = False
        GreenStarMacros.HandleChange Sh, Target
        Application.EnableEvents = True
    End If
    On Error GoTo 0
End Sub

Private Sub Workbook_Open()
    On Error Resume Next
    GreenStarMacros.InitWorkbook
    On Error GoTo 0
End Sub
'''

# Build VBA rules data as a const string (newline-separated entries)
rules_str = chr(10).join(cond_rules_vba)
meta_str = chr(10).join(credit_meta_vba)

# ── VBA GreenStarMacros module ──
MACROS_CODE = '''Attribute VB_Name = "GreenStarMacros"
Option Explicit

' ============================================================
' GREEN STAR BUILDINGS v1.1 - INTERACTIVE WORKBOOK MACROS
' ============================================================

' ── Data ──
Private Const RULES_DATA As String = "''' + rules_str.replace('"', '""') + '''"
Private Const META_DATA As String = "''' + meta_str.replace('"', '""') + '''"

' ── Colour Scheme ──
Private Const CLR_DARK_BG As Long = &H21201A
Private Const CLR_DARK_CARD As Long = &H342C28
Private Const CLR_DARK_TEXT As Long = &HE0E0E0
Private Const CLR_DARK_INPUT As Long = &H3A3130

Private gDarkMode As Boolean
Private gSearchSheet As String
Private gSearchRow As Long

' ============================================================
' INITIALISATION
' ============================================================
Public Sub InitWorkbook()
    SetupDashboard
End Sub

' ============================================================
' HANDLE CELL CHANGES (Response column = G)
' ============================================================
Public Sub HandleChange(Sh As Object, Target As Range)
    ' Apply conditional visibility rules
    ApplyConditionalRules Sh
    ' Update progress on Dashboard
    UpdateDashboardProgress
    ' Log to history
    LogChange Sh.Name, Target.Row, Target.Value
End Sub

' ============================================================
' CONDITIONAL VISIBILITY
' ============================================================
Public Sub ApplyConditionalRules(Sh As Object)
    Dim rules() As String
    Dim parts() As String
    Dim i As Long
    Dim sName As String

    If Len(RULES_DATA) = 0 Then Exit Sub
    rules = Split(RULES_DATA, vbLf)
    sName = Sh.Name

    For i = LBound(rules) To UBound(rules)
        If Len(rules(i)) = 0 Then GoTo NextRule
        parts = Split(rules(i), "|")
        If UBound(parts) < 3 Then GoTo NextRule
        If parts(0) <> sName Then GoTo NextRule

        Dim followerRef As String, gatewayRef As String, showWhen As String
        followerRef = parts(1)
        gatewayRef = parts(2)
        showWhen = parts(3)

        ' Find gateway row
        Dim gwRow As Long, fRow As Long
        gwRow = FindRefRow(Sh, gatewayRef)
        fRow = FindRefRow(Sh, followerRef)
        If gwRow = 0 Or fRow = 0 Then GoTo NextRule

        Dim gwVal As String
        gwVal = CStr(Sh.Cells(gwRow, 7).Value)

        If gwVal = showWhen Then
            If Sh.Rows(fRow).Hidden Then
                Sh.Rows(fRow).Hidden = False
            End If
        Else
            If Not Sh.Rows(fRow).Hidden Then
                Sh.Rows(fRow).Hidden = True
            End If
        End If
NextRule:
    Next i
End Sub

Private Function FindRefRow(Sh As Object, ref As String) As Long
    Dim lastRow As Long, r As Long
    lastRow = Sh.Cells(Sh.Rows.Count, 1).End(xlUp).Row
    For r = 2 To lastRow
        If CStr(Sh.Cells(r, 1).Value) = ref Then
            FindRefRow = r
            Exit Function
        End If
    Next r
    FindRefRow = 0
End Function

Public Sub ApplyAllConditionalRules()
    Dim ws As Worksheet
    For Each ws In ThisWorkbook.Worksheets
        If ws.Name <> "Dashboard" And ws.Name <> "History" And ws.Name <> "SearchResults" Then
            ApplyConditionalRules ws
        End If
    Next ws
End Sub

' ============================================================
' DASHBOARD PROGRESS
' ============================================================
Public Sub SetupDashboard()
    On Error Resume Next
    Dim dsh As Worksheet
    Set dsh = ThisWorkbook.Worksheets("Dashboard")
    If dsh Is Nothing Then Exit Sub

    UpdateDashboardProgress
    ApplyAllConditionalRules
    On Error GoTo 0
End Sub

Public Sub UpdateDashboardProgress()
    Dim dsh As Worksheet
    Set dsh = ThisWorkbook.Worksheets("Dashboard")
    If dsh Is Nothing Then Exit Sub

    Dim meta() As String
    Dim parts() As String
    Dim totalQ As Long, totalA As Long
    totalQ = 0: totalA = 0

    If Len(META_DATA) = 0 Then Exit Sub
    meta = Split(META_DATA, vbLf)

    Dim dashRow As Long
    dashRow = 5  ' First credit row on dashboard

    Dim i As Long
    For i = LBound(meta) To UBound(meta)
        If Len(meta(i)) = 0 Then GoTo NextMeta
        parts = Split(meta(i), "|")
        If UBound(parts) < 3 Then GoTo NextMeta

        Dim sName As String, qCount As Long
        sName = parts(0)
        qCount = CLng(parts(3))

        ' Count answered in this credit sheet
        Dim ws As Worksheet, answered As Long, visible As Long
        Set ws = Nothing
        On Error Resume Next
        Set ws = ThisWorkbook.Worksheets(sName)
        On Error GoTo 0
        If ws Is Nothing Then GoTo NextMeta

        answered = 0: visible = 0
        Dim r As Long
        For r = 2 To ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
            If Len(CStr(ws.Cells(r, 5).Value)) > 0 Then  ' Has question type = is a question row
                If Not ws.Rows(r).Hidden Then
                    visible = visible + 1
                    If Len(CStr(ws.Cells(r, 7).Value)) > 0 Then
                        answered = answered + 1
                    End If
                End If
            End If
        Next r

        totalQ = totalQ + visible
        totalA = totalA + answered

        ' Update dashboard row
        If dashRow <= dsh.Cells(dsh.Rows.Count, 1).End(xlUp).Row + 5 Then
            dsh.Cells(dashRow, 4).Value = answered
            dsh.Cells(dashRow, 5).Value = visible
            If visible > 0 Then
                dsh.Cells(dashRow, 6).Value = answered / visible
            Else
                dsh.Cells(dashRow, 6).Value = 0
            End If
            dashRow = dashRow + 1
        End If
NextMeta:
    Next i

    ' Update totals
    dsh.Cells(2, 4).Value = totalA
    dsh.Cells(2, 5).Value = totalQ
    If totalQ > 0 Then
        dsh.Cells(2, 6).Value = totalA / totalQ
    Else
        dsh.Cells(2, 6).Value = 0
    End If
End Sub

' ============================================================
' N/A TOGGLE
' ============================================================
Public Sub ToggleNA()
    Dim sName As String
    sName = ActiveSheet.Name
    If sName = "Dashboard" Or sName = "History" Or sName = "SearchResults" Then Exit Sub

    Dim ws As Worksheet
    Set ws = ActiveSheet

    ' Check current state - look at row 2 font color
    If ws.Cells(2, 1).Font.Color = RGB(180, 180, 180) Then
        ' Currently N/A - re-enable
        Dim r2 As Long
        For r2 = 1 To ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
            ws.Rows(r2).Font.Color = RGB(0, 0, 0)
        Next r2
        ws.Cells(1, 8).Value = ""
    Else
        ' Mark as N/A
        Dim r3 As Long
        For r3 = 2 To ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
            ws.Rows(r3).Font.Color = RGB(180, 180, 180)
        Next r3
        ws.Cells(1, 8).Value = "N/A"
    End If
    UpdateDashboardProgress
End Sub

' ============================================================
' REVIEW MODE - Highlight unanswered
' ============================================================
Public Sub ReviewMode()
    Dim ws As Worksheet
    Set ws = ActiveSheet
    If ws.Name = "Dashboard" Or ws.Name = "History" Or ws.Name = "SearchResults" Then Exit Sub

    Dim r As Long, unanswered As Long
    unanswered = 0

    For r = 2 To ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
        If Len(CStr(ws.Cells(r, 5).Value)) > 0 Then  ' Question row
            If Not ws.Rows(r).Hidden Then
                If Len(CStr(ws.Cells(r, 7).Value)) = 0 Then
                    ' Highlight unanswered
                    ws.Cells(r, 7).Interior.Color = RGB(255, 243, 224)
                    ws.Cells(r, 7).Borders.Color = RGB(255, 152, 0)
                    unanswered = unanswered + 1
                Else
                    ' Clear highlight
                    ws.Cells(r, 7).Interior.Color = RGB(255, 255, 255)
                    ws.Cells(r, 7).Borders.Color = RGB(200, 200, 200)
                End If
            End If
        End If
    Next r

    MsgBox unanswered & " unanswered question(s) highlighted in orange on " & ws.Name, vbInformation, "Review Mode"
End Sub

' ============================================================
' SEARCH
' ============================================================
Public Sub SearchQuestions()
    Dim query As String
    query = InputBox("Search across all questions:" & vbCrLf & vbCrLf & "Enter search term(s):", "Search Green Star Questions")
    If Len(query) = 0 Then Exit Sub

    Dim searchTerm As String
    searchTerm = LCase(Trim(query))

    ' Create or clear SearchResults sheet
    Dim sr As Worksheet
    On Error Resume Next
    Set sr = ThisWorkbook.Worksheets("SearchResults")
    On Error GoTo 0
    If sr Is Nothing Then
        Set sr = ThisWorkbook.Worksheets.Add(After:=ThisWorkbook.Worksheets(ThisWorkbook.Worksheets.Count))
        sr.Name = "SearchResults"
    End If
    sr.Cells.Clear

    ' Header
    sr.Cells(1, 1).Value = "Search Results for: """ & query & """"
    sr.Cells(1, 1).Font.Bold = True
    sr.Cells(1, 1).Font.Size = 14

    sr.Cells(2, 1).Value = "Credit"
    sr.Cells(2, 2).Value = "Ref"
    sr.Cells(2, 3).Value = "Question"
    sr.Cells(2, 4).Value = "Type"
    sr.Cells(2, 5).Value = "Current Response"
    Dim c As Long
    For c = 1 To 5
        sr.Cells(2, c).Font.Bold = True
        sr.Cells(2, c).Interior.Color = RGB(31, 78, 40)
        sr.Cells(2, c).Font.Color = RGB(255, 255, 255)
    Next c

    sr.Columns(1).ColumnWidth = 25
    sr.Columns(2).ColumnWidth = 8
    sr.Columns(3).ColumnWidth = 60
    sr.Columns(4).ColumnWidth = 16
    sr.Columns(5).ColumnWidth = 40

    Dim resultRow As Long
    resultRow = 3

    Dim ws As Worksheet
    For Each ws In ThisWorkbook.Worksheets
        If ws.Name = "Dashboard" Or ws.Name = "History" Or ws.Name = "SearchResults" Then GoTo NextSheet
        Dim r As Long
        For r = 2 To ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
            If Len(CStr(ws.Cells(r, 5).Value)) > 0 Then
                Dim haystack As String
                haystack = LCase(CStr(ws.Cells(r, 1).Value) & " " & CStr(ws.Cells(r, 6).Value) & " " & CStr(ws.Cells(r, 8).Value))
                If InStr(haystack, searchTerm) > 0 Then
                    sr.Cells(resultRow, 1).Value = ws.Name
                    sr.Cells(resultRow, 2).Value = ws.Cells(r, 1).Value
                    sr.Cells(resultRow, 3).Value = ws.Cells(r, 6).Value
                    sr.Cells(resultRow, 4).Value = ws.Cells(r, 5).Value
                    sr.Cells(resultRow, 5).Value = ws.Cells(r, 7).Value
                    ' Add hyperlink to jump to the question
                    sr.Hyperlinks.Add sr.Cells(resultRow, 2), "", "'" & ws.Name & "'!A" & r, "Go to question"
                    resultRow = resultRow + 1
                End If
            End If
        Next r
NextSheet:
    Next ws

    sr.Cells(1, 3).Value = (resultRow - 3) & " result(s) found"

    sr.Activate
End Sub

' ============================================================
' VERSION HISTORY
' ============================================================
Public Sub LogChange(sheetName As String, row As Long, newValue As Variant)
    On Error Resume Next
    Dim hsh As Worksheet
    Set hsh = ThisWorkbook.Worksheets("History")
    If hsh Is Nothing Then Exit Sub

    Dim nextRow As Long
    nextRow = hsh.Cells(hsh.Rows.Count, 1).End(xlUp).Row + 1
    If nextRow < 3 Then nextRow = 3

    ' Keep max 500 entries
    If nextRow > 502 Then
        hsh.Rows("3:103").Delete
        nextRow = hsh.Cells(hsh.Rows.Count, 1).End(xlUp).Row + 1
    End If

    hsh.Cells(nextRow, 1).Value = Now
    hsh.Cells(nextRow, 1).NumberFormat = "yyyy-mm-dd hh:mm:ss"
    hsh.Cells(nextRow, 2).Value = sheetName
    hsh.Cells(nextRow, 3).Value = "Row " & row
    hsh.Cells(nextRow, 4).Value = CStr(newValue)
    On Error GoTo 0
End Sub

Public Sub ShowHistory()
    On Error Resume Next
    ThisWorkbook.Worksheets("History").Activate
    On Error GoTo 0
End Sub

' ============================================================
' DARK MODE
' ============================================================
Public Sub ToggleDarkMode()
    gDarkMode = Not gDarkMode
    Dim ws As Worksheet

    If gDarkMode Then
        For Each ws In ThisWorkbook.Worksheets
            ws.Tab.Color = RGB(30, 33, 39)
            Dim lastR As Long, lastC As Long
            lastR = ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
            lastC = 8
            Dim r As Long, cl As Long
            For r = 1 To lastR
                For cl = 1 To lastC
                    If ws.Cells(r, cl).Interior.Color = RGB(255, 255, 255) Or _
                       ws.Cells(r, cl).Interior.ColorIndex = xlNone Then
                        ws.Cells(r, cl).Interior.Color = CLR_DARK_CARD
                        ws.Cells(r, cl).Font.Color = CLR_DARK_TEXT
                    End If
                Next cl
            Next r
        Next ws
        Application.StatusBar = "Dark mode ON"
    Else
        For Each ws In ThisWorkbook.Worksheets
            ws.Tab.Color = xlNone
            lastR = ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
            lastC = 8
            For r = 1 To lastR
                For cl = 1 To lastC
                    If ws.Cells(r, cl).Interior.Color = CLR_DARK_CARD Then
                        ws.Cells(r, cl).Interior.Color = RGB(255, 255, 255)
                        ws.Cells(r, cl).Font.Color = RGB(0, 0, 0)
                    End If
                Next cl
            Next r
        Next ws
        Application.StatusBar = "Dark mode OFF"
    End If
End Sub

' ============================================================
' NAVIGATION HELPERS
' ============================================================
Public Sub GoToDashboard()
    ThisWorkbook.Worksheets("Dashboard").Activate
End Sub

Public Sub RefreshAll()
    ApplyAllConditionalRules
    UpdateDashboardProgress
    MsgBox "All conditional rules applied and dashboard updated.", vbInformation, "Refresh Complete"
End Sub
'''

print("  VBA modules prepared")

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 6: Build the workbook with openpyxl
# ═══════════════════════════════════════════════════════════════════════════════
print("Building workbook...")

wb = Workbook()

# ── Styles ──
header_font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
credit_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
level_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
criteria_font = Font(name="Calibri", bold=True, size=11, color="1F4E28")
question_font = Font(name="Calibri", size=10)
data_flag_font = Font(name="Calibri", size=10, italic=True, color="2E75B6")
condition_font = Font(name="Calibri", bold=True, size=10, color="7030A0")
guidance_font = Font(name="Calibri", size=9, italic=True, color="666666")
wrap = Alignment(wrap_text=True, vertical="top")
center_wrap = Alignment(wrap_text=True, vertical="center", horizontal="center")

green_fill = PatternFill(start_color="1F4E28", end_color="1F4E28", fill_type="solid")
dark_green_fill = PatternFill(start_color="0D3318", end_color="0D3318", fill_type="solid")
level_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
criteria_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
question_fill = PatternFill(start_color="F1F8E9", end_color="F1F8E9", fill_type="solid")
condition_fill = PatternFill(start_color="EDE7F6", end_color="EDE7F6", fill_type="solid")
data_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
guidance_fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

COL_WIDTHS = {"A": 8, "B": 20, "C": 22, "D": 28, "E": 16, "F": 55, "G": 50, "H": 45}

yn_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
yn_dv.error = "Please select Yes or No"
yn_dv.errorTitle = "Invalid entry"
yn_dv.prompt = "Select Yes or No"
yn_dv.promptTitle = "Condition"

# ── Dashboard sheet ──
dsh = wb.active
dsh.title = "Dashboard"
dsh.sheet_properties.tabColor = "1F4E28"

# Dashboard header
dsh.merge_cells("A1:F1")
dsh.cell(row=1, column=1, value="Green Star Buildings v1.1 — Submission Dashboard")
dsh.cell(row=1, column=1).font = Font(name="Calibri", bold=True, size=18, color="FFFFFF")
dsh.cell(row=1, column=1).fill = green_fill
dsh.cell(row=1, column=1).alignment = Alignment(vertical="center")
dsh.row_dimensions[1].height = 45

# Summary row
labels = ["", "", "", "Answered", "Total", "Progress"]
for i, label in enumerate(labels):
    cell = dsh.cell(row=2, column=i + 1, value=label)
    cell.font = Font(name="Calibri", bold=True, size=11)
    cell.alignment = center_wrap

dsh.cell(row=2, column=1, value="TOTAL")
dsh.cell(row=2, column=1).font = Font(name="Calibri", bold=True, size=12, color="1F4E28")
dsh.cell(row=2, column=4, value=0)
dsh.cell(row=2, column=5, value=total_questions)
dsh.cell(row=2, column=6, value=0)
dsh.cell(row=2, column=6).number_format = '0%'
dsh.row_dimensions[2].height = 30

# Blank row
dsh.row_dimensions[3].height = 10

# Column headers for credit list
credit_headers = ["Credit", "Category", "Colour", "Answered", "Total Visible", "Progress"]
for i, h in enumerate(credit_headers):
    cell = dsh.cell(row=4, column=i + 1, value=h)
    cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    cell.fill = dark_green_fill
    cell.alignment = center_wrap
    cell.border = thin_border
dsh.row_dimensions[4].height = 25

# Credit rows
for ci, c in enumerate(all_credits):
    row = ci + 5
    sname = c["sheet_name"][:31]
    cat = c["category"]
    color = CATEGORY_COLORS.get(cat, "333333")
    qcount = len(c["questions"])

    # Credit name with hyperlink
    cell = dsh.cell(row=row, column=1, value=sname)
    cell.hyperlink = f"#{sname}!A1"
    cell.font = Font(name="Calibri", size=10, color="1F4E28", underline="single")
    cell.border = thin_border

    dsh.cell(row=row, column=2, value=cat).border = thin_border
    dsh.cell(row=row, column=2).font = Font(name="Calibri", size=10)

    # Category colour indicator
    color_cell = dsh.cell(row=row, column=3)
    color_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    color_cell.border = thin_border

    # Answered (VBA will update)
    dsh.cell(row=row, column=4, value=0).border = thin_border
    dsh.cell(row=row, column=4).alignment = Alignment(horizontal="center")

    # Total
    dsh.cell(row=row, column=5, value=qcount).border = thin_border
    dsh.cell(row=row, column=5).alignment = Alignment(horizontal="center")

    # Progress
    dsh.cell(row=row, column=6, value=0).border = thin_border
    dsh.cell(row=row, column=6).number_format = '0%'
    dsh.cell(row=row, column=6).alignment = Alignment(horizontal="center")

    dsh.row_dimensions[row].height = 22

# Dashboard column widths
dsh.column_dimensions["A"].width = 30
dsh.column_dimensions["B"].width = 15
dsh.column_dimensions["C"].width = 8
dsh.column_dimensions["D"].width = 12
dsh.column_dimensions["E"].width = 14
dsh.column_dimensions["F"].width = 12

# Toolbar buttons row
btn_row = total_credits + 6
dsh.cell(row=btn_row, column=1, value="Actions:").font = Font(bold=True, size=11)
buttons = [
    ("Review Mode", "ReviewMode"),
    ("Search", "SearchQuestions"),
    ("Dark Mode", "ToggleDarkMode"),
    ("History", "ShowHistory"),
    ("Refresh All", "RefreshAll"),
]
for i, (label, _) in enumerate(buttons):
    cell = dsh.cell(row=btn_row, column=i + 2, value=f"[ {label} ]")
    cell.font = Font(name="Calibri", size=10, color="1F4E28", bold=True)
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

# Add instructions
inst_row = btn_row + 2
instructions = [
    "HOW TO USE:",
    "• Click any credit name above to navigate to its questions",
    "• Fill in the Response column (G) for each question",
    "• Y/N questions use dropdown validation",
    "• Progress updates automatically via formulas (Answered / Total / %)",
    "• The Guidance column (H) shows submission guidelines, tips, and evidence requirements",
    "• Sheets are protected — only the Response column is editable (password: greenstar)",
    "",
    "OPTIONAL VBA MACROS (for advanced features):",
    "• Save as .xlsm, then import GreenStarMacros.bas via Developer > Visual Basic > File > Import",
    "• Adds: conditional row visibility, search, review mode, dark mode, version history",
]
for i, line in enumerate(instructions):
    cell = dsh.cell(row=inst_row + i, column=1, value=line)
    cell.font = Font(name="Calibri", size=10, color="555555", italic=(i > 0))
    if i == 0:
        cell.font = Font(name="Calibri", size=11, color="1F4E28", bold=True)

dsh.freeze_panes = "A5"

# ── Credit sheets ──
question_row_map = {}  # (sheet_name, ref) -> row number (for conditional rules)

for ci, credit in enumerate(all_credits):
    sname = credit["sheet_name"][:31]
    ws = wb.create_sheet(title=sname)
    cat = credit["category"]
    cat_color = CATEGORY_COLORS.get(cat, "333333")
    ws.sheet_properties.tabColor = cat_color

    # Column widths
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # Header row
    headers = ["Ref", "Credit", "Performance Level", "Criteria",
               "Question Type", "Question", "Response", "Guidance"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = dark_green_fill
        cell.alignment = center_wrap
        cell.border = thin_border
    ws.row_dimensions[1].height = 35
    ws.freeze_panes = "A2"

    # Add data validation for Y/N
    ws.add_data_validation(yn_dv)

    row = 2

    # Credit title row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    title = credit.get("title", sname)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = credit_font
    cell.fill = green_fill
    cell.alignment = wrap
    cell.border = thin_border
    ws.row_dimensions[row].height = 30
    row += 1

    # "Back to Dashboard" link row
    ws.cell(row=row, column=1, value="<< Dashboard")
    ws.cell(row=row, column=1).hyperlink = "#Dashboard!A1"
    ws.cell(row=row, column=1).font = Font(name="Calibri", size=9, color="1F4E28", underline="single")
    ws.row_dimensions[row].height = 18
    row += 1

    for section in credit["sections"]:
        # Level header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        cell = ws.cell(row=row, column=1, value=section["title"])
        cell.font = level_font
        cell.fill = level_fill
        cell.alignment = wrap
        cell.border = thin_border
        ws.row_dimensions[row].height = 22
        row += 1

        for crit in section["criteria"]:
            # Criteria header
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            cell = ws.cell(row=row, column=1, value=crit["name"])
            cell.font = criteria_font
            cell.fill = criteria_fill
            cell.alignment = wrap
            cell.border = thin_border
            ws.row_dimensions[row].height = 20
            row += 1

            for q in crit["questions"]:
                is_yn = q["type"] == "Condition (Y/N)"
                is_data = q["type"] == "Data"

                # Build guidance text
                guidance_text = get_guidance_text(
                    credit["sheet_name"], crit["name"], q["type"], q["data_note"]
                )

                values = [q["ref"], q["credit"], q["level"], q["criteria"],
                          q["type"], q["question"], "", guidance_text]

                for col, val in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.alignment = wrap
                    cell.border = thin_border

                    if col == 7:
                        # Response column - white, unlocked
                        cell.fill = white_fill
                        cell.font = question_font
                    elif col == 8:
                        # Guidance column
                        cell.fill = guidance_fill
                        cell.font = guidance_font
                    elif is_yn:
                        cell.fill = condition_fill
                        cell.font = condition_font if col == 5 else question_font
                    elif is_data and col == 5:
                        cell.fill = data_fill
                        cell.font = data_flag_font
                    else:
                        cell.fill = question_fill
                        cell.font = question_font

                # Y/N dropdown
                if is_yn:
                    yn_dv.add(ws.cell(row=row, column=7))

                ws.row_dimensions[row].height = 60

                # Track row for conditional rules
                question_row_map[(sname, q["ref"])] = row

                row += 1

    # Protect sheet - only column G (Response) is editable
    for r in range(1, row):
        for c in range(1, 9):
            ws.cell(row=r, column=c).protection = Protection(locked=True)
        # Unlock response column
        if ws.cell(row=r, column=5).value in ["Descriptive", "Data", "Condition (Y/N)"]:
            ws.cell(row=r, column=7).protection = Protection(locked=False)

    ws.protection.sheet = True
    ws.protection.password = "greenstar"
    ws.protection.enable()

# ── History sheet (hidden) ──
hsh = wb.create_sheet(title="History")
hsh.cell(row=1, column=1, value="Version History").font = Font(bold=True, size=14, color="1F4E28")
hsh.cell(row=2, column=1, value="Timestamp").font = Font(bold=True)
hsh.cell(row=2, column=2, value="Sheet").font = Font(bold=True)
hsh.cell(row=2, column=3, value="Location").font = Font(bold=True)
hsh.cell(row=2, column=4, value="New Value").font = Font(bold=True)
hsh.column_dimensions["A"].width = 20
hsh.column_dimensions["B"].width = 25
hsh.column_dimensions["C"].width = 12
hsh.column_dimensions["D"].width = 50
for c in range(1, 5):
    hsh.cell(row=2, column=c).fill = dark_green_fill
    hsh.cell(row=2, column=c).font = Font(bold=True, color="FFFFFF")
    hsh.cell(row=2, column=c).border = thin_border
hsh.freeze_panes = "A3"
hsh.sheet_state = "hidden"

# ── Dashboard protection ──
dsh.protection.sheet = True
dsh.protection.password = "greenstar"
dsh.protection.enable()

# ── Dashboard formulas: wire up progress via COUNTA ──
# Now that all credit sheets exist, add formulas referencing response columns
for ci, c in enumerate(all_credits):
    row = ci + 5
    sname = c["sheet_name"][:31]
    safe_name = "'" + sname.replace("'", "''") + "'"

    # Find the question rows for this credit (rows where col E has a value)
    ws = wb[sname]
    q_rows = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=5).value in ["Descriptive", "Data", "Condition (Y/N)"]:
            q_rows.append(r)

    if q_rows:
        # Build a COUNTA formula that counts non-blank cells in column G for question rows
        # Use a range approach: count non-blank in G column for question rows
        first_q = q_rows[0]
        last_q = q_rows[-1]
        # Answered = count non-blank responses in the question row range of column G
        # We use SUMPRODUCT to only count rows that are question rows (have col E non-blank)
        answered_formula = f'=SUMPRODUCT(({safe_name}!E{first_q}:E{last_q}<>"")*({safe_name}!G{first_q}:G{last_q}<>""))'
        total_formula = f'=COUNTA({safe_name}!E{first_q}:E{last_q})'

        dsh.cell(row=row, column=4, value=answered_formula)
        dsh.cell(row=row, column=5, value=total_formula)
        dsh.cell(row=row, column=6, value=f'=IF(E{row}>0,D{row}/E{row},0)')
        dsh.cell(row=row, column=6).number_format = '0%'

# Totals row formulas
dsh.cell(row=2, column=4, value=f'=SUM(D5:D{total_credits + 4})')
dsh.cell(row=2, column=5, value=f'=SUM(E5:E{total_credits + 4})')
dsh.cell(row=2, column=6, value='=IF(E2>0,D2/E2,0)')
dsh.cell(row=2, column=6).number_format = '0%'

print(f"  Dashboard formulas added")
print(f"  {total_credits} credit sheets created")
print(f"  {total_questions} questions with guidance")

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 7: Save workbook and export VBA as importable .bas file
# ═══════════════════════════════════════════════════════════════════════════════

print("Saving workbook...")
xlsx_path = "Green_Star_Buildings_v1.1_Interactive.xlsx"
wb.save(xlsx_path)
print(f"  Saved {xlsx_path} ({os.path.getsize(xlsx_path):,} bytes)")

# Export VBA code as a .bas file that can be imported into Excel
# (Developer tab > Visual Basic > File > Import File)
print("Exporting VBA module...")
bas_path = "GreenStarMacros.bas"
with open(bas_path, "w") as f:
    # Strip the Attribute lines from ThisWorkbook (not needed for .bas import)
    f.write(MACROS_CODE)

print(f"  Saved {bas_path}")

# Also write the ThisWorkbook code separately
twb_path = "ThisWorkbook.cls"
with open(twb_path, "w") as f:
    f.write(THISWORKBOOK_CODE)

print(f"  Saved {twb_path}")

# ═══════════════════════════════════════════════════════════════════════════════
# DONE
# ═══════════════════════════════════════════════════════════════════════════════
print(f"\nDone!")
print(f"  Credits: {total_credits}")
print(f"  Questions: {total_questions}")
print(f"  Conditional rules: {len(conditional_rules)}")
print(f"  Output: {xlsx_path} ({os.path.getsize(xlsx_path):,} bytes)")
print(f"  VBA module: {bas_path}")
print(f"  Protection password: greenstar")
print()
print("The .xlsx works immediately — Dashboard progress updates via formulas.")
print()
print("For advanced features (conditional visibility, search, review mode,")
print("dark mode, version history), import the VBA macros:")
print("  1. Open the .xlsx, save as .xlsm (macro-enabled)")
print("  2. Developer tab > Visual Basic > File > Import File")
print("  3. Import GreenStarMacros.bas")
print("  4. Double-click ThisWorkbook in the VBA editor, paste ThisWorkbook.cls")
print("  5. Save and enable macros")
