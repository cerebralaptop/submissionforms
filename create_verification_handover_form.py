#!/usr/bin/env python3
"""
Script to create Verification and Handover credit submission form for Green Star Buildings v1.1
With compliance tracking and tallying system for assessors.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule, CellIsRule

def create_submission_form():
    wb = Workbook()

    # Create Dashboard sheet first
    ws_dashboard = wb.active
    ws_dashboard.title = "Dashboard"

    # Create main assessment sheet
    ws = wb.create_sheet("Assessment")

    # Define styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    section_font = Font(bold=True, size=11)
    section_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    subsection_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    wrap_alignment = Alignment(wrap_text=True, vertical='top')

    # Conditional formatting colors
    complies_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")  # Green
    minor_nc_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")  # Yellow/Amber
    major_nc_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")  # Red
    na_fill = PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid")  # Grey

    # Set column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 45

    # Title
    ws.merge_cells('A1:G1')
    ws['A1'] = "GREEN STAR BUILDINGS v1.1 - VERIFICATION AND HANDOVER CREDIT SUBMISSION FORM"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Credit information
    ws.merge_cells('A2:G2')
    ws['A2'] = "Outcome: The building has been optimised and handed over to deliver a higher level of performance in operation."
    ws['A2'].font = Font(italic=True, size=10)
    ws['A2'].alignment = wrap_alignment

    # Headers
    row = 4
    headers = ['Criteria', 'Requirement', 'Submission Question', 'Compliance Status', 'Score', 'Evidence Provided', 'Documentation Required']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Data validation for compliance status dropdown
    dv = DataValidation(
        type="list",
        formula1='"Complies,Minor NC,Major NC,N/A"',
        allow_blank=True
    )
    dv.error = 'Please select from the dropdown'
    dv.errorTitle = 'Invalid Entry'
    dv.prompt = 'Select compliance status'
    dv.promptTitle = 'Compliance Status'
    ws.add_data_validation(dv)

    # Data structure for the form - now with section markers for tallying
    form_data = [
        # MINIMUM EXPECTATION SECTION
        {
            'criteria': 'MINIMUM EXPECTATION',
            'requirement': '',
            'question': '',
            'documentation': '',
            'is_section': True,
            'section_id': 'MIN_EXP'
        },
        # Metering and Monitoring
        {
            'criteria': 'Metering and Monitoring',
            'requirement': 'Metering Distribution',
            'question': 'Does the building have accessible energy and water metering for all distinct uses and major uses?',
            'documentation': 'As-built drawings showing location of all energy and water meters',
            'is_subsection': True,
            'section_id': 'M&M'
        },
        {
            'criteria': '',
            'requirement': '',
            'question': 'Is metering provided for each separate tenancy or unit?',
            'documentation': '',
            'section_id': 'M&M'
        },
        {
            'criteria': '',
            'requirement': '',
            'question': 'Is floor by floor metering provided where entire floors have a single use?',
            'documentation': '',
            'section_id': 'M&M'
        },
        {
            'criteria': '',
            'requirement': 'Metering Schedule',
            'question': 'Has a metering schedule been provided in accordance with CIBSE TM39?',
            'documentation': 'Metering schedule',
            'section_id': 'M&M'
        },
        {
            'criteria': '',
            'requirement': 'Metering Attributes',
            'question': 'Do all meters provide continual information (up to 1-hour interval readings)?',
            'documentation': 'Product data sheets demonstrating pattern approval by NMI or recognised standard',
            'section_id': 'M&M'
        },
        {
            'criteria': '',
            'requirement': '',
            'question': 'Are meters commissioned and validated per NABERS Metering and Consumption Rules?',
            'documentation': 'Completed metering validation documents',
            'section_id': 'M&M'
        },
        {
            'criteria': '',
            'requirement': '',
            'question': 'Are meters pattern approved by NMI or recognised standard?',
            'documentation': 'Product data sheets or certificates',
            'section_id': 'M&M'
        },
        {
            'criteria': '',
            'requirement': 'Automatic Monitoring',
            'question': 'Is an automatic monitoring system implemented that reports consumption trends?',
            'documentation': 'Letter of confirmation from contractor/metering provider',
            'section_id': 'M&M'
        },
        {
            'criteria': '',
            'requirement': '',
            'question': 'Does the system raise alarms when energy/water use exceeds parameters?',
            'documentation': 'Extracts from commissioning reports',
            'section_id': 'M&M',
            'is_section_end': True
        },
        # Commissioning and Tuning
        {
            'criteria': 'Commissioning and Tuning',
            'requirement': 'Environmental Performance Targets',
            'question': 'Were environmental performance targets documented prior to schematic design?',
            'documentation': 'Extracts from design intent report or OPR',
            'is_subsection': True,
            'section_id': 'C&T'
        },
        {
            'criteria': '',
            'requirement': '',
            'question': 'Does the report include targets for energy use, water consumption, IEQ and airtightness?',
            'documentation': '',
            'section_id': 'C&T'
        },
        {
            'criteria': '',
            'requirement': '',
            'question': 'Has the building owner signed off on the report/document?',
            'documentation': 'Signed design intent report or OPR',
            'section_id': 'C&T'
        },
        {
            'criteria': '',
            'requirement': 'Services and Maintainability Review',
            'question': 'Was a services and maintainability review conducted prior to construction?',
            'documentation': 'Evidence of review being conducted',
            'section_id': 'C&T'
        },
        {
            'criteria': '',
            'requirement': '',
            'question': 'Did the review involve building owner, design consultants, architect, FM, head contractor and ICA?',
            'documentation': '',
            'section_id': 'C&T'
        },
        {
            'criteria': '',
            'requirement': '',
            'question': 'Were all items addressed, closed-out and documented in Services and Maintainability Report?',
            'documentation': 'Services and Maintainability Report signed by all parties',
            'section_id': 'C&T'
        },
        {
            'criteria': '',
            'requirement': 'Building Commissioning',
            'question': 'Were commissioning requirements included in construction documentation prior to construction?',
            'documentation': 'Extracts from construction documentation',
            'section_id': 'C&T'
        },
        {
            'criteria': '',
            'requirement': '',
            'question': 'Was a commissioning plan developed prior to start of commissioning?',
            'documentation': 'Extracts of commissioning plan',
            'section_id': 'C&T'
        },
        {
            'criteria': '',
            'requirement': '',
            'question': 'Was commissioning completed per a recognised standard (AIRAH DA27, ASHRAE, CIBSE, SA TA 5342)?',
            'documentation': 'Extracts from commissioning report',
            'section_id': 'C&T'
        },
        {
            'criteria': '',
            'requirement': 'Airtightness',
            'question': 'Were airtightness targets set prior to schematic design based on ATTMA Australia Guide?',
            'documentation': 'Evidence of how airtightness targets were determined',
            'section_id': 'C&T'
        },
        {
            'criteria': '',
            'requirement': '',
            'question': 'Was the air barrier system schematic reviewed prior to end of design development?',
            'documentation': 'Evidence of air barrier system schematic review (marked up drawings)',
            'section_id': 'C&T'
        },
        {
            'criteria': '',
            'requirement': '',
            'question': 'Was an approach to delivering airtightness developed prior to construction?',
            'documentation': 'Airtightness testing and commissioning plan',
            'section_id': 'C&T'
        },
        {
            'criteria': '',
            'requirement': '',
            'question': 'Was airtightness testing carried out per AS/NZS ISO 9972 Method 1 by ATTMA member?',
            'documentation': 'Building airtightness testing report or ATTMA Green Star Commissioning report',
            'section_id': 'C&T'
        },
        {
            'criteria': '',
            'requirement': '',
            'question': 'Were test results shared with building owner (regardless of targets being met)?',
            'documentation': 'Signed confirmation from building owner',
            'section_id': 'C&T'
        },
        {
            'criteria': '',
            'requirement': '',
            'question': 'Were airtightness test results included in energy modelling?',
            'documentation': 'Extracts from energy modelling report',
            'section_id': 'C&T'
        },
        {
            'criteria': '',
            'requirement': 'Building Systems Tuning',
            'question': 'Has building owner/developer committed to tuning process with quarterly adjustments for 12 months?',
            'documentation': 'Commitment/contract from building owner and head contractor',
            'section_id': 'C&T'
        },
        {
            'criteria': '',
            'requirement': '',
            'question': 'Does commitment include a building tuning manual/plan with roles and responsibilities?',
            'documentation': 'Extracts of building tuning plan',
            'section_id': 'C&T'
        },
        {
            'criteria': '',
            'requirement': '',
            'question': 'Does the building tuning team include FM, ICA/owner rep, head contractor and relevant subcontractors?',
            'documentation': '',
            'section_id': 'C&T',
            'is_section_end': True
        },
        # Building Information
        {
            'criteria': 'Building Information',
            'requirement': 'Operations and Maintenance Information',
            'question': 'Has O&M information been provided to building owner for all nominated building systems?',
            'documentation': 'Extracts of operations and maintenance information',
            'is_subsection': True,
            'section_id': 'BI'
        },
        {
            'criteria': '',
            'requirement': '',
            'question': 'Does O&M info include maintenance instructions, procedures, schedules, service contacts, warranties and as-built drawings?',
            'documentation': '',
            'section_id': 'BI'
        },
        {
            'criteria': '',
            'requirement': '',
            'question': 'Do appropriate user groups have access to information for best practice environmental outcomes?',
            'documentation': '',
            'section_id': 'BI'
        },
        {
            'criteria': '',
            'requirement': '',
            'question': 'Does O&M info include guidance on keeping information up to date and validating alerts/faults?',
            'documentation': '',
            'section_id': 'BI'
        },
        {
            'criteria': '',
            'requirement': 'Building Logbook',
            'question': 'Has a building logbook been developed per CIBSE TM31 for all nominated building systems?',
            'documentation': 'Building logbook',
            'section_id': 'BI'
        },
        {
            'criteria': '',
            'requirement': '',
            'question': 'Was the logbook presented to building owner prior to occupation?',
            'documentation': 'Evidence of provision to building owner',
            'section_id': 'BI'
        },
        {
            'criteria': '',
            'requirement': 'Building User Information',
            'question': 'Has building user information been provided to building owner and FM prior to occupation?',
            'documentation': 'Building user information',
            'section_id': 'BI'
        },
        {
            'criteria': '',
            'requirement': '',
            'question': 'Is the building user information publicly available to intended users?',
            'documentation': '',
            'section_id': 'BI'
        },
        {
            'criteria': '',
            'requirement': '',
            'question': 'Is the information communicated in a way easily understood by the target audience?',
            'documentation': '',
            'section_id': 'BI'
        },
        {
            'criteria': '',
            'requirement': '',
            'question': 'Is the information in an editable, digital format accessible for updates by FM?',
            'documentation': 'Evidence of editable format provision',
            'section_id': 'BI',
            'is_section_end': True
        },
        # CREDIT ACHIEVEMENT SECTION
        {
            'criteria': 'CREDIT ACHIEVEMENT (1 point)',
            'requirement': '',
            'question': '',
            'documentation': '',
            'is_section': True,
            'section_id': 'CREDIT'
        },
        # Soft Landings Approach
        {
            'criteria': 'Soft Landings Approach',
            'requirement': 'CIBSE ANZ Framework',
            'question': 'Have Stages 1 to 4 of the CIBSE ANZ Soft Landings Framework been implemented?',
            'documentation': 'Evidence of CIBSE ANZ framework implementation',
            'is_subsection': True,
            'section_id': 'SL'
        },
        {
            'criteria': '',
            'requirement': '',
            'question': 'Are sample worksheets from CIBSE ANZ for Stages 1-3 completed?',
            'documentation': 'Completed worksheets',
            'section_id': 'SL'
        },
        {
            'criteria': '',
            'requirement': '',
            'question': 'Are actions for Stage 4 identified?',
            'documentation': '',
            'section_id': 'SL'
        },
        {
            'criteria': '',
            'requirement': 'FM Team Involvement',
            'question': 'Was the FM team (or building owner rep) involved in commissioning and handover process?',
            'documentation': '',
            'section_id': 'SL'
        },
        {
            'criteria': '',
            'requirement': '',
            'question': 'Did FM team take part in developing brief technical guide and O&M manual?',
            'documentation': '',
            'section_id': 'SL'
        },
        {
            'criteria': '',
            'requirement': '',
            'question': 'Did FM team sign-off on the O&M manual?',
            'documentation': 'Signed O&M manual',
            'section_id': 'SL'
        },
        {
            'criteria': '',
            'requirement': '',
            'question': 'Was FM team trained before handover (including BMS demonstration)?',
            'documentation': 'Training records',
            'section_id': 'SL'
        },
        {
            'criteria': '',
            'requirement': '',
            'question': 'Does FM team have continued access to design/construction team for 2 years post PC?',
            'documentation': 'Contractual evidence of continued access',
            'section_id': 'SL',
            'is_section_end': True
        },
        # Independent Commissioning Agent
        {
            'criteria': 'Independent Commissioning Agent',
            'requirement': 'ICA Appointment',
            'question': 'Was an ICA appointed prior to design development?',
            'documentation': 'Letter from building owner confirming ICA appointment',
            'is_subsection': True,
            'section_id': 'ICA'
        },
        {
            'criteria': '',
            'requirement': 'ICA Qualifications',
            'question': 'Is the ICA an advocate for and reporting directly to the project owner?',
            'documentation': 'CV of ICA with qualifications and experience',
            'section_id': 'ICA'
        },
        {
            'criteria': '',
            'requirement': '',
            'question': 'Is the ICA independent of any consultant/contractor involved in design or installation?',
            'documentation': '',
            'section_id': 'ICA'
        },
        {
            'criteria': '',
            'requirement': '',
            'question': 'Is the ICA a registered professional engineer or qualified technician with demonstrated commissioning competency?',
            'documentation': '',
            'section_id': 'ICA'
        },
        {
            'criteria': '',
            'requirement': '',
            'question': 'Does the ICA have experience commissioning at least 2 similar projects?',
            'documentation': '',
            'section_id': 'ICA'
        },
        {
            'criteria': '',
            'requirement': 'ICA Involvement',
            'question': 'Did ICA advise, monitor and verify commissioning/tuning throughout design development, tender, construction, commissioning and tuning phases?',
            'documentation': 'Evidence of ICA involvement from design stage through to tuning',
            'section_id': 'ICA',
            'is_section_end': True
        },
    ]

    # Track section row ranges for summary formulas
    section_ranges = {}
    current_section = None
    section_start_row = None

    # Write data to worksheet
    row = 5
    for item in form_data:
        is_section = item.get('is_section', False)
        is_subsection = item.get('is_subsection', False)
        is_section_end = item.get('is_section_end', False)
        section_id = item.get('section_id', '')

        if is_section:
            # Close previous section if exists
            if current_section and section_start_row:
                section_ranges[current_section] = (section_start_row, row - 1)

            ws.merge_cells(f'A{row}:G{row}')
            cell = ws.cell(row=row, column=1, value=item['criteria'])
            cell.font = section_font
            cell.fill = section_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            current_section = None
            section_start_row = None
        else:
            # Start new trackable section
            if is_subsection and section_id:
                if current_section and section_start_row:
                    section_ranges[current_section] = (section_start_row, row - 1)
                current_section = section_id
                section_start_row = row

            ws.cell(row=row, column=1, value=item['criteria']).alignment = wrap_alignment
            ws.cell(row=row, column=2, value=item['requirement']).alignment = wrap_alignment
            ws.cell(row=row, column=3, value=item['question']).alignment = wrap_alignment

            # Compliance Status column (D) - will have dropdown
            status_cell = ws.cell(row=row, column=4, value='')
            status_cell.alignment = Alignment(horizontal='center', vertical='top')
            dv.add(status_cell)

            # Score column (E) - formula to convert status to numeric score
            # Complies = 0, Minor NC = 1, Major NC = 3, N/A = blank
            score_formula = f'=IF(D{row}="Complies",0,IF(D{row}="Minor NC",1,IF(D{row}="Major NC",3,IF(D{row}="N/A","",""))))'
            ws.cell(row=row, column=5, value=score_formula).alignment = Alignment(horizontal='center', vertical='top')

            ws.cell(row=row, column=6, value='').alignment = wrap_alignment  # Evidence provided
            ws.cell(row=row, column=7, value=item['documentation']).alignment = wrap_alignment

            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                if is_subsection:
                    cell.fill = subsection_fill
                    if col == 1:
                        cell.font = Font(bold=True)

            # Close section at end marker
            if is_section_end and current_section:
                section_ranges[current_section] = (section_start_row, row)
                current_section = None
                section_start_row = None

        row += 1

    last_data_row = row - 1

    # Add conditional formatting to Compliance Status column
    # Complies - Green
    ws.conditional_formatting.add(
        f'D5:D{last_data_row}',
        CellIsRule(operator='equal', formula=['"Complies"'], fill=complies_fill)
    )
    # Minor NC - Yellow/Amber
    ws.conditional_formatting.add(
        f'D5:D{last_data_row}',
        CellIsRule(operator='equal', formula=['"Minor NC"'], fill=minor_nc_fill)
    )
    # Major NC - Red
    ws.conditional_formatting.add(
        f'D5:D{last_data_row}',
        CellIsRule(operator='equal', formula=['"Major NC"'], fill=major_nc_fill)
    )
    # N/A - Grey
    ws.conditional_formatting.add(
        f'D5:D{last_data_row}',
        CellIsRule(operator='equal', formula=['"N/A"'], fill=na_fill)
    )

    # Set row heights for better readability
    for r in range(5, row):
        ws.row_dimensions[r].height = 30

    # ==========================================
    # CREATE DASHBOARD SHEET
    # ==========================================

    ws_dashboard.column_dimensions['A'].width = 35
    ws_dashboard.column_dimensions['B'].width = 15
    ws_dashboard.column_dimensions['C'].width = 15
    ws_dashboard.column_dimensions['D'].width = 15
    ws_dashboard.column_dimensions['E'].width = 18
    ws_dashboard.column_dimensions['F'].width = 25

    # Dashboard Title
    ws_dashboard.merge_cells('A1:F1')
    ws_dashboard['A1'] = "VERIFICATION AND HANDOVER - COMPLIANCE DASHBOARD"
    ws_dashboard['A1'].font = Font(bold=True, size=16)
    ws_dashboard['A1'].alignment = Alignment(horizontal='center')

    ws_dashboard.merge_cells('A2:F2')
    ws_dashboard['A2'] = "Assessment Summary - Scores auto-calculate as you complete the Assessment sheet"
    ws_dashboard['A2'].font = Font(italic=True, size=10)
    ws_dashboard['A2'].alignment = Alignment(horizontal='center')

    # Scoring Legend
    ws_dashboard['A4'] = "SCORING LEGEND:"
    ws_dashboard['A4'].font = Font(bold=True)
    ws_dashboard['A5'] = "Complies = 0 points"
    ws_dashboard['A5'].fill = complies_fill
    ws_dashboard['A6'] = "Minor NC = 1 point"
    ws_dashboard['A6'].fill = minor_nc_fill
    ws_dashboard['A7'] = "Major NC = 3 points"
    ws_dashboard['A7'].fill = major_nc_fill
    ws_dashboard['A8'] = "N/A = Not counted"
    ws_dashboard['A8'].fill = na_fill

    ws_dashboard.merge_cells('C4:F4')
    ws_dashboard['C4'] = "THRESHOLDS: 3+ Minor NCs in a section = equivalent to 1 Major NC"
    ws_dashboard['C4'].font = Font(bold=True, italic=True)

    ws_dashboard.merge_cells('C5:F6')
    ws_dashboard['C5'] = "Section Status: ON TRACK (score 0-2) | AT RISK (score 3-5) | NOT COMPLIANT (score 6+, or any Major NC)"
    ws_dashboard['C5'].alignment = wrap_alignment

    # Section Summary Headers
    dash_row = 10
    dash_headers = ['Criteria Section', 'Complies', 'Minor NC', 'Major NC', 'Total Score', 'Section Status']
    for col, header in enumerate(dash_headers, 1):
        cell = ws_dashboard.cell(row=dash_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Section data with formulas
    sections_info = [
        ('Metering and Monitoring', 'M&M'),
        ('Commissioning and Tuning', 'C&T'),
        ('Building Information', 'BI'),
        ('Soft Landings Approach', 'SL'),
        ('Independent Commissioning Agent', 'ICA'),
    ]

    dash_row = 11
    for section_name, section_id in sections_info:
        if section_id in section_ranges:
            start_row, end_row = section_ranges[section_id]

            ws_dashboard.cell(row=dash_row, column=1, value=section_name).border = thin_border

            # Count Complies
            complies_formula = f'=COUNTIF(Assessment!D{start_row}:D{end_row},"Complies")'
            ws_dashboard.cell(row=dash_row, column=2, value=complies_formula).border = thin_border
            ws_dashboard.cell(row=dash_row, column=2).alignment = Alignment(horizontal='center')

            # Count Minor NC
            minor_formula = f'=COUNTIF(Assessment!D{start_row}:D{end_row},"Minor NC")'
            ws_dashboard.cell(row=dash_row, column=3, value=minor_formula).border = thin_border
            ws_dashboard.cell(row=dash_row, column=3).alignment = Alignment(horizontal='center')

            # Count Major NC
            major_formula = f'=COUNTIF(Assessment!D{start_row}:D{end_row},"Major NC")'
            ws_dashboard.cell(row=dash_row, column=4, value=major_formula).border = thin_border
            ws_dashboard.cell(row=dash_row, column=4).alignment = Alignment(horizontal='center')

            # Total Score (sum of score column, treating 3 minor = 1 major equivalent)
            score_formula = f'=SUM(Assessment!E{start_row}:E{end_row})'
            ws_dashboard.cell(row=dash_row, column=5, value=score_formula).border = thin_border
            ws_dashboard.cell(row=dash_row, column=5).alignment = Alignment(horizontal='center')

            # Section Status formula
            # ON TRACK: score 0-2 and no Major NC
            # AT RISK: score 3-5 and no Major NC
            # NOT COMPLIANT: score 6+ OR any Major NC
            status_formula = f'=IF(D{dash_row}>0,"NOT COMPLIANT",IF(E{dash_row}>=6,"NOT COMPLIANT",IF(E{dash_row}>=3,"AT RISK","ON TRACK")))'
            status_cell = ws_dashboard.cell(row=dash_row, column=6, value=status_formula)
            status_cell.border = thin_border
            status_cell.alignment = Alignment(horizontal='center')

            dash_row += 1

    # Add conditional formatting to status column
    status_range = f'F11:F{dash_row-1}'
    ws_dashboard.conditional_formatting.add(
        status_range,
        CellIsRule(operator='equal', formula=['"ON TRACK"'], fill=complies_fill)
    )
    ws_dashboard.conditional_formatting.add(
        status_range,
        CellIsRule(operator='equal', formula=['"AT RISK"'], fill=minor_nc_fill)
    )
    ws_dashboard.conditional_formatting.add(
        status_range,
        CellIsRule(operator='equal', formula=['"NOT COMPLIANT"'], fill=major_nc_fill)
    )

    # Overall Summary Section
    dash_row += 2
    ws_dashboard.merge_cells(f'A{dash_row}:F{dash_row}')
    ws_dashboard.cell(row=dash_row, column=1, value="OVERALL COMPLIANCE SUMMARY").font = Font(bold=True, size=12)
    ws_dashboard.cell(row=dash_row, column=1).fill = section_fill

    dash_row += 1
    # Minimum Expectation Status (must have M&M, C&T, and BI all on track or at risk with no Major NC)
    ws_dashboard.cell(row=dash_row, column=1, value="Minimum Expectation Status:").font = Font(bold=True)
    min_exp_formula = '=IF(OR(F11="NOT COMPLIANT",F12="NOT COMPLIANT",F13="NOT COMPLIANT"),"NOT COMPLIANT",IF(OR(F11="AT RISK",F12="AT RISK",F13="AT RISK"),"AT RISK","ON TRACK"))'
    min_exp_cell = ws_dashboard.cell(row=dash_row, column=2, value=min_exp_formula)
    min_exp_cell.alignment = Alignment(horizontal='center')
    min_exp_cell.font = Font(bold=True)
    ws_dashboard.conditional_formatting.add(
        f'B{dash_row}',
        CellIsRule(operator='equal', formula=['"ON TRACK"'], fill=complies_fill)
    )
    ws_dashboard.conditional_formatting.add(
        f'B{dash_row}',
        CellIsRule(operator='equal', formula=['"AT RISK"'], fill=minor_nc_fill)
    )
    ws_dashboard.conditional_formatting.add(
        f'B{dash_row}',
        CellIsRule(operator='equal', formula=['"NOT COMPLIANT"'], fill=major_nc_fill)
    )

    dash_row += 1
    ws_dashboard.cell(row=dash_row, column=1, value="Credit Achievement - Soft Landings:").font = Font(bold=True)
    sl_status_cell = ws_dashboard.cell(row=dash_row, column=2, value='=F14')
    sl_status_cell.alignment = Alignment(horizontal='center')
    sl_status_cell.font = Font(bold=True)
    ws_dashboard.conditional_formatting.add(
        f'B{dash_row}',
        CellIsRule(operator='equal', formula=['"ON TRACK"'], fill=complies_fill)
    )
    ws_dashboard.conditional_formatting.add(
        f'B{dash_row}',
        CellIsRule(operator='equal', formula=['"AT RISK"'], fill=minor_nc_fill)
    )
    ws_dashboard.conditional_formatting.add(
        f'B{dash_row}',
        CellIsRule(operator='equal', formula=['"NOT COMPLIANT"'], fill=major_nc_fill)
    )

    dash_row += 1
    ws_dashboard.cell(row=dash_row, column=1, value="Credit Achievement - ICA:").font = Font(bold=True)
    ica_status_cell = ws_dashboard.cell(row=dash_row, column=2, value='=F15')
    ica_status_cell.alignment = Alignment(horizontal='center')
    ica_status_cell.font = Font(bold=True)
    ws_dashboard.conditional_formatting.add(
        f'B{dash_row}',
        CellIsRule(operator='equal', formula=['"ON TRACK"'], fill=complies_fill)
    )
    ws_dashboard.conditional_formatting.add(
        f'B{dash_row}',
        CellIsRule(operator='equal', formula=['"AT RISK"'], fill=minor_nc_fill)
    )
    ws_dashboard.conditional_formatting.add(
        f'B{dash_row}',
        CellIsRule(operator='equal', formula=['"NOT COMPLIANT"'], fill=major_nc_fill)
    )

    # Notes
    dash_row += 2
    ws_dashboard.merge_cells(f'A{dash_row}:F{dash_row}')
    ws_dashboard.cell(row=dash_row, column=1, value="NOTES:").font = Font(bold=True)

    notes = [
        "1. For Minimum Expectation: All three criteria (M&M, C&T, BI) must be ON TRACK or AT RISK (no Major NCs).",
        "2. For Credit Achievement: Either Soft Landings OR ICA must be ON TRACK (plus Minimum Expectation).",
        "3. For projects >$20m Total Building Services Value: BOTH Soft Landings AND ICA required.",
        "4. AT RISK status indicates minor issues that should be addressed before final submission.",
        "5. Any Major NC in a section results in NOT COMPLIANT status for that section.",
        "6. 3 or more Minor NCs (score of 3+) without a Major NC results in AT RISK status.",
    ]

    for note in notes:
        dash_row += 1
        ws_dashboard.merge_cells(f'A{dash_row}:F{dash_row}')
        ws_dashboard.cell(row=dash_row, column=1, value=note).alignment = wrap_alignment

    # Save workbook
    filename = '/home/user/submissionforms/Verification_and_Handover_Submission_Form.xlsx'
    wb.save(filename)
    print(f"Created: {filename}")
    return filename

if __name__ == "__main__":
    create_submission_form()
