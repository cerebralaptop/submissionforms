#!/usr/bin/env python3
"""Parse the Green Star Buildings Excel and generate a complete submission website."""

import json
import html as html_mod
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ── Parse Excel ──────────────────────────────────────────────────────────────
wb = load_workbook("Green_Star_Buildings_v1.1_Submission_Questions.xlsx")

# Category mapping for each sheet
CATEGORIES = {
    "Responsible": [
        "Industry Development", "Responsible Construction",
        "Verification and Handover", "Responsible Resource Mgmt",
        "Responsible Procurement", "Responsible Structure",
        "Responsible Envelope", "Responsible Systems",
        "Responsible Finishes", "Impacts Disclosure",
    ],
    "Healthy": [
        "Clean Air", "Light Quality", "Acoustic Comfort",
        "Exposure to Toxins", "Amenity and Comfort", "Connection to Nature",
    ],
    "Resilient": [
        "Climate Resilience", "Operations Resilience",
        "Community Resilience", "Heat Resilience", "Grid Resilience",
    ],
    "Positive": [
        "Energy Source", "Energy Use", "Upfront Carbon Reduction",
        "Upfront Carbon Compensation", "Refrigerant Systems Impacts",
        "Low-Emissions Transport", "Design for Circularity", "Water Use",
    ],
    "Places": [
        "Movement and Place", "Enjoyable Places",
        "Contribution to Place", "Culture Heritage Identity",
    ],
    "People": [
        "Inclusive Construction", "First Nations Inclusion",
        "Procurement Workforce Inclusion", "Design for Equity",
    ],
    "Nature": [
        "Impacts to Nature", "Biodiversity Enhancement",
        "Nature Connectivity", "Nature Stewardship", "Waterway Protection",
    ],
    "Leadership": [
        "Market Transformation", "Leadership Challenges",
    ],
}

# Parse all sheets
all_credits = []
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    credit_data = {
        "sheet_name": sheet_name,
        "sections": [],   # list of {type, title, questions}
        "questions": [],
    }

    current_section = None
    for row_idx in range(2, ws.max_row + 1):
        a = ws.cell(row=row_idx, column=1).value
        b = ws.cell(row=row_idx, column=2).value
        c = ws.cell(row=row_idx, column=3).value
        d = ws.cell(row=row_idx, column=4).value
        e = ws.cell(row=row_idx, column=5).value
        f = ws.cell(row=row_idx, column=6).value
        h = ws.cell(row=row_idx, column=8).value

        # Check if it's a merged header row
        fill_color = ws.cell(row=row_idx, column=1).fill.start_color
        font = ws.cell(row=row_idx, column=1).font

        if a and not b and not e:
            # This is a header row (credit header, level header, or criteria header)
            text = str(a)
            if font.color and hasattr(font.color, 'rgb') and font.color.rgb:
                rgb = str(font.color.rgb)
                if rgb == "00FFFFFF" and font.size == 12:
                    # Credit header (green bg, white text, size 12)
                    credit_data["title"] = text
                    continue
                elif rgb == "00FFFFFF" and font.size == 11:
                    # Level header
                    current_section = {"type": "level", "title": text, "criteria": []}
                    credit_data["sections"].append(current_section)
                    continue
                elif "1F4E28" in rgb:
                    # Criteria header (light green bg, dark text)
                    if current_section is not None:
                        current_section["criteria"].append({"name": text, "questions": []})
                    continue
            # Fallback detection by font size/bold
            if font.bold and font.size and font.size >= 12:
                credit_data["title"] = text
                continue
            elif font.bold and font.size and font.size >= 11:
                if font.color and hasattr(font.color, 'rgb') and "1F4E28" in str(font.color.rgb):
                    if current_section is not None:
                        current_section["criteria"].append({"name": text, "questions": []})
                    continue
                current_section = {"type": "level", "title": text, "criteria": []}
                credit_data["sections"].append(current_section)
                continue

        # It's a question row if columns E and F have content
        if e and f:
            q = {
                "ref": str(a) if a else "",
                "credit": str(b) if b else "",
                "level": str(c) if c else "",
                "criteria": str(d) if d else "",
                "type": str(e) if e else "",
                "question": str(f) if f else "",
                "data_note": str(h) if h else "",
            }
            credit_data["questions"].append(q)
            # Add to current criteria; create a "General" one if none exists yet
            if current_section:
                if not current_section["criteria"]:
                    current_section["criteria"].append({"name": "General", "questions": []})
                current_section["criteria"][-1]["questions"].append(q)

    all_credits.append(credit_data)

# Map credits to categories
def find_category(sheet_name):
    for cat, sheets in CATEGORIES.items():
        for s in sheets:
            if s.lower().replace(" ", "") in sheet_name.lower().replace(" ", ""):
                return cat
    return "Other"

for c in all_credits:
    c["category"] = find_category(c["sheet_name"])

# Build JSON-serializable credit structure for the Excel exporter
credits_json_data = []
for i, c in enumerate(all_credits):
    credits_json_data.append({
        "id": f"credit-{i}",
        "sheet_name": c["sheet_name"],
        "title": c.get("title", c["sheet_name"]),
        "category": c["category"],
        "sections": [
            {
                "title": sec["title"],
                "criteria": [
                    {
                        "name": cr["name"],
                        "questions": [
                            {
                                "ref": q["ref"],
                                "credit": q["credit"],
                                "level": q["level"],
                                "criteria": q["criteria"],
                                "type": q["type"],
                                "question": q["question"],
                                "data_note": q["data_note"],
                                "input_id": f"credit-{i}-{q['ref'].replace('.', '-')}",
                            }
                            for q in cr["questions"]
                        ],
                    }
                    for cr in sec["criteria"]
                ],
            }
            for sec in c["sections"]
        ],
    })
credits_json_str = json.dumps(credits_json_data)

# ── Conditional visibility rules ─────────────────────────────────────────────
# Maps: question_input_id -> { "depends_on": gateway_input_id, "show_when": "Yes"|"No" }
# Built from the analysis of all Y/N gateway -> follow-up patterns.
#
# We generate rules keyed by credit index + ref to match the input IDs
# Format of input IDs: credit-{idx}-{ref with dots replaced by dashes}
def make_id(credit_idx, ref):
    return f"credit-{credit_idx}-{ref.replace('.', '-')}"

conditional_rules = {}
for ci, c in enumerate(all_credits):
    sheet = c["sheet_name"]
    qs = c["questions"]
    refs = {q["ref"]: q for q in qs}
    ref_list = [q["ref"] for q in qs]

    # Build ref -> index for this credit
    ref_idx = {r: i for i, r in enumerate(ref_list)}

    # Define rules per credit based on analysis
    # Explicit "If yes/no" follow-ups
    rules_for_credit = []

    # Helper: add rule(s) for gateway_ref -> [follow_ref1, ...] shown when answer is val
    def add_rules(gateway, followers, val="Yes"):
        for f in followers:
            if gateway in ref_idx and f in ref_idx:
                rules_for_credit.append((f, gateway, val))

    sn = sheet.lower().replace(" ", "")

    if "industrydevelopment" in sn:
        add_rules("ID.5", ["ID.6"], "Yes")
    elif "responsibleconstruction" in sn:
        add_rules("RC.1", ["RC.3"], "Yes")
        add_rules("RC.1", ["RC.2"], "No")
        add_rules("RC.4", ["RC.5"], "Yes")
    elif "verificationandhandover" in sn:
        add_rules("VH.5", ["VH.6"], "Yes")
        add_rules("VH.7", ["VH.8"], "Yes")
        add_rules("VH.26", ["VH.27"], "Yes")
    elif "responsibleresource" in sn:
        add_rules("RRM.2", ["RRM.3"], "Yes")
        add_rules("RRM.5", ["RRM.6"], "Yes")
        add_rules("RRM.7", ["RRM.8"], "Yes")
        add_rules("RRM.12", ["RRM.13"], "Yes")
    elif "responsibleprocurement" in sn:
        add_rules("RP.12", ["RP.13"], "Yes")
    elif "cleanair" in sn:
        add_rules("CA.9", ["CA.10"], "Yes")
    elif "lightquality" in sn:
        add_rules("LQ.7", ["LQ.8"], "Yes")
        add_rules("LQ.7", ["LQ.9", "LQ.10", "LQ.11"], "No")
    elif "exposuretotoxins" in sn:
        add_rules("ET.1", ["ET.2", "ET.3"], "Yes")
    elif "amenityandcomfort" in sn or "amenity" in sn:
        add_rules("AmC.3", ["AmC.4"], "Yes")
    elif "connectiontonature" in sn:
        add_rules("CN.4", ["CN.5", "CN.6"], "Yes")
    elif "climateresilience" in sn:
        add_rules("CR.1", ["CR.2", "CR.3", "CR.4"], "Yes")
    elif "operationsresilience" in sn:
        add_rules("OR.4", ["OR.5"], "Yes")
        add_rules("OR.6", ["OR.7"], "Yes")
    elif "communityresilience" in sn:
        add_rules("CoR.1", ["CoR.2", "CoR.3", "CoR.4"], "Yes")
    elif "gridresilience" in sn:
        add_rules("GR.1", ["GR.2", "GR.3"], "Yes")
        add_rules("GR.4", ["GR.5", "GR.6"], "Yes")
        add_rules("GR.7", ["GR.8"], "Yes")
    elif "energysource" in sn:
        add_rules("ES.5", ["ES.6"], "Yes")
    elif "upfrontcarbonreduction" in sn:
        add_rules("UCR.2", ["UCR.3", "UCR.4", "UCR.5"], "Yes")
    elif "wateruse" in sn:
        add_rules("WU.3", ["WU.4"], "Yes")
        add_rules("WU.5", ["WU.6"], "Yes")
    elif "contributiontoplace" in sn:
        add_rules("CP.1", ["CP.2"], "Yes")
    elif "cultureheritage" in sn:
        add_rules("CHI.1", ["CHI.2"], "Yes")
    elif "firstnations" in sn:
        add_rules("FNI.1", ["FNI.2", "FNI.3"], "Yes")
    elif "designforequity" in sn:
        add_rules("DE.4", ["DE.5"], "Yes")
    elif "impactstonature" in sn:
        add_rules("IN.1", ["IN.2", "IN.3"], "Yes")
        add_rules("IN.7", ["IN.8"], "Yes")
    elif "natureconnectivity" in sn:
        add_rules("NC.1", ["NC.2"], "Yes")
        add_rules("NC.5", ["NC.6"], "Yes")
    elif "naturestewardship" in sn:
        add_rules("NS.1", ["NS.2", "NS.3"], "Yes")
    elif "markettransformation" in sn:
        add_rules("MT.4", ["MT.5"], "Yes")
    elif "waterwayprotection" in sn:
        add_rules("WP.5", ["WP.6"], "Yes")
    elif "impactsdisclosure" in sn:
        add_rules("ID2.5", ["ID2.6"], "Yes")

    for follower_ref, gateway_ref, show_val in rules_for_credit:
        fid = make_id(ci, follower_ref)
        gid = make_id(ci, gateway_ref)
        conditional_rules[fid] = {"depends_on": gid, "show_when": show_val}

conditional_rules_json = json.dumps(conditional_rules)

# Build search index for client-side search
search_index = []
for ci, c in enumerate(all_credits):
    credit_id = f"credit-{ci}"
    for q in c["questions"]:
        q_id = f"{credit_id}-{q['ref'].replace('.', '-')}"
        search_index.append({
            "ref": q["ref"],
            "credit": c["sheet_name"],
            "creditId": credit_id,
            "cardId": f"card-{q_id}",
            "type": q["type"],
            "question": q["question"],
            "note": q["data_note"],
        })
search_index_json = json.dumps(search_index)

# ── Generate HTML ────────────────────────────────────────────────────────────
def esc(s):
    return html_mod.escape(str(s)) if s else ""

category_colors = {
    "Responsible": {"bg": "#1F4E28", "light": "#E8F5E9", "mid": "#A5D6A7"},
    "Healthy": {"bg": "#1565C0", "light": "#E3F2FD", "mid": "#90CAF9"},
    "Resilient": {"bg": "#E65100", "light": "#FFF3E0", "mid": "#FFCC80"},
    "Positive": {"bg": "#2E7D32", "light": "#F1F8E9", "mid": "#C5E1A5"},
    "Places": {"bg": "#6A1B9A", "light": "#F3E5F5", "mid": "#CE93D8"},
    "People": {"bg": "#C62828", "light": "#FFEBEE", "mid": "#EF9A9A"},
    "Nature": {"bg": "#00695C", "light": "#E0F2F1", "mid": "#80CBC4"},
    "Leadership": {"bg": "#F57F17", "light": "#FFFDE7", "mid": "#FFF176"},
}

category_icons = {
    "Responsible": "&#9878;",   # recycling
    "Healthy": "&#9829;",      # heart
    "Resilient": "&#9730;",    # umbrella
    "Positive": "&#9889;",     # lightning
    "Places": "&#9962;",       # building
    "People": "&#9823;",       # person
    "Nature": "&#9752;",       # leaf
    "Leadership": "&#9733;",   # star
}

# Build sidebar and pages
sidebar_html = ""
pages_html = ""
credit_index = 0

for cat_name, cat_sheets in CATEGORIES.items():
    colors = category_colors[cat_name]
    icon = category_icons[cat_name]
    cat_credits = [c for c in all_credits if c["category"] == cat_name]

    sidebar_html += f'''
    <div class="sidebar-category">
      <div class="sidebar-category-header" style="background:{colors['bg']}" onclick="toggleCategory(this)">
        <span>{icon} {esc(cat_name)}</span>
        <span class="arrow">&#9662;</span>
      </div>
      <div class="sidebar-category-items">'''

    for credit in cat_credits:
        credit_id = f"credit-{credit_index}"
        q_count = len(credit["questions"])
        sidebar_html += f'''
        <div class="sidebar-item" data-credit="{credit_id}" id="sidebar-{credit_id}">
          <span class="sidebar-item-name" onclick="showCredit('{credit_id}')">{esc(credit["sheet_name"])}</span>
          <span class="sidebar-progress-ring" id="ring-{credit_id}"><svg width="18" height="18" viewBox="0 0 18 18"><circle cx="9" cy="9" r="7" fill="none" stroke="#e0e0e0" stroke-width="2"/><circle cx="9" cy="9" r="7" fill="none" stroke="{colors['bg']}" stroke-width="2" stroke-dasharray="44" stroke-dashoffset="44" stroke-linecap="round" transform="rotate(-90 9 9)" class="ring-fill"/></svg></span>
          <button class="na-toggle" onclick="toggleNA('{credit_id}', event)" title="Mark as Not Applicable">N/A</button>
        </div>'''

        # Build credit page
        title = credit.get("title", credit["sheet_name"])
        pages_html += f'''
    <div class="credit-page" id="{credit_id}" style="display:none">
      <div class="credit-header" style="background:{colors['bg']}">
        <div class="credit-header-top">
          <span class="credit-category-tag" style="background:{colors['mid']};color:{colors['bg']}">{esc(cat_name)}</span>
          <div class="credit-header-right">
            <button class="wizard-toggle" onclick="toggleWizard('{credit_id}')" title="Step-by-step mode">Step-by-step</button>
            <button class="na-btn-header" onclick="toggleNA('{credit_id}', event)">Mark N/A</button>
          </div>
        </div>
        <h2>{esc(title)}</h2>
      </div>
      <div class="credit-progress-bar">
        <div class="credit-progress-fill" id="{credit_id}-progress" style="background:{colors['bg']}"></div>
      </div>
      <div class="credit-progress-text">
        <span id="{credit_id}-progress-text">0 of {q_count} answered</span>
      </div>
      <div class="wizard-nav" id="{credit_id}-wizard-nav" style="display:none">
        <button class="wizard-btn" onclick="wizardPrev('{credit_id}')">&#8592; Back</button>
        <span class="wizard-step-text" id="{credit_id}-wizard-step">1 / {q_count}</span>
        <button class="wizard-btn wizard-btn-next" onclick="wizardNext('{credit_id}')">Next &#8594;</button>
      </div>
      <div class="credit-body">
        <div class="gaps-panel" id="{credit_id}-gaps">
          <h3>Unanswered Questions</h3>
          <div class="gaps-count" id="{credit_id}-gaps-count"></div>
          <ul class="gaps-list" id="{credit_id}-gaps-list"></ul>
        </div>'''

        for section in credit["sections"]:
            pages_html += f'''
        <div class="level-header" style="background:{colors['bg']}">{esc(section["title"])}</div>'''

            for crit in section["criteria"]:
                pages_html += f'''
        <div class="criteria-header" style="border-left-color:{colors['bg']};background:{colors['light']}">{esc(crit["name"])}</div>'''

                for q in crit["questions"]:
                    q_id = f"{credit_id}-{q['ref'].replace('.', '-')}"
                    type_class = ""
                    input_html = ""

                    if q["type"] == "Condition (Y/N)":
                        type_class = "q-condition"
                        input_html = f'''
              <div class="response-field">
                <select id="{q_id}" class="yn-select" onchange="onAnswer('{credit_id}')" data-credit="{credit_id}">
                  <option value="">-- Select --</option>
                  <option value="Yes">Yes</option>
                  <option value="No">No</option>
                </select>
              </div>'''
                    elif q["type"] == "Data":
                        type_class = "q-data"
                        input_html = f'''
              <div class="response-field">
                <textarea id="{q_id}" class="data-input" rows="3" placeholder="Enter data..." oninput="onAnswer('{credit_id}')" data-credit="{credit_id}"></textarea>
              </div>'''
                    else:
                        type_class = "q-descriptive"
                        input_html = f'''
              <div class="response-field">
                <textarea id="{q_id}" class="desc-input" rows="4" placeholder="Describe..." oninput="onAnswer('{credit_id}')" data-credit="{credit_id}"></textarea>
              </div>'''

                    data_note_html = ""
                    if q["data_note"]:
                        data_note_html = f'''
              <div class="guidance-wrapper">
                <button class="guidance-toggle" onclick="this.parentElement.classList.toggle('open')">
                  <span class="guidance-icon">?</span> Guidance
                  <span class="guidance-arrow">&#9662;</span>
                </button>
                <div class="guidance-content">{esc(q["data_note"])}</div>
              </div>'''

                    type_badge = q["type"]
                    # Check if this question has a conditional dependency
                    dep_attrs = ""
                    is_conditional = q_id in conditional_rules
                    if is_conditional:
                        rule = conditional_rules[q_id]
                        dep_attrs = f' data-depends-on="{rule["depends_on"]}" data-show-when="{rule["show_when"]}"'

                    hidden_class = " q-hidden" if is_conditional else ""
                    pages_html += f'''
        <div class="question-card {type_class}{hidden_class}" id="card-{q_id}"{dep_attrs}>
          <div class="question-header">
            <span class="question-ref">{esc(q["ref"])}</span>
            <span class="question-type-badge {type_class}-badge">{esc(type_badge)}</span>
          </div>
          <div class="question-text">{esc(q["question"])}</div>
          {input_html}
          {data_note_html}
        </div>'''

        pages_html += '''
      </div>
    </div>'''
        credit_index += 1

    sidebar_html += '''
      </div>
    </div>'''

# Build dashboard summary
total_questions = sum(len(c["questions"]) for c in all_credits)
total_credits = len(all_credits)

dashboard_cards = ""
for cat_name, cat_sheets in CATEGORIES.items():
    colors = category_colors[cat_name]
    icon = category_icons[cat_name]
    cat_credits = [c for c in all_credits if c["category"] == cat_name]
    cat_q_count = sum(len(c["questions"]) for c in cat_credits)

    dashboard_cards += f'''
      <div class="dash-card" style="border-top:4px solid {colors['bg']}">
        <div class="dash-card-icon" style="color:{colors['bg']}">{icon}</div>
        <div class="dash-card-title">{esc(cat_name)}</div>
        <div class="dash-card-stats">
          <span>{len(cat_credits)} credits</span>
          <span>{cat_q_count} questions</span>
        </div>
        <div class="dash-card-bar">
          <div class="dash-card-bar-fill" id="dash-{cat_name.lower()}-bar" style="background:{colors['bg']}"></div>
        </div>
        <div class="dash-card-pct" id="dash-{cat_name.lower()}-pct">0% complete</div>
      </div>'''

# ── Full HTML ────────────────────────────────────────────────────────────────
html = f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Green Star Buildings v1.1 — Submission Forms</title>
<style>
*, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}

:root {{
  --sidebar-width: 300px;
  --header-height: 60px;
  --bg: #f5f7f5;
  --text: #1a1a1a;
  --text-light: #555;
  --border: #e0e0e0;
  --white: #ffffff;
  --green-dark: #0D3318;
  --green-primary: #1F4E28;
  --green-mid: #2E7D32;
  --green-light: #E8F5E9;
  --card-bg: #ffffff;
  --input-bg: #ffffff;
  --hover-bg: #E8F5E9;
  --shadow: rgba(0,0,0,0.06);
}}

html.dark {{
  --bg: #1a1d21;
  --text: #e0e0e0;
  --text-light: #999;
  --border: #333;
  --white: #23272e;
  --green-dark: #0f1f14;
  --green-light: #1a2e1f;
  --card-bg: #282c34;
  --input-bg: #2c313a;
  --hover-bg: #1e3325;
  --shadow: rgba(0,0,0,0.2);
}}
html.dark .app-header {{ background: #0a1210; }}
html.dark .sidebar {{ background: #1e2127; }}
html.dark .question-card {{ background: var(--card-bg); border-color: var(--border); }}
html.dark .q-descriptive {{ border-left-color: #388E3C; }}
html.dark .q-data {{ border-left-color: #1E88E5; }}
html.dark .q-condition {{ border-left-color: #9575CD; }}
html.dark .question-card:hover {{ box-shadow: 0 2px 8px rgba(0,0,0,0.2); }}
html.dark textarea, html.dark select {{
  background: var(--input-bg) !important;
  color: var(--text);
  border-color: var(--border);
}}
html.dark textarea:focus, html.dark select:focus {{
  border-color: var(--green-mid);
  box-shadow: 0 0 0 2px rgba(46,125,50,0.25);
}}
html.dark .dash-stat, html.dark .dash-card {{ background: var(--card-bg); box-shadow: 0 1px 4px var(--shadow); }}
html.dark .modal {{ background: var(--card-bg); color: var(--text); }}
html.dark .modal-btn {{ background: var(--input-bg); color: var(--text); border-color: var(--border); }}
html.dark .export-option {{ border-color: var(--border); }}
html.dark .export-option:hover {{ border-color: var(--green-mid); }}
html.dark .history-panel {{ background: var(--card-bg); }}
html.dark .guidance-content {{ background: #2a2520; border-color: #3d3520; color: #bbb; }}
html.dark .guidance-icon {{ background: #3d3520; border-color: #5a4a20; }}
html.dark .gaps-panel {{ background: #2d2518; border-color: #4a3a20; }}
html.dark .criteria-header {{ background: var(--green-light) !important; }}
html.dark .q-descriptive-badge {{ background: #1a2e1f; color: #66BB6A; }}
html.dark .q-data-badge {{ background: #1a2535; color: #64B5F6; }}
html.dark .q-condition-badge {{ background: #251a35; color: #B39DDB; }}
html.dark .credit-page.credit-na::after {{ color: var(--text-light); }}
html.dark ::-webkit-scrollbar-thumb {{ background: #444; }}
html.dark ::-webkit-scrollbar-thumb:hover {{ background: #555; }}
html.dark .save-toast {{ background: var(--green-mid); }}

html {{ height: 100%; }}
body {{
  font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
  background: var(--bg);
  color: var(--text);
  height: 100%;
  overflow: hidden;
}}

/* ── Header ── */
.app-header {{
  position: fixed; top: 0; left: 0; right: 0;
  height: var(--header-height);
  background: var(--green-dark);
  color: white;
  display: flex;
  align-items: center;
  padding: 0 24px;
  z-index: 100;
  box-shadow: 0 2px 8px rgba(0,0,0,0.15);
}}
.app-header h1 {{
  font-size: 18px;
  font-weight: 600;
  letter-spacing: -0.3px;
}}
.app-header .subtitle {{
  margin-left: 12px;
  font-size: 13px;
  opacity: 0.7;
  font-weight: 400;
}}
.header-actions {{
  margin-left: auto;
  display: flex;
  gap: 10px;
}}
.header-btn {{
  background: rgba(255,255,255,0.15);
  color: white;
  border: 1px solid rgba(255,255,255,0.2);
  padding: 6px 14px;
  border-radius: 6px;
  cursor: pointer;
  font-size: 13px;
  transition: background 0.2s;
}}
.header-btn:hover {{ background: rgba(255,255,255,0.25); }}
.header-btn.primary {{ background: var(--green-mid); border-color: var(--green-mid); }}
.header-btn.primary:hover {{ background: #388E3C; }}
.menu-toggle {{
  display: none;
  background: none;
  border: none;
  color: white;
  font-size: 22px;
  cursor: pointer;
  margin-right: 12px;
  padding: 4px;
}}

/* ── Sidebar ── */
.sidebar {{
  position: fixed;
  top: var(--header-height);
  left: 0;
  bottom: 0;
  width: var(--sidebar-width);
  background: var(--white);
  border-right: 1px solid var(--border);
  overflow-y: auto;
  z-index: 50;
  transition: transform 0.3s;
}}
.sidebar-category {{ border-bottom: 1px solid var(--border); }}
.sidebar-category-header {{
  display: flex;
  justify-content: space-between;
  align-items: center;
  padding: 12px 16px;
  color: white;
  font-weight: 600;
  font-size: 13px;
  cursor: pointer;
  user-select: none;
  text-transform: uppercase;
  letter-spacing: 0.5px;
}}
.sidebar-category-header .arrow {{
  transition: transform 0.2s;
  font-size: 10px;
}}
.sidebar-category-header.collapsed .arrow {{
  transform: rotate(-90deg);
}}
.sidebar-category-items {{
  max-height: 800px;
  overflow: hidden;
  transition: max-height 0.3s ease;
}}
.sidebar-category-header.collapsed + .sidebar-category-items {{
  max-height: 0;
}}
.sidebar-item {{
  display: flex;
  justify-content: space-between;
  align-items: center;
  padding: 10px 16px 10px 24px;
  cursor: pointer;
  font-size: 13px;
  color: var(--text);
  transition: background 0.15s;
  border-left: 3px solid transparent;
}}
.sidebar-item:hover {{ background: var(--green-light); }}
.sidebar-item.active {{
  background: var(--green-light);
  border-left-color: var(--green-primary);
  font-weight: 600;
}}
.sidebar-item-name {{
  overflow: hidden;
  text-overflow: ellipsis;
  white-space: nowrap;
  flex: 1;
}}
.sidebar-badge {{
  background: var(--border);
  color: var(--text-light);
  font-size: 11px;
  padding: 2px 7px;
  border-radius: 10px;
  margin-left: 8px;
  flex-shrink: 0;
}}

/* ── Main Content ── */
.main-content {{
  margin-left: var(--sidebar-width);
  margin-top: var(--header-height);
  height: calc(100vh - var(--header-height));
  overflow-y: auto;
  padding: 0;
}}

/* ── Dashboard ── */
.dashboard {{
  padding: 32px;
}}
.dash-title {{
  font-size: 28px;
  font-weight: 700;
  color: var(--green-dark);
  margin-bottom: 4px;
}}
.dash-subtitle {{
  font-size: 14px;
  color: var(--text-light);
  margin-bottom: 24px;
}}
.dash-overview {{
  display: flex;
  gap: 16px;
  margin-bottom: 32px;
  flex-wrap: wrap;
}}
.dash-stat {{
  background: var(--white);
  border-radius: 10px;
  padding: 20px 28px;
  box-shadow: 0 1px 4px rgba(0,0,0,0.06);
  text-align: center;
  flex: 1;
  min-width: 140px;
}}
.dash-stat-number {{
  font-size: 36px;
  font-weight: 700;
  color: var(--green-primary);
}}
.dash-stat-label {{
  font-size: 12px;
  color: var(--text-light);
  text-transform: uppercase;
  letter-spacing: 0.5px;
  margin-top: 4px;
}}
.dash-grid {{
  display: grid;
  grid-template-columns: repeat(auto-fill, minmax(240px, 1fr));
  gap: 16px;
}}
.dash-card {{
  background: var(--white);
  border-radius: 10px;
  padding: 20px;
  box-shadow: 0 1px 4px rgba(0,0,0,0.06);
}}
.dash-card-icon {{ font-size: 28px; margin-bottom: 8px; }}
.dash-card-title {{
  font-size: 16px;
  font-weight: 700;
  margin-bottom: 8px;
}}
.dash-card-stats {{
  display: flex;
  gap: 12px;
  font-size: 12px;
  color: var(--text-light);
  margin-bottom: 12px;
}}
.dash-card-bar {{
  height: 6px;
  background: #eee;
  border-radius: 3px;
  overflow: hidden;
  margin-bottom: 6px;
}}
.dash-card-bar-fill {{
  height: 100%;
  width: 0%;
  border-radius: 3px;
  transition: width 0.4s;
}}
.dash-card-pct {{
  font-size: 12px;
  color: var(--text-light);
}}

/* ── Credit Pages ── */
.credit-page {{ padding: 0; }}
.credit-header {{
  padding: 28px 32px;
  color: white;
}}
.credit-header-top {{
  display: flex;
  justify-content: space-between;
  align-items: center;
  margin-bottom: 10px;
}}
.credit-category-tag {{
  font-size: 11px;
  font-weight: 700;
  padding: 3px 10px;
  border-radius: 4px;
  text-transform: uppercase;
  letter-spacing: 0.5px;
}}
.credit-count {{
  font-size: 13px;
  opacity: 0.8;
}}
.credit-header h2 {{
  font-size: 24px;
  font-weight: 700;
  line-height: 1.3;
}}
.credit-progress-bar {{
  height: 4px;
  background: #eee;
}}
.credit-progress-fill {{
  height: 100%;
  width: 0%;
  transition: width 0.4s;
}}
.credit-progress-text {{
  padding: 8px 32px;
  font-size: 12px;
  color: var(--text-light);
  background: var(--white);
  border-bottom: 1px solid var(--border);
}}
.credit-body {{ padding: 24px 32px 48px; }}

.level-header {{
  color: white;
  padding: 10px 16px;
  border-radius: 6px;
  font-weight: 600;
  font-size: 14px;
  margin: 24px 0 12px 0;
}}
.level-header:first-child {{ margin-top: 0; }}

.criteria-header {{
  padding: 10px 16px;
  border-left: 4px solid;
  border-radius: 0 6px 6px 0;
  font-weight: 600;
  font-size: 13px;
  margin: 16px 0 10px 0;
  color: var(--text);
}}

/* ── Question Cards ── */
.question-card {{
  background: var(--white);
  border: 1px solid var(--border);
  border-radius: 8px;
  padding: 16px 20px;
  margin-bottom: 10px;
  transition: box-shadow 0.2s;
}}
.question-card:hover {{ box-shadow: 0 2px 8px rgba(0,0,0,0.08); }}

.question-header {{
  display: flex;
  justify-content: space-between;
  align-items: center;
  margin-bottom: 8px;
}}
.question-ref {{
  font-size: 12px;
  font-weight: 700;
  color: var(--text-light);
  font-family: monospace;
}}
.question-type-badge {{
  font-size: 10px;
  padding: 2px 8px;
  border-radius: 4px;
  font-weight: 600;
  text-transform: uppercase;
  letter-spacing: 0.3px;
}}
.q-descriptive-badge {{ background: #F1F8E9; color: #33691E; }}
.q-data-badge {{ background: #E3F2FD; color: #1565C0; }}
.q-condition-badge {{ background: #EDE7F6; color: #4A148C; }}

.question-text {{
  font-size: 14px;
  line-height: 1.5;
  color: var(--text);
  margin-bottom: 12px;
}}

.q-condition {{ border-left: 3px solid #7E57C2; }}
.q-data {{ border-left: 3px solid #42A5F5; }}
.q-descriptive {{ border-left: 3px solid #66BB6A; }}

.response-field {{ margin-bottom: 8px; }}

.yn-select {{
  width: 180px;
  padding: 8px 12px;
  border: 1px solid var(--border);
  border-radius: 6px;
  font-size: 14px;
  background: var(--white);
  cursor: pointer;
  color: var(--text);
}}
.yn-select:focus {{ outline: none; border-color: var(--green-primary); box-shadow: 0 0 0 2px rgba(31,78,40,0.15); }}

.desc-input, .data-input {{
  width: 100%;
  padding: 10px 12px;
  border: 1px solid var(--border);
  border-radius: 6px;
  font-size: 13px;
  font-family: inherit;
  resize: vertical;
  color: var(--text);
  line-height: 1.5;
}}
.desc-input:focus, .data-input:focus {{
  outline: none;
  border-color: var(--green-primary);
  box-shadow: 0 0 0 2px rgba(31,78,40,0.15);
}}
.data-input {{ background: #FAFBFF; }}

.data-note {{
  font-size: 12px;
  color: #666;
  background: #FFFDE7;
  padding: 8px 12px;
  border-radius: 4px;
  line-height: 1.4;
  border: 1px solid #FFF9C4;
}}
.data-note-icon {{ margin-right: 4px; }}
.data-note-label {{ font-weight: 600; color: #F57F17; }}

/* ── Export Modal ── */
.modal-overlay {{
  display: none;
  position: fixed;
  top: 0; left: 0; right: 0; bottom: 0;
  background: rgba(0,0,0,0.4);
  z-index: 200;
  justify-content: center;
  align-items: center;
}}
.modal-overlay.active {{ display: flex; }}
.modal {{
  background: white;
  border-radius: 12px;
  padding: 32px;
  width: 500px;
  max-width: 90%;
  box-shadow: 0 20px 60px rgba(0,0,0,0.2);
}}
.modal h3 {{ font-size: 18px; margin-bottom: 16px; }}
.modal p {{ font-size: 14px; color: var(--text-light); margin-bottom: 16px; line-height: 1.5; }}
.modal-actions {{
  display: flex;
  gap: 10px;
  justify-content: flex-end;
}}
.modal-btn {{
  padding: 8px 20px;
  border-radius: 6px;
  font-size: 14px;
  cursor: pointer;
  border: 1px solid var(--border);
  background: white;
  color: var(--text);
}}
.modal-btn:hover {{ background: #f5f5f5; }}
.modal-btn.primary {{
  background: var(--green-primary);
  color: white;
  border-color: var(--green-primary);
}}
.modal-btn.primary:hover {{ background: var(--green-mid); }}

.export-options {{
  display: flex;
  gap: 12px;
  margin-bottom: 20px;
}}
.export-option {{
  flex: 1;
  border: 1px solid var(--border);
  border-radius: 8px;
  padding: 16px;
  cursor: pointer;
  text-align: center;
  transition: border-color 0.2s, box-shadow 0.2s;
}}
.export-option:hover {{
  border-color: var(--green-primary);
  box-shadow: 0 2px 8px rgba(0,0,0,0.08);
}}
.export-option-icon {{ font-size: 28px; margin-bottom: 8px; }}
.export-option-label {{ font-weight: 600; font-size: 14px; margin-bottom: 4px; }}
.export-option-desc {{ font-size: 11px; color: var(--text-light); line-height: 1.4; }}

/* ── Conditional visibility ── */
.q-hidden {{
  display: none !important;
}}
.q-revealed {{
  animation: fadeSlideIn 0.3s ease;
}}
@keyframes fadeSlideIn {{
  from {{ opacity: 0; transform: translateY(-8px); }}
  to {{ opacity: 1; transform: translateY(0); }}
}}

/* ── Review mode ── */
body.review-mode .question-card {{
  display: none !important;
}}
body.review-mode .question-card.q-answered {{
  display: block !important;
}}
body.review-mode .question-card .response-field {{
  pointer-events: none;
  opacity: 0.7;
}}
body.review-mode .question-card textarea,
body.review-mode .question-card select {{
  background: #f9f9f9;
}}
body.review-mode .level-header,
body.review-mode .criteria-header {{
  display: none;
}}
body.review-mode .criteria-header.has-answers {{
  display: block;
}}
body.review-mode .level-header.has-answers {{
  display: block;
}}
.review-active {{
  background: rgba(255,255,255,0.35) !important;
  font-weight: 700 !important;
}}

/* ── N/A toggle ── */
.na-toggle {{
  background: none;
  border: 1px solid #ccc;
  color: #999;
  font-size: 9px;
  padding: 1px 5px;
  border-radius: 3px;
  cursor: pointer;
  flex-shrink: 0;
  margin-left: 4px;
  line-height: 1.3;
  transition: all 0.15s;
}}
.na-toggle:hover {{ border-color: #C62828; color: #C62828; }}
.sidebar-item.credit-na .na-toggle {{
  background: #C62828;
  color: white;
  border-color: #C62828;
}}
.sidebar-item.credit-na .sidebar-item-name {{
  text-decoration: line-through;
  opacity: 0.4;
}}
.sidebar-item.credit-na .sidebar-progress-ring {{ opacity: 0.2; }}
.credit-page.credit-na .credit-body {{ display: none; }}
.credit-page.credit-na .credit-progress-bar {{ display: none; }}
.credit-page.credit-na .credit-progress-text {{ display: none; }}
.credit-page.credit-na .wizard-nav {{ display: none !important; }}
.credit-page.credit-na::after {{
  content: 'This credit has been marked as Not Applicable';
  display: block;
  text-align: center;
  padding: 60px 20px;
  color: var(--text-light);
  font-size: 16px;
  font-style: italic;
}}
.na-btn-header {{
  background: rgba(255,255,255,0.15);
  color: white;
  border: 1px solid rgba(255,255,255,0.3);
  padding: 4px 10px;
  border-radius: 4px;
  cursor: pointer;
  font-size: 11px;
}}
.na-btn-header:hover {{ background: rgba(255,255,255,0.3); }}
.credit-page.credit-na .na-btn-header {{ background: rgba(255,255,255,0.3); }}
.credit-header-right {{ display: flex; gap: 8px; align-items: center; }}

/* ── Sidebar progress ring ── */
.sidebar-progress-ring {{
  flex-shrink: 0;
  margin-left: auto;
  display: flex;
  align-items: center;
}}
.sidebar-item-name {{ cursor: pointer; }}

/* ── Expandable guidance ── */
.guidance-wrapper {{ margin-top: 6px; }}
.guidance-toggle {{
  background: none;
  border: none;
  color: #F57F17;
  font-size: 12px;
  cursor: pointer;
  padding: 4px 0;
  display: flex;
  align-items: center;
  gap: 4px;
  font-weight: 600;
}}
.guidance-toggle:hover {{ color: #E65100; }}
.guidance-icon {{
  display: inline-flex;
  align-items: center;
  justify-content: center;
  width: 16px; height: 16px;
  border-radius: 50%;
  background: #FFF8E1;
  border: 1px solid #FFD54F;
  font-size: 10px;
  font-weight: 700;
  color: #F57F17;
}}
.guidance-arrow {{ font-size: 9px; transition: transform 0.2s; }}
.guidance-wrapper.open .guidance-arrow {{ transform: rotate(180deg); }}
.guidance-content {{
  max-height: 0;
  overflow: hidden;
  transition: max-height 0.25s ease, padding 0.25s ease;
  font-size: 12px;
  color: #666;
  line-height: 1.5;
  background: #FFFDE7;
  border: 1px solid #FFF9C4;
  border-radius: 4px;
  padding: 0 10px;
}}
.guidance-wrapper.open .guidance-content {{ max-height: 200px; padding: 8px 10px; }}

/* ── Wizard mode ── */
.wizard-toggle {{
  background: rgba(255,255,255,0.15);
  color: white;
  border: 1px solid rgba(255,255,255,0.3);
  padding: 4px 10px;
  border-radius: 4px;
  cursor: pointer;
  font-size: 11px;
}}
.wizard-toggle:hover {{ background: rgba(255,255,255,0.3); }}
.wizard-toggle.active {{ background: rgba(255,255,255,0.35); font-weight: 700; }}
.wizard-nav {{
  display: flex;
  align-items: center;
  justify-content: space-between;
  padding: 10px 32px;
  background: var(--white);
  border-bottom: 1px solid var(--border);
}}
.wizard-btn {{
  background: var(--green-primary);
  color: white;
  border: none;
  padding: 8px 20px;
  border-radius: 6px;
  cursor: pointer;
  font-size: 13px;
  transition: background 0.2s;
}}
.wizard-btn:hover {{ background: var(--green-mid); }}
.wizard-btn:disabled {{ opacity: 0.3; cursor: default; }}
.wizard-btn-next {{ background: var(--green-mid); }}
.wizard-step-text {{ font-size: 13px; color: var(--text-light); font-weight: 600; }}

body.wizard-active .credit-page:not(.wizard-target) .question-card {{ }}
.wizard-target .question-card {{ display: none !important; }}
.wizard-target .question-card.wizard-current {{ display: block !important; }}
.wizard-target .question-card.wizard-current.q-hidden {{ display: none !important; }}
.wizard-target .level-header,
.wizard-target .criteria-header {{ display: none; }}

/* ── Validation warnings ── */
.q-unanswered-warn {{
  border-color: #FFCC80 !important;
}}
.q-unanswered-warn .question-ref::after {{
  content: ' (unanswered)';
  color: #FF9800;
  font-weight: 400;
  font-family: inherit;
}}
.question-card {{ position: relative; }}

/* Gaps summary panel in review mode */
.gaps-panel {{
  display: none;
  background: #FFF3E0;
  border: 1px solid #FFCC80;
  border-radius: 8px;
  padding: 20px;
  margin-bottom: 20px;
}}
body.review-mode .gaps-panel {{ display: block; }}
.gaps-panel h3 {{ font-size: 15px; color: #E65100; margin-bottom: 10px; }}
.gaps-list {{ list-style: none; padding: 0; }}
.gaps-list li {{
  padding: 6px 0;
  font-size: 13px;
  color: var(--text);
  border-bottom: 1px solid #FFE0B2;
  cursor: pointer;
}}
.gaps-list li:hover {{ color: var(--green-primary); }}
.gaps-list li:last-child {{ border-bottom: none; }}
.gaps-list-ref {{
  font-family: monospace;
  font-weight: 600;
  color: #E65100;
  margin-right: 6px;
}}
.gaps-count {{ font-size: 13px; color: var(--text-light); margin-bottom: 10px; }}

/* ── Version history ── */
.history-overlay {{
  display: none;
  position: fixed;
  top: 0; left: 0; right: 0; bottom: 0;
  background: rgba(0,0,0,0.4);
  z-index: 200;
  justify-content: center;
  align-items: center;
}}
.history-overlay.active {{ display: flex; }}
.history-panel {{
  background: white;
  border-radius: 12px;
  width: 600px;
  max-width: 90vw;
  max-height: 80vh;
  display: flex;
  flex-direction: column;
  box-shadow: 0 20px 60px rgba(0,0,0,0.2);
}}
.history-header {{
  padding: 20px 24px;
  border-bottom: 1px solid var(--border);
  display: flex;
  justify-content: space-between;
  align-items: center;
}}
.history-header h3 {{ font-size: 16px; }}
.history-close {{ background: none; border: none; font-size: 20px; cursor: pointer; color: var(--text-light); }}
.history-body {{ flex: 1; overflow-y: auto; padding: 16px 24px; }}
.history-entry {{ padding: 12px 0; border-bottom: 1px solid var(--border); }}
.history-entry:last-child {{ border-bottom: none; }}
.history-time {{ font-size: 11px; color: var(--text-light); margin-bottom: 4px; }}
.history-changes {{ font-size: 13px; }}
.history-change {{ display: flex; gap: 6px; padding: 3px 0; }}
.history-change-ref {{ font-family: monospace; font-weight: 600; color: var(--green-primary); min-width: 50px; }}
.history-restore {{
  background: var(--green-primary);
  color: white;
  border: none;
  padding: 4px 10px;
  border-radius: 4px;
  cursor: pointer;
  font-size: 11px;
  margin-top: 6px;
}}
.history-restore:hover {{ background: var(--green-mid); }}
.history-empty {{ text-align: center; padding: 40px; color: var(--text-light); font-size: 14px; }}
.autosave-indicator {{ font-size: 11px; opacity: 0.6; transition: opacity 0.3s; }}
.autosave-indicator.saving {{ opacity: 1; }}

/* ── Dark mode toggle ── */
.dark-toggle {{
  background: none;
  border: 1px solid rgba(255,255,255,0.2);
  color: white;
  width: 30px;
  height: 30px;
  border-radius: 50%;
  cursor: pointer;
  font-size: 15px;
  display: flex;
  align-items: center;
  justify-content: center;
  transition: background 0.2s;
  padding: 0;
  flex-shrink: 0;
}}
.dark-toggle:hover {{ background: rgba(255,255,255,0.15); }}

/* ── Search ── */
.search-container {{
  padding: 10px 12px;
  border-bottom: 1px solid var(--border);
  position: relative;
}}
.search-input {{
  width: 100%;
  padding: 8px 10px 8px 30px;
  border: 1px solid var(--border);
  border-radius: 6px;
  font-size: 13px;
  background: var(--bg);
  color: var(--text);
  outline: none;
  transition: border-color 0.2s;
}}
.search-input:focus {{ border-color: var(--green-primary); }}
.search-input::placeholder {{ color: var(--text-light); }}
.search-icon {{
  position: absolute;
  left: 22px;
  top: 50%;
  transform: translateY(-50%);
  font-size: 13px;
  color: var(--text-light);
  pointer-events: none;
}}
.search-clear {{
  position: absolute;
  right: 20px;
  top: 50%;
  transform: translateY(-50%);
  background: none;
  border: none;
  font-size: 16px;
  color: var(--text-light);
  cursor: pointer;
  display: none;
  padding: 2px 4px;
}}
.search-clear.visible {{ display: block; }}
.search-results {{
  display: none;
  max-height: calc(100vh - var(--header-height) - 100px);
  overflow-y: auto;
  padding: 0;
}}
.search-results.active {{ display: block; }}
.search-result-item {{
  padding: 10px 16px;
  border-bottom: 1px solid var(--border);
  cursor: pointer;
  transition: background 0.1s;
}}
.search-result-item:hover {{ background: var(--hover-bg); }}
.search-result-ref {{
  font-family: monospace;
  font-size: 11px;
  font-weight: 700;
  color: var(--green-primary);
}}
.search-result-credit {{
  font-size: 10px;
  color: var(--text-light);
  margin-left: 8px;
}}
.search-result-text {{
  font-size: 12px;
  color: var(--text);
  margin-top: 3px;
  line-height: 1.4;
}}
.search-result-text mark {{
  background: #FFEE58;
  color: #000;
  border-radius: 2px;
  padding: 0 1px;
}}
html.dark .search-result-text mark {{
  background: #5a4a10;
  color: #eee;
}}
.search-result-count {{
  padding: 8px 16px;
  font-size: 12px;
  color: var(--text-light);
  border-bottom: 1px solid var(--border);
}}
.search-no-results {{
  padding: 20px 16px;
  text-align: center;
  color: var(--text-light);
  font-size: 13px;
}}

/* ── Responsive ── */
@media (max-width: 768px) {{
  .menu-toggle {{ display: block; }}
  .sidebar {{
    transform: translateX(-100%);
    width: 280px;
    box-shadow: 4px 0 20px rgba(0,0,0,0.15);
  }}
  .sidebar.open {{ transform: translateX(0); }}
  .main-content {{ margin-left: 0; }}
  .credit-header {{ padding: 20px; }}
  .credit-body {{ padding: 16px; }}
  .dashboard {{ padding: 20px; }}
  .dash-overview {{ flex-direction: column; }}
  .app-header .subtitle {{ display: none; }}
}}

/* ── Scrollbar ── */
::-webkit-scrollbar {{ width: 8px; }}
::-webkit-scrollbar-track {{ background: transparent; }}
::-webkit-scrollbar-thumb {{ background: #ccc; border-radius: 4px; }}
::-webkit-scrollbar-thumb:hover {{ background: #aaa; }}

/* Save notification */
.save-toast {{
  position: fixed;
  bottom: 24px;
  right: 24px;
  background: var(--green-primary);
  color: white;
  padding: 12px 20px;
  border-radius: 8px;
  font-size: 13px;
  box-shadow: 0 4px 12px rgba(0,0,0,0.15);
  transform: translateY(80px);
  opacity: 0;
  transition: all 0.3s;
  z-index: 300;
}}
.save-toast.show {{ transform: translateY(0); opacity: 1; }}
</style>
<script src="https://cdn.sheetjs.com/xlsx-0.20.3/package/dist/xlsx.full.min.js"></script>
</head>
<body>

<!-- Header -->
<header class="app-header">
  <button class="menu-toggle" onclick="toggleSidebar()">&#9776;</button>
  <h1>Green Star Buildings v1.1</h1>
  <span class="subtitle">Submission Forms</span>
  <div class="header-actions">
    <span class="autosave-indicator" id="autosave-indicator">Saved</span>
    <button class="header-btn" onclick="showDashboard()">Dashboard</button>
    <button class="header-btn" id="review-btn" onclick="toggleReview()">Review</button>
    <button class="header-btn" onclick="showHistory()">History</button>
    <button class="header-btn" onclick="saveAllResponses()">Save</button>
    <button class="header-btn primary" onclick="showExportModal()">Export</button>
    <button class="dark-toggle" id="dark-toggle" onclick="toggleDark()" title="Toggle dark mode">&#9790;</button>
  </div>
</header>

<!-- Sidebar -->
<nav class="sidebar" id="sidebar">
  <div class="search-container">
    <span class="search-icon">&#128269;</span>
    <input type="text" class="search-input" id="search-input" placeholder="Search questions..." oninput="onSearch(this.value)" onfocus="onSearchFocus()" />
    <button class="search-clear" id="search-clear" onclick="clearSearch()">&times;</button>
  </div>
  <div class="search-results" id="search-results"></div>
  <div id="sidebar-nav">
    <div class="sidebar-item active" data-credit="dashboard" onclick="showDashboard()" style="padding:14px 16px;font-weight:600;border-bottom:1px solid var(--border);">
      <span class="sidebar-item-name">&#9638; Dashboard</span>
    </div>
    {sidebar_html}
  </div>
</nav>

<!-- Main Content -->
<main class="main-content" id="main-content">

  <!-- Dashboard -->
  <div class="dashboard" id="dashboard">
    <div class="dash-title">Submission Dashboard</div>
    <div class="dash-subtitle">Green Star Buildings v1.1 — Track your progress across all {total_credits} credits</div>

    <div class="dash-overview">
      <div class="dash-stat">
        <div class="dash-stat-number">{total_credits}</div>
        <div class="dash-stat-label">Credits</div>
      </div>
      <div class="dash-stat">
        <div class="dash-stat-number">{total_questions}</div>
        <div class="dash-stat-label">Questions</div>
      </div>
      <div class="dash-stat">
        <div class="dash-stat-number" id="dash-answered">0</div>
        <div class="dash-stat-label">Answered</div>
      </div>
      <div class="dash-stat">
        <div class="dash-stat-number" id="dash-pct">0%</div>
        <div class="dash-stat-label">Complete</div>
      </div>
    </div>

    <div class="dash-grid">
      {dashboard_cards}
    </div>
  </div>

  <!-- Credit Pages -->
  {pages_html}
</main>

<!-- Export Modal -->
<div class="modal-overlay" id="export-modal">
  <div class="modal">
    <h3>Export Responses</h3>
    <p>Download your responses in your preferred format. Excel includes full formatting matching the original spreadsheet.</p>
    <div class="export-options">
      <div class="export-option" onclick="exportExcel()">
        <div class="export-option-icon" style="color:#217346">&#128196;</div>
        <div class="export-option-label">Excel (.xlsx)</div>
        <div class="export-option-desc">Formatted spreadsheet with colored headers, dropdowns, and all responses</div>
      </div>
      <div class="export-option" onclick="exportResponses()">
        <div class="export-option-icon" style="color:#F57F17">&#123; &#125;</div>
        <div class="export-option-label">JSON</div>
        <div class="export-option-desc">Raw data file for backup and re-import</div>
      </div>
      <div class="export-option" onclick="importResponses()">
        <div class="export-option-icon" style="color:#1565C0">&#128229;</div>
        <div class="export-option-label">Import JSON</div>
        <div class="export-option-desc">Restore a previous session from exported JSON</div>
      </div>
    </div>
    <div class="modal-actions">
      <button class="modal-btn" onclick="closeExportModal()">Cancel</button>
    </div>
  </div>
</div>

<!-- Save Toast -->
<div class="save-toast" id="save-toast">Responses saved to browser storage.</div>

<!-- Hidden file input for import -->
<input type="file" id="import-file" accept=".json" style="display:none" onchange="handleImport(event)">

<!-- Version history modal -->
<div class="history-overlay" id="history-overlay" onclick="if(event.target===this)closeHistory()">
  <div class="history-panel">
    <div class="history-header">
      <h3>Version History</h3>
      <button class="history-close" onclick="closeHistory()">&times;</button>
    </div>
    <div class="history-body" id="history-body">
      <div class="history-empty">No history yet. Changes are tracked as you work.</div>
    </div>
  </div>
</div>

<script>
// ── Conditional rules ──
const CONDITIONAL_RULES = {conditional_rules_json};

// ── State ──
let wizardCredits = {{}};  // creditId -> {{ active: bool, step: int }}
let naCredits = new Set();
let versionHistory = [];
let lastSnapshot = {{}};

// ── Navigation ──
function showCredit(id) {{
  document.querySelectorAll('.credit-page').forEach(p => p.style.display = 'none');
  document.getElementById('dashboard').style.display = 'none';
  const page = document.getElementById(id);
  if (page) {{
    page.style.display = 'block';
    document.getElementById('main-content').scrollTop = 0;
  }}
  document.querySelectorAll('.sidebar-item').forEach(i => i.classList.remove('active'));
  const item = document.querySelector(`#sidebar-${{id}}`);
  if (item) item.classList.add('active');
  document.getElementById('sidebar').classList.remove('open');
  if (wizardCredits[id] && wizardCredits[id].active) renderWizard(id);
}}

function showDashboard() {{
  document.querySelectorAll('.credit-page').forEach(p => p.style.display = 'none');
  document.getElementById('dashboard').style.display = 'block';
  document.getElementById('main-content').scrollTop = 0;
  document.querySelectorAll('.sidebar-item').forEach(i => i.classList.remove('active'));
  document.querySelector('.sidebar-item[data-credit="dashboard"]').classList.add('active');
  updateDashboard();
  document.getElementById('sidebar').classList.remove('open');
}}

function toggleCategory(el) {{ el.classList.toggle('collapsed'); }}
function toggleSidebar() {{ document.getElementById('sidebar').classList.toggle('open'); }}

// ── N/A toggle ──
function toggleNA(creditId, event) {{
  event.stopPropagation();
  const page = document.getElementById(creditId);
  const sidebar = document.getElementById('sidebar-' + creditId);
  if (!page) return;
  const isNA = page.classList.toggle('credit-na');
  if (sidebar) sidebar.classList.toggle('credit-na', isNA);
  if (isNA) {{ naCredits.add(creditId); }} else {{ naCredits.delete(creditId); }}
  saveNAState();
  updateDashboard();
  updateAllProgress();
}}

function saveNAState() {{
  localStorage.setItem('greenstar_na_credits', JSON.stringify([...naCredits]));
}}

function loadNAState() {{
  try {{
    const raw = localStorage.getItem('greenstar_na_credits');
    if (!raw) return;
    const arr = JSON.parse(raw);
    arr.forEach(id => {{
      naCredits.add(id);
      const page = document.getElementById(id);
      const sidebar = document.getElementById('sidebar-' + id);
      if (page) page.classList.add('credit-na');
      if (sidebar) sidebar.classList.add('credit-na');
    }});
  }} catch(e) {{}}
}}

// ── Conditional visibility engine ──
function applyConditionalRules() {{
  for (const [inputId, rule] of Object.entries(CONDITIONAL_RULES)) {{
    const card = document.getElementById('card-' + inputId);
    if (!card) continue;
    const gateway = document.getElementById(rule.depends_on);
    if (!gateway) continue;
    const gatewayVal = gateway.value;
    const shouldShow = gatewayVal === rule.show_when;
    if (shouldShow) {{
      if (card.classList.contains('q-hidden')) {{
        card.classList.remove('q-hidden');
        card.classList.add('q-revealed');
      }}
    }} else {{
      card.classList.add('q-hidden');
      card.classList.remove('q-revealed');
    }}
  }}
}}

// ── Unified answer handler ──
function onAnswer(creditId) {{
  applyConditionalRules();
  updateProgress(creditId);
  updateSidebarRing(creditId);
  // Update autosave indicator
  const indicator = document.getElementById('autosave-indicator');
  indicator.textContent = 'Unsaved';
  indicator.classList.add('saving');
  clearTimeout(window._saveTimer);
  window._saveTimer = setTimeout(() => {{
    saveAllResponses();
    indicator.textContent = 'Saved';
    indicator.classList.remove('saving');
  }}, 2000);
  // Update wizard if active
  if (wizardCredits[creditId] && wizardCredits[creditId].active) renderWizard(creditId);
}}

// ── Progress tracking (only counts visible, non-N/A questions) ──
function updateProgress(creditId) {{
  const page = document.getElementById(creditId);
  if (!page || page.classList.contains('credit-na')) return;
  const cards = page.querySelectorAll('.question-card');
  let visible = 0, answered = 0;
  cards.forEach(card => {{
    if (card.classList.contains('q-hidden')) return;
    visible++;
    const input = card.querySelector('select, textarea');
    if (!input) return;
    const val = input.tagName === 'SELECT' ? input.value : input.value.trim();
    if (val) {{
      answered++;
      card.classList.add('q-answered');
      card.classList.remove('q-unanswered-warn');
    }} else {{
      card.classList.remove('q-answered');
    }}
  }});
  const pct = visible > 0 ? (answered / visible) * 100 : 0;
  const bar = document.getElementById(`${{creditId}}-progress`);
  const text = document.getElementById(`${{creditId}}-progress-text`);
  if (bar) bar.style.width = pct + '%';
  if (text) text.textContent = `${{answered}} of ${{visible}} answered`;
}}

function updateSidebarRing(creditId) {{
  const page = document.getElementById(creditId);
  if (!page) return;
  const cards = page.querySelectorAll('.question-card');
  let visible = 0, answered = 0;
  cards.forEach(card => {{
    if (card.classList.contains('q-hidden')) return;
    visible++;
    const input = card.querySelector('select, textarea');
    if (!input) return;
    const val = input.tagName === 'SELECT' ? input.value : input.value.trim();
    if (val) answered++;
  }});
  const pct = visible > 0 ? answered / visible : 0;
  const ring = document.querySelector(`#ring-${{creditId}} .ring-fill`);
  if (ring) {{
    const circumference = 44;
    ring.style.strokeDashoffset = circumference * (1 - pct);
  }}
}}

function updateAllProgress() {{
  document.querySelectorAll('.credit-page').forEach(page => {{
    updateProgress(page.id);
    updateSidebarRing(page.id);
  }});
}}

function updateDashboard() {{
  let totalVisible = 0, totalAnswered = 0;
  document.querySelectorAll('.credit-page').forEach(page => {{
    if (page.classList.contains('credit-na')) return;
    page.querySelectorAll('.question-card').forEach(card => {{
      if (card.classList.contains('q-hidden')) return;
      totalVisible++;
      const input = card.querySelector('select, textarea');
      if (!input) return;
      const val = input.tagName === 'SELECT' ? input.value : input.value.trim();
      if (val) totalAnswered++;
    }});
  }});

  document.getElementById('dash-answered').textContent = totalAnswered;
  document.getElementById('dash-pct').textContent = (totalVisible > 0 ? Math.round((totalAnswered / totalVisible) * 100) : 0) + '%';

  const catMap = {json.dumps({cat: [c["sheet_name"] for c in all_credits if c["category"] == cat] for cat in CATEGORIES})};
  const creditMap = {json.dumps({c["sheet_name"]: {"id": f"credit-{i}", "count": len(c["questions"])} for i, c in enumerate(all_credits)})};

  for (const [cat, sheets] of Object.entries(catMap)) {{
    let catVisible = 0, catAnswered = 0;
    sheets.forEach(s => {{
      const info = creditMap[s];
      if (!info) return;
      const page = document.getElementById(info.id);
      if (!page || page.classList.contains('credit-na')) return;
      page.querySelectorAll('.question-card').forEach(card => {{
        if (card.classList.contains('q-hidden')) return;
        catVisible++;
        const input = card.querySelector('select, textarea');
        if (!input) return;
        const val = input.tagName === 'SELECT' ? input.value : input.value.trim();
        if (val) catAnswered++;
      }});
    }});
    const pct = catVisible > 0 ? Math.round((catAnswered / catVisible) * 100) : 0;
    const bar = document.getElementById(`dash-${{cat.toLowerCase()}}-bar`);
    const pctEl = document.getElementById(`dash-${{cat.toLowerCase()}}-pct`);
    if (bar) bar.style.width = pct + '%';
    if (pctEl) pctEl.textContent = pct + '% complete';
  }}
}}

// ── Wizard mode ──
function toggleWizard(creditId) {{
  if (!wizardCredits[creditId]) wizardCredits[creditId] = {{ active: false, step: 0 }};
  wizardCredits[creditId].active = !wizardCredits[creditId].active;
  wizardCredits[creditId].step = 0;
  const page = document.getElementById(creditId);
  const nav = document.getElementById(creditId + '-wizard-nav');
  const btn = page.querySelector('.wizard-toggle');
  if (wizardCredits[creditId].active) {{
    page.classList.add('wizard-target');
    nav.style.display = 'flex';
    btn.classList.add('active');
    btn.textContent = 'Show all';
    renderWizard(creditId);
  }} else {{
    page.classList.remove('wizard-target');
    nav.style.display = 'none';
    btn.classList.remove('active');
    btn.textContent = 'Step-by-step';
    page.querySelectorAll('.question-card').forEach(c => c.classList.remove('wizard-current'));
  }}
}}

function getVisibleCards(creditId) {{
  const page = document.getElementById(creditId);
  if (!page) return [];
  return Array.from(page.querySelectorAll('.question-card')).filter(
    c => !c.classList.contains('q-hidden')
  );
}}

function renderWizard(creditId) {{
  const wiz = wizardCredits[creditId];
  if (!wiz || !wiz.active) return;
  const cards = getVisibleCards(creditId);
  if (cards.length === 0) return;
  if (wiz.step >= cards.length) wiz.step = cards.length - 1;
  if (wiz.step < 0) wiz.step = 0;
  const page = document.getElementById(creditId);
  page.querySelectorAll('.question-card').forEach(c => c.classList.remove('wizard-current'));
  cards[wiz.step].classList.add('wizard-current');
  const stepText = document.getElementById(creditId + '-wizard-step');
  if (stepText) stepText.textContent = `${{wiz.step + 1}} / ${{cards.length}}`;
  // Scroll card into view
  document.getElementById('main-content').scrollTop = 0;
}}

function wizardNext(creditId) {{
  const wiz = wizardCredits[creditId];
  if (!wiz) return;
  const cards = getVisibleCards(creditId);
  if (wiz.step < cards.length - 1) {{ wiz.step++; renderWizard(creditId); }}
}}

function wizardPrev(creditId) {{
  const wiz = wizardCredits[creditId];
  if (!wiz) return;
  if (wiz.step > 0) {{ wiz.step--; renderWizard(creditId); }}
}}

// ── Review mode with validation ──
let reviewMode = false;
function toggleReview() {{
  reviewMode = !reviewMode;
  document.body.classList.toggle('review-mode', reviewMode);
  document.getElementById('review-btn').classList.toggle('review-active', reviewMode);

  if (reviewMode) {{
    document.querySelectorAll('.credit-page').forEach(page => {{
      if (page.classList.contains('credit-na')) return;
      const body = page.querySelector('.credit-body');
      if (!body) return;

      // Build gaps list
      const gapsList = document.getElementById(page.id + '-gaps-list');
      const gapsCount = document.getElementById(page.id + '-gaps-count');
      if (!gapsList) return;
      gapsList.innerHTML = '';
      let gaps = 0;

      // Mark unanswered visible questions
      page.querySelectorAll('.question-card').forEach(card => {{
        if (card.classList.contains('q-hidden')) return;
        const input = card.querySelector('select, textarea');
        if (!input) return;
        const val = input.tagName === 'SELECT' ? input.value : input.value.trim();
        if (!val) {{
          card.classList.add('q-unanswered-warn');
          gaps++;
          const ref = card.querySelector('.question-ref');
          const qtext = card.querySelector('.question-text');
          const li = document.createElement('li');
          li.innerHTML = `<span class="gaps-list-ref">${{ref ? ref.textContent : ''}}</span>${{qtext ? qtext.textContent.substring(0, 80) + '...' : ''}}`;
          li.onclick = function() {{
            toggleReview();
            card.scrollIntoView({{ behavior: 'smooth', block: 'center' }});
          }};
          gapsList.appendChild(li);
        }}
      }});

      if (gapsCount) gapsCount.textContent = gaps > 0 ? `${{gaps}} question${{gaps > 1 ? 's' : ''}} still need${{gaps > 1 ? '' : 's'}} a response` : 'All questions answered!';

      // Mark headers with answers
      let currentHeaders = [];
      Array.from(body.children).forEach(el => {{
        if (el.classList.contains('level-header') || el.classList.contains('criteria-header')) {{
          el.classList.remove('has-answers');
          currentHeaders.push(el);
        }} else if (el.classList.contains('question-card') && (el.classList.contains('q-answered') || el.classList.contains('q-unanswered-warn'))) {{
          currentHeaders.forEach(h => h.classList.add('has-answers'));
        }}
      }});
    }});
  }} else {{
    // Clear warning highlights
    document.querySelectorAll('.q-unanswered-warn').forEach(c => c.classList.remove('q-unanswered-warn'));
  }}
}}

// ── Save / Load ──
function saveAllResponses() {{
  const data = {{}};
  document.querySelectorAll('[data-credit]').forEach(el => {{
    if (el.id) data[el.id] = el.tagName === 'SELECT' ? el.value : el.value;
  }});
  try {{
    // Track version history (diff against last snapshot)
    recordHistory(data);
    localStorage.setItem('greenstar_responses', JSON.stringify(data));
    const indicator = document.getElementById('autosave-indicator');
    indicator.textContent = 'Saved';
    indicator.classList.remove('saving');
  }} catch(e) {{
    console.error('Save failed', e);
  }}
}}

function loadResponses() {{
  try {{
    const raw = localStorage.getItem('greenstar_responses');
    if (!raw) return;
    const data = JSON.parse(raw);
    for (const [id, val] of Object.entries(data)) {{
      const el = document.getElementById(id);
      if (el) el.value = val;
    }}
    lastSnapshot = {{ ...data }};
    applyConditionalRules();
    updateAllProgress();
    updateDashboard();
  }} catch(e) {{
    console.error('Load failed', e);
  }}
  // Load version history
  try {{
    const h = localStorage.getItem('greenstar_history');
    if (h) versionHistory = JSON.parse(h);
  }} catch(e) {{}}
}}

function showToast() {{
  const t = document.getElementById('save-toast');
  t.classList.add('show');
  setTimeout(() => t.classList.remove('show'), 2000);
}}

// ── Version history ──
function recordHistory(newData) {{
  const changes = [];
  const allKeys = new Set([...Object.keys(lastSnapshot), ...Object.keys(newData)]);
  allKeys.forEach(key => {{
    const oldVal = lastSnapshot[key] || '';
    const newVal = newData[key] || '';
    if (oldVal !== newVal) {{
      // Extract ref from id like "credit-0-ID-1"
      const parts = key.split('-');
      const ref = parts.length >= 3 ? parts.slice(2).join('.') : key;
      changes.push({{ id: key, ref: ref, from: oldVal, to: newVal }});
    }}
  }});
  if (changes.length > 0) {{
    versionHistory.push({{
      time: new Date().toISOString(),
      changes: changes,
      snapshot: {{ ...newData }}
    }});
    // Keep last 50 entries
    if (versionHistory.length > 50) versionHistory = versionHistory.slice(-50);
    try {{
      localStorage.setItem('greenstar_history', JSON.stringify(versionHistory));
    }} catch(e) {{}}
  }}
  lastSnapshot = {{ ...newData }};
}}

function showHistory() {{
  document.getElementById('history-overlay').classList.add('active');
  renderHistory();
}}

function closeHistory() {{
  document.getElementById('history-overlay').classList.remove('active');
}}

function renderHistory() {{
  const body = document.getElementById('history-body');
  if (versionHistory.length === 0) {{
    body.innerHTML = '<div class="history-empty">No history yet. Changes are tracked as you work.</div>';
    return;
  }}
  let html = '';
  // Show newest first
  for (let i = versionHistory.length - 1; i >= 0; i--) {{
    const entry = versionHistory[i];
    const d = new Date(entry.time);
    const timeStr = d.toLocaleDateString() + ' ' + d.toLocaleTimeString();
    html += `<div class="history-entry">`;
    html += `<div class="history-time">${{timeStr}} — ${{entry.changes.length}} change${{entry.changes.length > 1 ? 's' : ''}}</div>`;
    html += `<div class="history-changes">`;
    entry.changes.slice(0, 5).forEach(ch => {{
      const display = ch.to ? `"${{ch.to.substring(0, 60)}}${{ch.to.length > 60 ? '...' : ''}}"` : '<em>cleared</em>';
      html += `<div class="history-change"><span class="history-change-ref">${{ch.ref}}</span><span class="history-change-val">${{display}}</span></div>`;
    }});
    if (entry.changes.length > 5) html += `<div style="font-size:11px;color:#999">+ ${{entry.changes.length - 5}} more</div>`;
    html += `</div>`;
    html += `<button class="history-restore" onclick="restoreVersion(${{i}})">Restore this version</button>`;
    html += `</div>`;
  }}
  body.innerHTML = html;
}}

function restoreVersion(idx) {{
  if (idx < 0 || idx >= versionHistory.length) return;
  if (!confirm('Restore all responses to this point? Current answers will be saved in history first.')) return;
  // Save current state before restoring
  saveAllResponses();
  const snapshot = versionHistory[idx].snapshot;
  // Clear all inputs first
  document.querySelectorAll('[data-credit]').forEach(el => {{
    if (el.tagName === 'SELECT') el.value = '';
    else el.value = '';
  }});
  // Apply snapshot
  for (const [id, val] of Object.entries(snapshot)) {{
    const el = document.getElementById(id);
    if (el) el.value = val;
  }}
  localStorage.setItem('greenstar_responses', JSON.stringify(snapshot));
  lastSnapshot = {{ ...snapshot }};
  applyConditionalRules();
  updateAllProgress();
  updateDashboard();
  closeHistory();
  showToast();
}}

// ── Export / Import ──
const CREDITS_DATA = {credits_json_str};

function showExportModal() {{
  document.getElementById('export-modal').classList.add('active');
}}
function closeExportModal() {{
  document.getElementById('export-modal').classList.remove('active');
}}

function exportResponses() {{
  const data = {{}};
  document.querySelectorAll('[data-credit]').forEach(el => {{
    if (el.id) data[el.id] = el.tagName === 'SELECT' ? el.value : el.value;
  }});
  const blob = new Blob([JSON.stringify(data, null, 2)], {{type: 'application/json'}});
  const a = document.createElement('a');
  a.href = URL.createObjectURL(blob);
  a.download = 'greenstar_v1.1_responses.json';
  a.click();
  closeExportModal();
}}

function exportExcel() {{
  if (typeof XLSX === 'undefined') {{
    alert('Excel library is still loading. Please try again in a moment.');
    return;
  }}
  const wb = XLSX.utils.book_new();

  // Style helpers - hex color without #
  function hexFill(hex) {{ return {{ fgColor: {{ rgb: hex.replace('#','') }} }}; }}
  function makeStyle(opts) {{
    const s = {{}};
    if (opts.fill) s.fill = {{ patternType: 'solid', ...hexFill(opts.fill) }};
    if (opts.font) {{
      s.font = {{ name: 'Calibri', ...opts.font }};
      if (opts.font.color) s.font.color = {{ rgb: opts.font.color.replace('#','') }};
    }}
    if (opts.alignment) s.alignment = opts.alignment;
    s.border = {{
      top: {{ style: 'thin', color: {{ rgb: 'CCCCCC' }} }},
      bottom: {{ style: 'thin', color: {{ rgb: 'CCCCCC' }} }},
      left: {{ style: 'thin', color: {{ rgb: 'CCCCCC' }} }},
      right: {{ style: 'thin', color: {{ rgb: 'CCCCCC' }} }},
    }};
    return s;
  }}

  const headerStyle = makeStyle({{ fill:'#0D3318', font:{{ bold:true, sz:14, color:'#FFFFFF' }}, alignment:{{ wrapText:true, vertical:'center', horizontal:'center' }} }});
  const creditStyle = makeStyle({{ fill:'#1F4E28', font:{{ bold:true, sz:12, color:'#FFFFFF' }}, alignment:{{ wrapText:true }} }});
  const levelStyle = makeStyle({{ fill:'#2E7D32', font:{{ bold:true, sz:11, color:'#FFFFFF' }}, alignment:{{ wrapText:true }} }});
  const criteriaStyle = makeStyle({{ fill:'#C8E6C9', font:{{ bold:true, sz:11, color:'#1F4E28' }}, alignment:{{ wrapText:true }} }});
  const questionStyle = makeStyle({{ fill:'#F1F8E9', font:{{ sz:10 }}, alignment:{{ wrapText:true, vertical:'top' }} }});
  const conditionStyle = makeStyle({{ fill:'#EDE7F6', font:{{ bold:true, sz:10, color:'#7030A0' }}, alignment:{{ wrapText:true, vertical:'top' }} }});
  const dataStyle = makeStyle({{ fill:'#E3F2FD', font:{{ sz:10, italic:true, color:'#2E75B6' }}, alignment:{{ wrapText:true, vertical:'top' }} }});
  const responseStyle = makeStyle({{ fill:'#FFFFFF', font:{{ sz:10 }}, alignment:{{ wrapText:true, vertical:'top' }} }});

  const headers = ['Ref','Credit','Performance Level','Criteria','Question Type','Question','Response','Data Collection / Research Notes'];
  const colWidths = [8, 20, 22, 28, 16, 55, 50, 40];

  CREDITS_DATA.forEach(credit => {{
    const rows = [];
    const styles = [];
    const merges = [];

    // Header row
    rows.push(headers);
    styles.push(headers.map(() => headerStyle));

    // Credit title row
    const creditRow = [credit.title, '', '', '', '', '', '', ''];
    rows.push(creditRow);
    styles.push(creditRow.map(() => creditStyle));
    merges.push({{ s: {{ r: rows.length - 1, c: 0 }}, e: {{ r: rows.length - 1, c: 7 }} }});

    credit.sections.forEach(section => {{
      // Level header
      const lvlRow = [section.title, '', '', '', '', '', '', ''];
      rows.push(lvlRow);
      styles.push(lvlRow.map(() => levelStyle));
      merges.push({{ s: {{ r: rows.length - 1, c: 0 }}, e: {{ r: rows.length - 1, c: 7 }} }});

      section.criteria.forEach(crit => {{
        // Criteria header
        const critRow = [crit.name, '', '', '', '', '', '', ''];
        rows.push(critRow);
        styles.push(critRow.map(() => criteriaStyle));
        merges.push({{ s: {{ r: rows.length - 1, c: 0 }}, e: {{ r: rows.length - 1, c: 7 }} }});

        crit.questions.forEach(q => {{
          // Get response value from the form
          const el = document.getElementById(q.input_id);
          const response = el ? (el.tagName === 'SELECT' ? el.value : el.value) : '';

          const isYN = q.type === 'Condition (Y/N)';
          const isData = q.type === 'Data';
          const qRow = [q.ref, q.credit, q.level, q.criteria, q.type, q.question, response, q.data_note];
          rows.push(qRow);

          const rowStyles = qRow.map((_, ci) => {{
            if (ci === 6) return responseStyle;
            if (ci === 7 && q.data_note) return dataStyle;
            if (isYN) return conditionStyle;
            return questionStyle;
          }});
          styles.push(rowStyles);
        }});
      }});
    }});

    // Create worksheet from array
    const ws = XLSX.utils.aoa_to_sheet(rows);

    // Apply column widths
    ws['!cols'] = colWidths.map(w => ({{ wch: w }}));

    // Apply merges
    ws['!merges'] = merges;

    // Apply styles (works with xlsx-style or pro; basic xlsx ignores but structure is there)
    for (let r = 0; r < rows.length; r++) {{
      for (let c = 0; c < rows[r].length; c++) {{
        const addr = XLSX.utils.encode_cell({{ r, c }});
        if (ws[addr]) {{
          ws[addr].s = styles[r][c];
        }}
      }}
    }}

    // Apply row heights
    ws['!rows'] = rows.map((_, r) => {{
      if (r === 0) return {{ hpt: 35 }};
      if (r === 1) return {{ hpt: 30 }};
      return {{ hpt: 45 }};
    }});

    // Truncate sheet name to 31 chars (Excel limit)
    const sheetName = credit.sheet_name.substring(0, 31);
    XLSX.utils.book_append_sheet(wb, ws, sheetName);
  }});

  XLSX.writeFile(wb, 'Green_Star_Buildings_v1.1_Submission_Responses.xlsx');
  closeExportModal();
}}

function importResponses() {{
  document.getElementById('import-file').click();
}}

function handleImport(event) {{
  const file = event.target.files[0];
  if (!file) return;
  const reader = new FileReader();
  reader.onload = function(e) {{
    try {{
      const data = JSON.parse(e.target.result);
      for (const [id, val] of Object.entries(data)) {{
        const el = document.getElementById(id);
        if (el) el.value = val;
      }}
      localStorage.setItem('greenstar_responses', JSON.stringify(data));
      loadResponses();
      closeExportModal();
      showToast();
    }} catch(err) {{
      alert('Invalid file format.');
    }}
  }};
  reader.readAsText(file);
  event.target.value = '';
}}

// ── Dark mode ──
function toggleDark() {{
  document.documentElement.classList.toggle('dark');
  const isDark = document.documentElement.classList.contains('dark');
  localStorage.setItem('greenstar_dark', isDark ? '1' : '0');
  document.getElementById('dark-toggle').textContent = isDark ? '\\u2600' : '\\u263E';
}}

function loadDarkMode() {{
  const pref = localStorage.getItem('greenstar_dark');
  // Also check system preference if no saved pref
  const prefersDark = pref === '1' || (!pref && window.matchMedia('(prefers-color-scheme: dark)').matches);
  if (prefersDark) {{
    document.documentElement.classList.add('dark');
    document.getElementById('dark-toggle').textContent = '\\u2600';
  }}
}}

// ── Search ──
const SEARCH_INDEX = {search_index_json};
let searchTimeout = null;

function onSearch(query) {{
  const clearBtn = document.getElementById('search-clear');
  clearBtn.classList.toggle('visible', query.length > 0);

  clearTimeout(searchTimeout);
  if (!query.trim()) {{
    hideSearchResults();
    return;
  }}
  searchTimeout = setTimeout(() => performSearch(query.trim()), 150);
}}

function onSearchFocus() {{
  const q = document.getElementById('search-input').value.trim();
  if (q) performSearch(q);
}}

function performSearch(query) {{
  const terms = query.toLowerCase().split(/\\s+/).filter(t => t.length > 1);
  if (terms.length === 0) {{ hideSearchResults(); return; }}

  const results = [];
  SEARCH_INDEX.forEach(item => {{
    const haystack = (item.ref + ' ' + item.credit + ' ' + item.question + ' ' + item.note).toLowerCase();
    const matches = terms.every(t => haystack.includes(t));
    if (matches) results.push(item);
  }});

  renderSearchResults(results, terms, query);
}}

function renderSearchResults(results, terms, query) {{
  const container = document.getElementById('search-results');
  const navEl = document.getElementById('sidebar-nav');

  if (results.length === 0) {{
    container.innerHTML = '<div class="search-no-results">No questions match your search.</div>';
    container.classList.add('active');
    navEl.style.display = 'none';
    return;
  }}

  // Limit displayed results
  const shown = results.slice(0, 30);
  let html = `<div class="search-result-count">${{results.length}} result${{results.length > 1 ? 's' : ''}}${{results.length > 30 ? ' (showing first 30)' : ''}}</div>`;

  shown.forEach(item => {{
    // Highlight matched terms in question text
    let text = escapeHtml(item.question);
    terms.forEach(t => {{
      const re = new RegExp('(' + t.replace(/[.*+?^${{}}()|[\\]\\\\]/g, '\\\\$&') + ')', 'gi');
      text = text.replace(re, '<mark>$1</mark>');
    }});

    html += `<div class="search-result-item" onclick="goToQuestion('${{item.creditId}}', '${{item.cardId}}')">
      <div><span class="search-result-ref">${{escapeHtml(item.ref)}}</span><span class="search-result-credit">${{escapeHtml(item.credit)}}</span></div>
      <div class="search-result-text">${{text}}</div>
    </div>`;
  }});

  container.innerHTML = html;
  container.classList.add('active');
  navEl.style.display = 'none';
}}

function escapeHtml(str) {{
  const d = document.createElement('div');
  d.textContent = str;
  return d.innerHTML;
}}

function goToQuestion(creditId, cardId) {{
  clearSearch();
  showCredit(creditId);
  // Small delay to let the page render
  setTimeout(() => {{
    const card = document.getElementById(cardId);
    if (card) {{
      card.scrollIntoView({{ behavior: 'smooth', block: 'center' }});
      card.style.outline = '2px solid var(--green-mid)';
      card.style.outlineOffset = '2px';
      setTimeout(() => {{ card.style.outline = ''; card.style.outlineOffset = ''; }}, 2000);
    }}
  }}, 100);
}}

function clearSearch() {{
  document.getElementById('search-input').value = '';
  document.getElementById('search-clear').classList.remove('visible');
  hideSearchResults();
}}

function hideSearchResults() {{
  document.getElementById('search-results').classList.remove('active');
  document.getElementById('search-results').innerHTML = '';
  document.getElementById('sidebar-nav').style.display = '';
}}

// ── Keyboard shortcut ──
document.addEventListener('keydown', function(e) {{
  if ((e.ctrlKey || e.metaKey) && e.key === 's') {{
    e.preventDefault();
    saveAllResponses();
  }}
  // Ctrl/Cmd+K to focus search
  if ((e.ctrlKey || e.metaKey) && e.key === 'k') {{
    e.preventDefault();
    document.getElementById('search-input').focus();
  }}
}});

// ── Init ──
window.addEventListener('DOMContentLoaded', function() {{
  loadDarkMode();
  loadNAState();
  loadResponses();
  applyConditionalRules();
  updateAllProgress();
  updateDashboard();
}});
</script>
</body>
</html>'''

with open("index.html", "w") as f:
    f.write(html)

print(f"Generated index.html")
print(f"  Credits: {total_credits}")
print(f"  Questions: {total_questions}")
print(f"  Categories: {len(CATEGORIES)}")
print(f"  File size: {len(html):,} bytes")
